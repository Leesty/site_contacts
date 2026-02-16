import asyncio
import csv
import io
import os
from datetime import datetime, timedelta, timezone
from typing import List, Set, Dict, Optional
from zoneinfo import ZoneInfo

from aiogram import Bot, Dispatcher, F
from aiogram.filters import CommandStart, Command, StateFilter
from aiogram.types import (
    Message,
    CallbackQuery,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    KeyboardButton,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    BufferedInputFile,
)
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook


# ============ НАСТРОЙКИ ============

# Секретная команда для получения админ-доступа
ADMIN_SECRET_COMMAND = "get_bd_access_9876"

# Файл со списком админов (Telegram ID)
ADMINS_FILE = "admins.txt"

# Группа поддержки с топиками (Forum)
SUPPORT_GROUP_ID = -1003702935049

# Файл для хранения связей user_id -> topic_id
SUPPORT_TOPICS_FILE = "support_topics.csv"

# Типы баз данных и их настройки
# key: внутреннее имя, name: отображаемое имя, csv: файл, limit: лимит на пользователя
BASE_TYPES = {
    "telegram": {"name": "Telegram", "csv": "base_telegram.csv", "limit": 50},
    "whatsapp": {"name": "WhatsApp", "csv": "base_whatsapp.csv", "limit": 35},
    "max": {"name": "Max", "csv": "base_max.csv", "limit": 35},
    "viber": {"name": "Viber", "csv": "base_viber.csv", "limit": 35},
    "instagram": {"name": "Нельзяграм (там где Reels)", "csv": "base_instagram.csv", "limit": 300},
    "vk": {"name": "ВКонтакте", "csv": "base_vk.csv", "limit": 250},
    "ok": {"name": "Одноклассники", "csv": "base_ok.csv", "limit": 250},
    "email": {"name": "Почта", "csv": "base_email.csv", "limit": 100},
}

# Файл для хранения пользователей (счётчик)
USERS_FILE = "users.txt"

# Файл для хранения дополнительных лимитов (user_id, base_type, extra_limit)
USER_LIMITS_FILE = "user_limits.csv"

# Файл для хранения статусов пользователей (pending/approved/banned)
USER_STATUS_FILE = "user_status.csv"

# ID топика для заявок (создаётся автоматически или указать вручную)
REQUESTS_TOPIC_ID = None  # Будет создан автоматически

# ID топика «Отчёт» в чате админов
REPORTS_TOPIC_ID = 156

# ID топика «Лиды авто» в чате админов
LEADS_TOPIC_ID = 769

# Часовой пояс и граница «дня» для лидов (после 20:00 — новый день)
LEAD_TIMEZONE = "Europe/Moscow"
LEAD_DAY_CUTOFF_HOUR = 20

# Пауза между сообщениями в группу (защита от Flood control)
FLOOD_DELAY = 0.4

# Лимит лидов в одном отчёте (защита от спама и Flood control)
REPORT_LEADS_LIMIT = 5

# Карта названий листов Excel -> внутренние ключи (для загрузки через админку)
EXCEL_SHEET_MAP = {
    # Короткие названия
    "Тг": "telegram",
    "ТГ": "telegram",
    "Вотсап": "whatsapp",
    "Макс": "max",
    "Вайбер": "viber",
    "Инст": "instagram",
    "ВК": "vk",
    "Ок": "ok",
    "Почта": "email",
    # Полные названия
    "Telegram": "telegram",
    "telegram": "telegram",
    "WhatsApp": "whatsapp",
    "Whatsapp": "whatsapp",
    "whatsapp": "whatsapp",
    "Max": "max",
    "max": "max",
    "Viber": "viber",
    "viber": "viber",
    "Нельзяграм": "instagram",
    "Нельзяграм (там где Reels)": "instagram",
    "Instagram": "instagram",
    "instagram": "instagram",
    "ВКонтакте": "vk",
    "Вконтакте": "vk",
    "вконтакте": "vk",
    "VK": "vk",
    "Одноклассники": "ok",
    "одноклассники": "ok",
    "OK": "ok",
    "Ok": "ok",
    "Email": "email",
    "email": "email",
    "Почты": "email",
}

# Типы лидов (структура для хранения)
LEAD_TYPES = {
    "telegram": {"name": "Telegram", "csv": "leads_telegram.csv"},
    "whatsapp": {"name": "WhatsApp", "csv": "leads_whatsapp.csv"},
    "max": {"name": "Max", "csv": "leads_max.csv"},
    "viber": {"name": "Viber", "csv": "leads_viber.csv"},
    "instagram": {"name": "Нельзяграм", "csv": "leads_instagram.csv"},
    "vk": {"name": "ВКонтакте", "csv": "leads_vk.csv"},
    "ok": {"name": "Одноклассники", "csv": "leads_ok.csv"},
    "email": {"name": "Почта", "csv": "leads_email.csv"},
    "avito": {"name": "Авито", "csv": "leads_avito.csv"},
    "yula": {"name": "Юла", "csv": "leads_yula.csv"},
    "kwork": {"name": "Кворк", "csv": "leads_kwork.csv"},
    "other_social": {"name": "Прочие соц. сети", "csv": "leads_other_social.csv"},
    "self": {"name": "Самостоятельные лиды", "csv": "leads_self.csv"},
}

# ============ НАЧАЛЬНАЯ ЗАГРУЗКА (ОТКЛЮЧЕНА) ============
# Раскомментируй для автозагрузки из файла при первом запуске:
# INITIAL_EXCEL_PATH = "Новая таблица.xlsx"
# INITIAL_LOAD_ENABLED = True


# ============ СОСТОЯНИЯ FSM ============

class AdminStates(StatesGroup):
    waiting_upload_choice = State()  # Ожидание выбора типа базы для загрузки
    waiting_file = State()  # Ожидание файла от админа
    waiting_delete_confirm = State()  # Ожидание подтверждения удаления базы


class ReportStates(StatesGroup):
    waiting_report = State()  # Сбор файлов отчёта
    waiting_category = State()  # Выбор категории для лида из отчёта


class ManualLeadStates(StatesGroup):
    waiting_contact = State()  # Ожидание контакта лида
    waiting_category = State()  # Ожидание выбора категории


class DeleteLeadStates(StatesGroup):
    waiting_contact = State()  # Ожидание контакта для удаления


class SupportStates(StatesGroup):
    active = State()  # Пользователь нажал «Написать в поддержку» и может отправлять сообщения


# ============ ГЛОБАЛЬНЫЕ ПЕРЕМЕННЫЕ ============

csv_lock = asyncio.Lock()
processing_users: Dict[str, Set[int]] = {key: set() for key in BASE_TYPES}


# ============ РАБОТА С АДМИНАМИ ============

def load_admins() -> Set[int]:
    """Загружает список админов из файла."""
    if not os.path.exists(ADMINS_FILE):
        return set()
    with open(ADMINS_FILE, "r", encoding="utf-8") as f:
        admins = set()
        for line in f:
            line = line.strip()
            if line:
                try:
                    admins.add(int(line))
                except ValueError:
                    pass
        return admins


def save_admin(user_id: int) -> None:
    """Добавляет админа в файл."""
    admins = load_admins()
    if user_id not in admins:
        with open(ADMINS_FILE, "a", encoding="utf-8") as f:
            f.write(f"{user_id}\n")


def is_admin(user_id: int) -> bool:
    """Проверяет, является ли пользователь админом."""
    return user_id in load_admins()


# ============ СЧЁТЧИК ПОЛЬЗОВАТЕЛЕЙ ============

def load_users() -> Set[int]:
    """Загружает список пользователей из файла."""
    if not os.path.exists(USERS_FILE):
        return set()
    with open(USERS_FILE, "r", encoding="utf-8") as f:
        users = set()
        for line in f:
            line = line.strip()
            if line:
                try:
                    users.add(int(line))
                except ValueError:
                    pass
        return users


def save_user(user_id: int) -> None:
    """Добавляет пользователя в файл (если ещё нет)."""
    users = load_users()
    if user_id not in users:
        with open(USERS_FILE, "a", encoding="utf-8") as f:
            f.write(f"{user_id}\n")


def get_users_count() -> int:
    """Возвращает количество пользователей."""
    return len(load_users())


# ============ ДОПОЛНИТЕЛЬНЫЕ ЛИМИТЫ ============

def load_user_limits() -> Dict[tuple, int]:
    """Загружает дополнительные лимиты: {(user_id, base_type): extra_limit}."""
    limits = {}
    if not os.path.exists(USER_LIMITS_FILE):
        return limits
    with open(USER_LIMITS_FILE, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        next(reader, None)  # Пропускаем заголовок
        for row in reader:
            if len(row) >= 3:
                try:
                    user_id = int(row[0])
                    base_type = row[1]
                    extra = int(row[2])
                    limits[(user_id, base_type)] = extra
                except ValueError:
                    pass
    return limits


def get_user_extra_limit(user_id: int, base_type: str) -> int:
    """Возвращает дополнительный лимит для пользователя по типу базы."""
    limits = load_user_limits()
    return limits.get((user_id, base_type), 0)


def set_user_extra_limit(user_id: int, base_type: str, value: int) -> None:
    """Устанавливает дополнительный лимит для пользователя."""
    limits = load_user_limits()
    key = (user_id, base_type)
    limits[key] = value
    
    # Сохраняем
    with open(USER_LIMITS_FILE, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["user_id", "base_type", "extra_limit"])
        for (uid, btype), extra in limits.items():
            writer.writerow([uid, btype, extra])


# ============ СТАТУСЫ ПОЛЬЗОВАТЕЛЕЙ ============
# Статусы: pending (ожидает), approved (одобрен), banned (забанен)

def load_user_statuses() -> Dict[int, str]:
    """Загружает статусы пользователей: {user_id: status}."""
    statuses = {}
    if not os.path.exists(USER_STATUS_FILE):
        return statuses
    with open(USER_STATUS_FILE, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        next(reader, None)
        for row in reader:
            if len(row) >= 2:
                try:
                    user_id = int(row[0])
                    status = row[1]
                    statuses[user_id] = status
                except ValueError:
                    pass
    return statuses


def get_user_status(user_id: int) -> Optional[str]:
    """Возвращает статус пользователя (pending/approved/banned) или None если не зарегистрирован."""
    statuses = load_user_statuses()
    return statuses.get(user_id)


def set_user_status(user_id: int, status: str) -> None:
    """Устанавливает статус пользователя."""
    statuses = load_user_statuses()
    statuses[user_id] = status
    
    with open(USER_STATUS_FILE, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["user_id", "status"])
        for uid, st in statuses.items():
            writer.writerow([uid, st])


def is_user_approved(user_id: int) -> bool:
    """Проверяет, одобрен ли пользователь."""
    return get_user_status(user_id) == "approved"


def is_user_banned(user_id: int) -> bool:
    """Проверяет, забанен ли пользователь."""
    return get_user_status(user_id) == "banned"


def is_user_pending(user_id: int) -> bool:
    """Проверяет, ожидает ли пользователь одобрения."""
    return get_user_status(user_id) == "pending"


# ============ РАБОТА С ТОПИКАМИ ПОДДЕРЖКИ ============

def load_support_topics() -> Dict[int, int]:
    """Загружает связи user_id -> topic_id из файла."""
    topics = {}
    if not os.path.exists(SUPPORT_TOPICS_FILE):
        return topics
    with open(SUPPORT_TOPICS_FILE, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        next(reader, None)  # Пропускаем заголовок
        for row in reader:
            if len(row) >= 2:
                try:
                    user_id = int(row[0])
                    topic_id = int(row[1])
                    topics[user_id] = topic_id
                except ValueError:
                    pass
    return topics


def save_support_topic(user_id: int, topic_id: int) -> None:
    """Сохраняет связь user_id -> topic_id."""
    topics = load_support_topics()
    topics[user_id] = topic_id
    
    with open(SUPPORT_TOPICS_FILE, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["user_id", "topic_id"])
        for uid, tid in topics.items():
            writer.writerow([uid, tid])


def get_user_by_topic(topic_id: int) -> Optional[int]:
    """Находит user_id по topic_id."""
    topics = load_support_topics()
    for uid, tid in topics.items():
        if tid == topic_id:
            return uid
    return None


# ============ ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ============

def clean_value(val) -> Optional[str]:
    """Убирает .0 у чисел, знак = в начале, возвращает строку."""
    if val is None:
        return None
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    s = str(val).strip()
    # Убираем знак = в начале (Excel иногда добавляет для формул)
    if s.startswith("="):
        s = s[1:]
    return s if s else None


def ensure_csv_exists() -> None:
    """Проверяет наличие CSV-файлов. Создаёт пустые, если нет."""
    for key, info in BASE_TYPES.items():
        csv_path = info["csv"]
        if not os.path.exists(csv_path):
            with open(csv_path, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Value", "ID", "Username", "Date"])
            print(f"Создан пустой файл: {csv_path}")


LEADS_CSV_HEADER = ["Value", "User_ID", "Username", "Date", "Источник", "Ссылка"]


def get_current_lead_day() -> str:
    """Возвращает дату текущего «дня» для лидов (20:00 — граница, после неё новый день)."""
    tz = ZoneInfo(LEAD_TIMEZONE)
    now = datetime.now(tz)
    if now.hour >= LEAD_DAY_CUTOFF_HOUR:
        next_day = now.date() + timedelta(days=1)
        return next_day.strftime("%Y-%m-%d")
    return now.date().strftime("%Y-%m-%d")


def get_yesterday_lead_day() -> str:
    """Возвращает дату вчерашнего «дня» для лидов."""
    today = get_current_lead_day()
    d = datetime.strptime(today, "%Y-%m-%d").date() - timedelta(days=1)
    return d.strftime("%Y-%m-%d")


def _get_daily_leads_path(lead_type: str, date: str) -> str:
    """Путь к дневному CSV для категории лидов."""
    info = LEAD_TYPES.get(lead_type)
    if not info:
        return ""
    base_csv = info["csv"]
    # leads_telegram.csv -> leads_telegram_2025-01-28.csv
    base_name = base_csv.removesuffix(".csv")
    return f"{base_name}_{date}.csv"


def ensure_leads_csv_exists() -> None:
    """Проверяет наличие CSV-файлов для лидов. Создаёт пустые, если нет."""
    for key, info in LEAD_TYPES.items():
        csv_path = info["csv"]
        if not os.path.exists(csv_path):
            with open(csv_path, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(LEADS_CSV_HEADER)
            print(f"Создан пустой файл лидов: {csv_path}")


# ============ РАБОТА С ЛИДАМИ ============

import re


def normalize_contact(contact: str) -> str:
    """Нормализует контакт для сравнения (убирает @, ссылки, нормализует номера)."""
    if not contact:
        return ""
    
    c = contact.strip().lower()
    
    # Полные ссылки Юла/mail.ru — оставляем как есть для сравнения дубликатов
    if "mail.ru" in c or "youla.ru" in c:
        return c.replace("https://", "").replace("http://", "").replace("www.", "")
    
    # Убираем протоколы и www
    c = c.replace("https://", "").replace("http://", "").replace("www.", "")
    
    # Убираем @ и домены для username
    c = c.replace("@", "").replace("t.me/", "").replace("vk.com/", "").replace("vk.ru/", "").replace("instagram.com/", "").replace("avito.ru/", "")
    
    # Для номеров: убираем пробелы, скобки, дефисы
    c_digits = re.sub(r'[\s\-\(\)\+]', '', c)
    
    # Если это номер (только цифры)
    if c_digits.isdigit():
        # 8XXXXXXXXXX -> 7XXXXXXXXXX
        if c_digits.startswith("8") and len(c_digits) == 11:
            c_digits = "7" + c_digits[1:]
        # Убираем + если есть
        if c_digits.startswith("7") and len(c_digits) == 11:
            return c_digits  # 7XXXXXXXXXX
        return c_digits
    
    # Иначе возвращаем как username (без @ и доменов)
    return c


# Ключевые слова: если есть в тексте рядом с лидом — категория "Самостоятельные лиды"
SELF_LEAD_KEYWORDS = re.compile(r'\b(сам|сама|самостоятельно)\b', re.IGNORECASE)

# Ключевые слова: если есть в тексте рядом с лидом — категория "Юла"
YULA_LEAD_KEYWORDS = re.compile(r'\bюла\b', re.IGNORECASE)

# Ключевые слова: если есть в тексте рядом с лидом — категория "Кворк"
KWORK_LEAD_KEYWORDS = re.compile(r'\bкворк\b', re.IGNORECASE)


def extract_contacts_from_text(text: str) -> List[str]:
    """Извлекает контакты из текста: @username, номера телефонов, ссылки."""
    contacts = []
    if not text:
        return contacts
    
    # @username (Telegram/Instagram) — минимум 4 символа
    usernames = re.findall(r'@([a-zA-Z0-9_]{4,32})', text)
    contacts.extend([u for u in usernames])
    
    # t.me/username или https://t.me/username
    tg_links = re.findall(r'(?:https?://)?t\.me/([a-zA-Z0-9_]+)', text, re.IGNORECASE)
    contacts.extend([u for u in tg_links])
    
    # vk.com/id123, vk.ru/o.kornilova2015 и т.д. — включая точки в username
    vk_links = re.findall(r'(?:https?://)?(?:www\.)?vk\.(com|ru)/([a-zA-Z0-9_.\-]+)', text, re.IGNORECASE)
    for domain, username in vk_links:
        clean_id = username.split("?")[0].strip()  # убираем query-параметры
        if clean_id:
            contacts.append(f"vk.{domain.lower()}/{clean_id}")
    
    # avito.ru/... (объявления, бренды и т.д.)
    avito_links = re.findall(r'(?:https?://)?(?:www\.)?avito\.ru/([a-zA-Z0-9_/\-]+)', text, re.IGNORECASE)
    # Убираем query-параметры и сохраняем путь
    for path in avito_links:
        path_clean = path.split("?")[0].strip("/")
        if path_clean:
            contacts.append(f"avito.ru/{path_clean}")
    
    # instagram.com/username (включая l.instagram.com, ?igsh=...)
    ig_links = re.findall(r'(?:https?://)?(?:[a-zA-Z0-9\-]+\.)?instagram\.com/([a-zA-Z0-9_.\-]+)', text, re.IGNORECASE)
    for u in ig_links:
        clean_u = u.split("?")[0].strip().rstrip("/")
        if clean_u:
            contacts.append(f"instagram.com/{clean_u}")
    
    # Юла / mail.ru (trk.mail.ru, la.youla.ru, m.youla.ru, youla.ru и др.) — сохраняем полную ссылку
    yula_links = re.findall(
        r'https?://(?:trk\.mail\.ru/[\S]+|(?:[a-zA-Z0-9\-]+\.)?youla\.ru/[\S]+)',
        text,
        re.IGNORECASE,
    )
    for url in yula_links:
        url_clean = url.rstrip('.,;:!?')
        if url_clean:
            contacts.append(url_clean)
    
    # ok.ru (Одноклассники) — profile/ID и username
    ok_profile_ids = re.findall(
        r'(?:https?://)?(?:www\.)?ok\.ru/profile/(\d+)',
        text,
        re.IGNORECASE,
    )
    for pid in ok_profile_ids:
        contacts.append(f"ok.ru/profile/{pid}")
    ok_usernames = re.findall(
        r'(?:https?://)?(?:www\.)?ok\.ru/([a-zA-Z0-9_.\-]+)',
        text,
        re.IGNORECASE,
    )
    for u in ok_usernames:
        if u == "profile" or u.startswith("profile/"):
            continue  # уже добавлены через ok_profile_ids
        clean = u.split("?")[0].rstrip("/")
        if clean and f"ok.ru/{clean}" not in contacts:
            contacts.append(f"ok.ru/{clean}")
    
    # Телефонные номера — исключаем те, что являются подстрокой ID из ok.ru (586438915595 → 86438915595)
    phones = re.findall(r'[\+]?[78][\s\-]?[\(]?\d{3}[\)]?[\s\-]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}', text)
    for p in phones:
        digits = re.sub(r'\D', '', p)
        if not any(digits in pid or pid in digits for pid in ok_profile_ids):
            contacts.append(p)
    
    # kwork.ru/user/username или kwork.ru/username (Кворк)
    kwork_links = re.findall(r'(?:https?://)?(?:www\.)?kwork\.ru/([a-zA-Z0-9_/\-]+)', text, re.IGNORECASE)
    for u in kwork_links:
        clean = u.split("?")[0].strip("/")
        if clean:
            contacts.append(f"kwork.ru/{clean}")
    
    # Если в тексте есть "кворк" — извлекаем также plain username/ID (ElenaTuz и т.д.)
    if KWORK_LEAD_KEYWORDS.search(text):
        words = text.split()
        for word in words:
            clean_word = re.sub(r'[^\w]', '', word)
            if re.match(r'^[a-zA-Z0-9_]{4,32}$', clean_word):
                cw_lower = clean_word.lower()
                if cw_lower not in {'https', 'http', 'kwork', 'кворк'}:
                    if not any(x in cw_lower for x in ('http', 'www', 'tme', 'vkru', 'avitoru')):
                        contacts.append(clean_word)
    
    # Если не нашли — пробуем первый токен («@LinaSmirnov тг» → берём @LinaSmirnov)
    if not contacts and text.strip():
        tokens = text.strip().split()
        if len(tokens) > 1:
            contacts = extract_contacts_from_text(tokens[0])
    
    # Убираем дубликаты с учётом нормализации.
    # При коллизии (одинаковый username на разных платформах) приоритет у формы с явной платформой
    # (instagram.com/, vk.ru/ и т.д.), чтобы категория определялась верно.
    def has_platform_prefix(s: str) -> bool:
        return any(s.lower().startswith(p) for p in ("instagram.com/", "vk.com/", "vk.ru/", "t.me/", "avito.ru/", "kwork.ru/", "ok.ru/"))
    unique = {}
    for c in contacts:
        normalized = normalize_contact(c)
        if not normalized:
            continue
        existing = unique.get(normalized)
        if existing is None:
            unique[normalized] = c
        elif has_platform_prefix(c) and not has_platform_prefix(existing):
            unique[normalized] = c
    return list(unique.values())


def determine_contact_type(contact: str, user_id: int) -> Optional[str]:
    """Определяет тип контакта по выданным пользователю базам (или всей базе)."""
    # Ссылки на Авито — сразу категория avito
    if contact and ("avito.ru" in contact.lower() or contact.lower().startswith("avito")):
        return "avito"
    
    # Ссылки на Юлу (mail.ru, youla.ru)
    if contact and ("mail.ru" in contact.lower() or "youla.ru" in contact.lower()):
        return "yula"
    
    # Ссылки на Кворк (kwork.ru)
    if contact and "kwork.ru" in contact.lower():
        return "kwork"
    
    # Ссылки на Instagram (instagram.com, l.instagram.com и т.д.)
    if contact and "instagram.com" in contact.lower():
        return "instagram"
    
    # Ссылки на VK — сразу категория ВКонтакте (как Instagram)
    if contact and ("vk.com" in contact.lower() or "vk.ru" in contact.lower()):
        return "vk"
    
    # Ссылки на Одноклассники (ok.ru)
    if contact and "ok.ru" in contact.lower():
        return "ok"
    
    # Остальное — проверяем в базах выдачи
    # Проверяем и в базе: возможно выдан пользователю
    contact_normalized = normalize_contact(contact)
    
    # Сначала проверяем выданные конкретному пользователю
    for key, info in BASE_TYPES.items():
        csv_path = info["csv"]
        if not os.path.exists(csv_path):
            continue
        
        rows = _read_csv(csv_path)
        for row in rows[1:]:  # Пропускаем заголовок
            if len(row) < 4:
                continue
            value, assigned_id, *_ = row
            
            # Проверяем, выдан ли этот контакт пользователю
            if assigned_id and str(assigned_id).strip():
                try:
                    if int(assigned_id) == user_id:
                        # Нормализуем и сравниваем
                        value_clean = clean_value(value) or ""
                        value_normalized = normalize_contact(value_clean)
                        
                        if value_normalized == contact_normalized:
                            return key
                except (ValueError, AttributeError):
                    pass
    
    # Если не нашли в выданных пользователю — ищем по всей базе
    for key, info in BASE_TYPES.items():
        csv_path = info["csv"]
        if not os.path.exists(csv_path):
            continue
        
        rows = _read_csv(csv_path)
        for row in rows[1:]:
            if len(row) < 1:
                continue
            value = row[0]
            value_clean = clean_value(value) or ""
            value_normalized = normalize_contact(value_clean)
            
            if value_normalized == contact_normalized:
                return key
    
    return None


def check_lead_duplicate(contact: str) -> Optional[tuple]:
    """Проверяет, существует ли лид в базе. Возвращает (lead_type, user_id, username) если найден."""
    contact_normalized = normalize_contact(contact)
    
    for key, info in LEAD_TYPES.items():
        csv_path = info["csv"]
        if not os.path.exists(csv_path):
            continue
        
        rows = _read_csv(csv_path)
        for row in rows[1:]:
            if len(row) < 4:
                continue
            value, user_id_str, username, *_ = row
            
            value_normalized = normalize_contact(value or "")
            if value_normalized == contact_normalized:
                return (key, user_id_str, username)
    
    return None


def add_lead(contact: str, lead_type: str, user_id: int, username: str, source: str = "", message_link: str = "") -> bool:
    """Добавляет лид в базу. source: '' | 'база' | 'самостоятельный'. message_link: ссылка на сообщение отчёта/поддержки."""
    info = LEAD_TYPES.get(lead_type)
    if not info:
        return False

    csv_path = info["csv"]
    rows = _read_csv(csv_path)

    # Убеждаемся, что у всех строк 6 колонок (для совместимости со старыми файлами)
    if rows and len(rows[0]) < 6 and rows[0][0] == "Value":
        rows[0] = LEADS_CSV_HEADER
    for i in range(1, len(rows)):
        while len(rows[i]) < 6:
            rows[i].append("")

    now = datetime.now().strftime("%Y.%m.%d %H:%M:%S")
    new_row = [contact, user_id, username or "нет", now, source or "", message_link or ""]
    rows.append(new_row)

    _write_csv(csv_path, rows)

    # Дубликаты проверяются по общей базе; лид также добавляем в базу дня (20:00 — граница дня)
    daily_path = _get_daily_leads_path(lead_type, get_current_lead_day())
    if daily_path:
        if not os.path.exists(daily_path):
            with open(daily_path, "w", encoding="utf-8", newline="") as f:
                csv.writer(f).writerow(LEADS_CSV_HEADER)
        with open(daily_path, "a", encoding="utf-8", newline="") as f:
            csv.writer(f).writerow(new_row)

    return True


def delete_lead(contact: str) -> Optional[tuple]:
    """Удаляет лид из базы. Возвращает (lead_type, contact) если успешно, None если не найден."""
    contact_clean = contact.strip().lower().replace("@", "").replace("t.me/", "")
    
    for key, info in LEAD_TYPES.items():
        csv_path = info["csv"]
        if not os.path.exists(csv_path):
            continue
        
        rows = _read_csv(csv_path)
        new_rows = [rows[0]]  # Заголовок
        found = False
        found_value = None
        
        for row in rows[1:]:
            if len(row) < 4:
                new_rows.append(row)
                continue
            
            value = row[0] or ""
            value_clean = value.strip().lower().replace("@", "").replace("t.me/", "")
            
            if value_clean == contact_clean:
                found = True
                found_value = value
                # Пропускаем эту строку (удаляем)
            else:
                new_rows.append(row)
        
        if found:
            _write_csv(csv_path, new_rows)
            return (key, found_value)
    
    return None


# ============ РАБОТА С CSV ============

def _read_csv(path: str) -> List[List[str]]:
    """Читает CSV и возвращает список строк."""
    if not os.path.exists(path):
        return [["Value", "ID", "Username", "Date"]]
    with open(path, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        return list(reader)


def _write_csv(path: str, rows: List[List[str]]) -> None:
    """Записывает список строк в CSV."""
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def _count_user_records(rows: List[List[str]], user_id: int) -> int:
    """Считает, сколько записей выдано пользователю."""
    count = 0
    for row in rows[1:]:
        if len(row) > 1 and row[1]:
            try:
                if int(row[1]) == user_id:
                    count += 1
            except (ValueError, TypeError):
                continue
    return count


def _assign_records_csv(
    rows: List[List[str]],
    count: int,
    user_id: int,
    username: str,
) -> List[str]:
    """Берёт свободные записи, помечает как выданные."""
    taken: List[str] = []
    now = datetime.now(timezone.utc).strftime("%Y.%m.%d %H:%M:%S")

    for row in rows[1:]:
        if len(taken) >= count:
            break

        if len(row) < 4:
            row.extend([""] * (4 - len(row)))

        if row[1]:  # Уже выдано
            continue

        value = row[0].strip()
        # Убираем знак = в начале (Excel иногда добавляет)
        if value.startswith("="):
            value = value[1:]
        if not value:
            continue

        row[1] = str(user_id)
        row[2] = username or ""
        row[3] = now
        taken.append(value)

    return taken


def _get_existing_values(rows: List[List[str]]) -> Set[str]:
    """Возвращает множество всех значений в базе."""
    values = set()
    for row in rows[1:]:
        if row and row[0]:
            values.add(row[0].strip().lower())
    return values


def _add_new_values(csv_path: str, new_values: List[str]) -> int:
    """
    Добавляет новые значения в CSV с проверкой на дубликаты.
    Возвращает количество добавленных записей.
    """
    rows = _read_csv(csv_path)
    existing = _get_existing_values(rows)

    added = 0
    for val in new_values:
        val_clean = clean_value(val)
        if val_clean and val_clean.lower() not in existing:
            rows.append([val_clean, "", "", ""])
            existing.add(val_clean.lower())
            added += 1

    if added > 0:
        _write_csv(csv_path, rows)

    return added


def _process_excel_upload_sync(file_bytes: bytes, upload_type: str) -> tuple[List[str], Optional[str]]:
    """
    Синхронная обработка загруженного Excel (запускается в отдельном потоке,
    чтобы не блокировать бота при больших файлах). Возвращает (список результатов, ошибка или None).
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
        results = []

        if upload_type == "all":
            for sheet_name in wb.sheetnames:
                base_key = EXCEL_SHEET_MAP.get(sheet_name)
                if not base_key:
                    results.append(f"⚠️ Лист «{sheet_name}» — неизвестный тип, пропущен")
                    continue

                ws = wb[sheet_name]
                new_values = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    val = clean_value(row[0] if row else None)
                    if val and val.lower() not in ("value", "значение", "контакт", "данные"):
                        new_values.append(val)

                if new_values:
                    csv_path = BASE_TYPES[base_key]["csv"]
                    added = _add_new_values(csv_path, new_values)
                    info = BASE_TYPES[base_key]
                    results.append(
                        f"✅ «{info['name']}» — добавлено {added} из {len(new_values)}"
                    )
                else:
                    results.append(f"⚠️ Лист «{sheet_name}» — пустой")
        else:
            ws = wb.active
            new_values = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = clean_value(row[0] if row else None)
                if val and val.lower() not in ("value", "значение", "контакт", "данные"):
                    new_values.append(val)

            if new_values:
                csv_path = BASE_TYPES[upload_type]["csv"]
                added = _add_new_values(csv_path, new_values)
                info = BASE_TYPES[upload_type]
                results.append(
                    f"✅ «{info['name']}» — добавлено {added} из {len(new_values)} "
                    f"(дубликатов пропущено: {len(new_values) - added})"
                )
            else:
                results.append("⚠️ Файл пустой или не содержит данных в первом столбце")

        wb.close()
        return (results, None)
    except Exception as e:
        return ([], str(e))


# ============ ВЫДАЧА ДАННЫХ ============

async def allocate_for_user(base_key: str, user_id: int, username: str) -> tuple[List[str], str]:
    """
    Универсальная функция выдачи данных из любой базы.
    
    Возвращает кортеж: (список_контактов, причина_отказа)
    - причина: None — успех, "already_got" — уже получил, "not_enough" — недостаточно контактов
    """
    info = BASE_TYPES[base_key]
    csv_path = info["csv"]
    base_limit = info["limit"]
    
    # Учитываем дополнительный лимит от менеджера
    extra_limit = get_user_extra_limit(user_id, base_key)
    total_allowed = base_limit + extra_limit

    async with csv_lock:
        def _worker() -> tuple[List[str], str]:
            rows = _read_csv(csv_path)

            # Проверяем, сколько уже выдано этому пользователю
            current = _count_user_records(rows, user_id)
            if current >= total_allowed:
                return ([], "already_got")

            # Сколько ещё можно выдать
            can_give = total_allowed - current

            # Считаем свободные контакты (где нет ID)
            free_count = sum(1 for r in rows if len(r) < 2 or not r[1])
            if free_count < can_give:
                return ([], "not_enough")

            taken = _assign_records_csv(rows, can_give, user_id, username)

            if taken:
                _write_csv(csv_path, rows)

            return (taken, None)

        return await asyncio.to_thread(_worker)


# ============ СОЗДАНИЕ ФАЙЛОВ ============

def _create_txt_file(values: List[str], prefix: str) -> tuple[io.BytesIO, str]:
    """Создаёт txt-файл в памяти."""
    content = "\n".join(values)
    buffer = io.BytesIO(content.encode("utf-8"))
    filename = f"{prefix}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.txt"
    return buffer, filename


def _create_full_excel() -> tuple[io.BytesIO, str]:
    """Собирает все CSV-базы в один Excel-файл."""
    wb = Workbook()
    first = True

    for key, info in BASE_TYPES.items():
        csv_path = info["csv"]
        sheet_name = info["name"]

        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        if os.path.exists(csv_path):
            rows = _read_csv(csv_path)
            for row in rows:
                ws.append(row)
        else:
            ws.append(["Value", "ID", "Username", "Date"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"full_base_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return buffer, filename


def _create_leads_excel() -> tuple[io.BytesIO, str]:
    """Собирает все CSV-базы лидов в один Excel-файл."""
    wb = Workbook()
    first = True

    for key, info in LEAD_TYPES.items():
        csv_path = info["csv"]
        sheet_name = info["name"]

        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        if os.path.exists(csv_path):
            rows = _read_csv(csv_path)
            for row in rows:
                ws.append(row)
        else:
            ws.append(LEADS_CSV_HEADER)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"leads_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return buffer, filename


def _create_daily_leads_excel(date: str) -> tuple[io.BytesIO, str]:
    """Собирает дневные CSV-базы лидов в один Excel. date: YYYY-MM-DD."""
    wb = Workbook()
    first = True

    for key, info in LEAD_TYPES.items():
        daily_path = _get_daily_leads_path(key, date)
        sheet_name = f"{info['name']} ({date})"

        if first:
            ws = wb.active
            ws.title = sheet_name[:31]
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name[:31])

        if daily_path and os.path.exists(daily_path):
            rows = _read_csv(daily_path)
            for row in rows:
                ws.append(row)
        else:
            ws.append(LEADS_CSV_HEADER)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"leads_day_{date}.xlsx"
    return buffer, filename


def _create_user_leads_excel(user_id: int, all_time: bool, date: str = "") -> tuple[io.BytesIO, str]:
    """Создаёт Excel с лидами пользователя. all_time=True — из основных CSV, False — из дневных."""
    wb = Workbook()
    first = True
    user_id_str = str(user_id)

    for key, info in LEAD_TYPES.items():
        if all_time:
            csv_path = info["csv"]
            rows = _read_csv(csv_path) if os.path.exists(csv_path) else [LEADS_CSV_HEADER]
        else:
            daily_path = _get_daily_leads_path(key, date)
            rows = _read_csv(daily_path) if daily_path and os.path.exists(daily_path) else [LEADS_CSV_HEADER]

        user_rows = [rows[0]]
        for row in rows[1:]:
            if len(row) >= 2 and str(row[1]).strip() == user_id_str:
                user_rows.append(row)

        if len(user_rows) <= 1:
            continue

        sheet_name = f"{info['name']} ({len(user_rows) - 1})"
        if first:
            ws = wb.active
            ws.title = sheet_name[:31]
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name[:31])
        for row in user_rows:
            ws.append(row)

    if first:
        ws = wb.active
        ws.append(LEADS_CSV_HEADER)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    if all_time:
        filename = f"leads_user_{user_id}_all.xlsx"
    else:
        filename = f"leads_user_{user_id}_day_{date}.xlsx"
    return buffer, filename


# ============ КЛАВИАТУРЫ ============

def get_main_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📦 Получить списки контактов")],
            [KeyboardButton(text="📋 Отчёт по лидам")],
            [KeyboardButton(text="💬 Написать в поддержку")],
            [KeyboardButton(text="📊 Статистика лидов")],
        ],
        resize_keyboard=True,
    )


def get_support_keyboard() -> ReplyKeyboardMarkup:
    """Клавиатура в режиме поддержки."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="⬅️ Назад")],
        ],
        resize_keyboard=True,
    )


def get_report_keyboard() -> ReplyKeyboardMarkup:
    """Клавиатура при сдаче отчёта."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📤 Отправить отчёт")],
            [KeyboardButton(text="❌ Отмена")],
        ],
        resize_keyboard=True,
    )


def get_registration_keyboard() -> ReplyKeyboardMarkup:
    """Клавиатура для регистрации. one_time_keyboard — скрывается после первого нажатия."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="✅ Отправить приглашение")],
        ],
        resize_keyboard=True,
        one_time_keyboard=True,
    )


def get_user_choice_keyboard() -> ReplyKeyboardMarkup:
    """Клавиатура выбора типа контактов для пользователя."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="📱 Telegram"),
                KeyboardButton(text="💬 WhatsApp"),
            ],
            [
                KeyboardButton(text="📨 Max"),
                KeyboardButton(text="📞 Viber"),
            ],
            [
                KeyboardButton(text="📷 Нельзяграм"),
                KeyboardButton(text="👥 ВКонтакте"),
            ],
            [
                KeyboardButton(text="🟠 Одноклассники"),
                KeyboardButton(text="📧 Почта"),
            ],
            [KeyboardButton(text="🆕 Получить новые контакты")],
            [KeyboardButton(text="⬅️ Назад")],
        ],
        resize_keyboard=True,
    )


def get_admin_keyboard() -> ReplyKeyboardMarkup:
    """Клавиатура админа."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📥 Загрузить Базу данных")],
            [KeyboardButton(text="📤 Выкачать Базу данных")],
            [KeyboardButton(text="🗑 Удалить всю базу данных")],
            [KeyboardButton(text="⬅️ Выход из админки")],
        ],
        resize_keyboard=True,
    )


def get_delete_confirm_keyboard() -> ReplyKeyboardMarkup:
    """Клавиатура подтверждения удаления."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="✅ Да, удалить всё")],
            [KeyboardButton(text="❌ Отмена")],
        ],
        resize_keyboard=True,
    )


def get_admin_upload_choice_keyboard() -> ReplyKeyboardMarkup:
    """Клавиатура выбора типа базы для загрузки."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="📱 Загрузить Telegram"),
                KeyboardButton(text="💬 Загрузить WhatsApp"),
            ],
            [
                KeyboardButton(text="📨 Загрузить Max"),
                KeyboardButton(text="📞 Загрузить Viber"),
            ],
            [
                KeyboardButton(text="📷 Загрузить Нельзяграм"),
                KeyboardButton(text="👥 Загрузить ВКонтакте"),
            ],
            [
                KeyboardButton(text="🟠 Загрузить Одноклассники"),
                KeyboardButton(text="📧 Загрузить Почта"),
            ],
            [KeyboardButton(text="📚 Загрузить ВСЕ листы из файла")],
            [KeyboardButton(text="⬅️ Отмена")],
        ],
        resize_keyboard=True,
    )


def get_lead_category_keyboard() -> ReplyKeyboardMarkup:
    """Клавиатура выбора категории (Reply — только для личных чатов)."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="📱 Telegram"),
                KeyboardButton(text="💬 WhatsApp"),
            ],
            [
                KeyboardButton(text="📨 Max"),
                KeyboardButton(text="📞 Viber"),
            ],
            [
                KeyboardButton(text="📷 Нельзяграм"),
                KeyboardButton(text="👥 ВКонтакте"),
            ],
            [
                KeyboardButton(text="🟠 Одноклассники"),
                KeyboardButton(text="📧 Почта"),
            ],
            [KeyboardButton(text="🟢 Авито")],
            [KeyboardButton(text="🔵 Самостоятельные лиды")],
            [KeyboardButton(text="⬅️ Отмена")],
        ],
        resize_keyboard=True,
    )


def get_lead_category_inline_keyboard() -> InlineKeyboardMarkup:
    """Inline-клавиатура выбора категории (работает в группах)."""
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="📱 Telegram", callback_data="lead_cat_telegram"),
            InlineKeyboardButton(text="💬 WhatsApp", callback_data="lead_cat_whatsapp"),
        ],
        [
            InlineKeyboardButton(text="📨 Max", callback_data="lead_cat_max"),
            InlineKeyboardButton(text="📞 Viber", callback_data="lead_cat_viber"),
        ],
        [
            InlineKeyboardButton(text="📷 Нельзяграм", callback_data="lead_cat_instagram"),
            InlineKeyboardButton(text="👥 ВКонтакте", callback_data="lead_cat_vk"),
        ],
        [
            InlineKeyboardButton(text="🟠 Одноклассники", callback_data="lead_cat_ok"),
            InlineKeyboardButton(text="📧 Почта", callback_data="lead_cat_email"),
        ],
        [
            InlineKeyboardButton(text="🟢 Авито", callback_data="lead_cat_avito"),
            InlineKeyboardButton(text="🟡 Юла", callback_data="lead_cat_yula"),
        ],
        [InlineKeyboardButton(text="🟣 Кворк", callback_data="lead_cat_kwork")],
        [InlineKeyboardButton(text="🌐 Прочие соц. сети", callback_data="lead_cat_other_social")],
        [InlineKeyboardButton(text="🔵 Самостоятельные лиды", callback_data="lead_cat_self")],
        [InlineKeyboardButton(text="⬅️ Отмена", callback_data="lead_cat_cancel")],
    ])


def get_report_category_inline_keyboard(idx: int) -> InlineKeyboardMarkup:
    """Inline-клавиатура выбора категории для лида в отчёте."""
    prefix = f"report_cat_{idx}_"
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="📱 Telegram", callback_data=prefix + "telegram"),
            InlineKeyboardButton(text="💬 WhatsApp", callback_data=prefix + "whatsapp"),
        ],
        [
            InlineKeyboardButton(text="📨 Max", callback_data=prefix + "max"),
            InlineKeyboardButton(text="📞 Viber", callback_data=prefix + "viber"),
        ],
        [
            InlineKeyboardButton(text="📷 Нельзяграм", callback_data=prefix + "instagram"),
            InlineKeyboardButton(text="👥 ВКонтакте", callback_data=prefix + "vk"),
        ],
        [
            InlineKeyboardButton(text="🟠 Одноклассники", callback_data=prefix + "ok"),
            InlineKeyboardButton(text="📧 Почта", callback_data=prefix + "email"),
        ],
        [
            InlineKeyboardButton(text="🟢 Авито", callback_data=prefix + "avito"),
            InlineKeyboardButton(text="🟡 Юла", callback_data=prefix + "yula"),
        ],
        [InlineKeyboardButton(text="🟣 Кворк", callback_data=prefix + "kwork")],
        [InlineKeyboardButton(text="🌐 Прочие соц. сети", callback_data=prefix + "other_social")],
        [InlineKeyboardButton(text="⬅️ Отмена отчёта", callback_data=prefix + "cancel")],
    ])


# ============ МАППИНГ КНОПОК ============

# Кнопки пользователя -> ключ базы
USER_BUTTON_MAP = {
    "📱 Telegram": "telegram",
    "💬 WhatsApp": "whatsapp",
    "📨 Max": "max",
    "📞 Viber": "viber",
    "📷 Нельзяграм": "instagram",
    "👥 ВКонтакте": "vk",
    "🟠 Одноклассники": "ok",
    "📧 Почта": "email",
}

# Кнопки админа для загрузки -> ключ базы
ADMIN_UPLOAD_MAP = {
    "📱 Загрузить Telegram": "telegram",
    "💬 Загрузить WhatsApp": "whatsapp",
    "📨 Загрузить Max": "max",
    "📞 Загрузить Viber": "viber",
    "📷 Загрузить Нельзяграм": "instagram",
    "👥 Загрузить ВКонтакте": "vk",
    "🟠 Загрузить Одноклассники": "ok",
    "📧 Загрузить Почта": "email",
    "📚 Загрузить ВСЕ листы из файла": "all",
}


# ============ ХЕНДЛЕРЫ ============

async def on_start(message: Message, state: FSMContext, bot: Bot) -> None:
    await state.clear()
    
    user = message.from_user
    if not user:
        return
    
    user_id = user.id
    status = get_user_status(user_id)
    
    # Сохраняем пользователя в счётчик
    save_user(user_id)
    
    # Проверяем статус пользователя
    if status == "banned":
        await message.answer(
            "🚫 Ваш аккаунт заблокирован.\n\n"
            "Обратитесь к администратору для разблокировки."
        )
        return
    
    if status == "approved":
        # Пользователь одобрен — показываем главное меню
        text = (
            "Привет!\n\n"
            "Этот бот выдаёт тебе списки контактов по которым нужно отправлять сообщения.\n\n"
            "Нажми кнопку ниже, затем выбери соц сеть или мессенджер где тебе удобнее работать."
        )
        await message.answer(text, reply_markup=get_main_keyboard())
        return
    
    if status == "pending":
        # Уже отправил заявку — ждёт одобрения, кнопку убираем
        await message.answer(
            "⏳ Ваша заявка уже отправлена!\n\n"
            "Ожидайте подтверждения от администратора.",
            reply_markup=ReplyKeyboardRemove(remove_keyboard=True),
        )
        return
    
    # Новый пользователь — показываем экран регистрации
    text = (
        "Если вы получили доступ к данному боту, значит вы уже прошли собеседование.\n\n"
        "Нажмите на кнопку ниже, админ примет приглашение и начнем ✅"
    )
    await message.answer(text, reply_markup=get_registration_keyboard())


async def on_send_request(message: Message, bot: Bot) -> None:
    """Пользователь нажал 'Отправить приглашение'."""
    user = message.from_user
    if not user:
        return
    
    # Проверяем что это личный чат
    if message.chat.type != "private":
        return
    
    user_id = user.id
    status = get_user_status(user_id)
    
    if status == "approved":
        await message.answer("Вы уже зарегистрированы!", reply_markup=get_main_keyboard())
        return
    
    if status == "pending":
        await message.answer(
            "⏳ Ваша заявка уже отправлена! Ожидайте подтверждения.",
            reply_markup=ReplyKeyboardRemove(remove_keyboard=True),
        )
        return
    
    if status == "banned":
        await message.answer("🚫 Ваш аккаунт заблокирован.")
        return
    
    # Создаём заявку
    set_user_status(user_id, "pending")
    
    # Создаём топик для пользователя
    user_name = user.full_name or f"User {user_id}"
    if user.username:
        user_name += f" (@{user.username})"
    
    try:
        forum_topic = await bot.create_forum_topic(
            chat_id=SUPPORT_GROUP_ID,
            name=f"📝 {user_name[:120]}",
        )
        topic_id = forum_topic.message_thread_id
        save_support_topic(user_id, topic_id)
        
        # Отправляем заявку в топик
        await bot.send_message(
            chat_id=SUPPORT_GROUP_ID,
            message_thread_id=topic_id,
            text=(
                f"📝 НОВАЯ ЗАЯВКА!\n\n"
                f"👤 Пользователь: {user.full_name}\n"
                f"🆔 ID: {user_id}\n"
                f"📱 Username: @{user.username or 'нет'}\n\n"
                f"Для одобрения: /add\n"
                f"Для бана: /ban"
            ),
        )
        
        await message.answer(
            "✅ Заявка отправлена!\n\n"
            "Ожидайте подтверждения от администратора.\n"
            "Вам придёт уведомление когда заявка будет одобрена.",
            reply_markup=ReplyKeyboardRemove(remove_keyboard=True),
        )
    except Exception as e:
        set_user_status(user_id, None)  # Откатываем статус
        await message.answer(f"❌ Ошибка при отправке заявки: {e}")


async def on_add_user(message: Message, bot: Bot) -> None:
    """Команда /add — одобрить пользователя."""
    if message.chat.id != SUPPORT_GROUP_ID:
        return
    
    topic_id = message.message_thread_id
    if not topic_id:
        await message.answer("❌ Эта команда работает только в топике пользователя.")
        return
    
    user_id = get_user_by_topic(topic_id)
    if not user_id:
        await message.answer("❌ Пользователь не найден для этого топика.")
        return
    
    status = get_user_status(user_id)
    if status == "approved":
        await message.answer("ℹ️ Пользователь уже одобрен.")
        return
    
    set_user_status(user_id, "approved")
    await message.answer(f"✅ Пользователь {user_id} одобрен!")
    
    # Уведомляем пользователя
    try:
        await bot.send_message(
            chat_id=user_id,
            text=(
                "🎉 Ваша заявка одобрена!\n\n"
                "Теперь вы можете пользоваться ботом.\n"
                "Нажмите /start чтобы начать."
            ),
            reply_markup=ReplyKeyboardRemove(remove_keyboard=True),
        )
    except Exception:
        pass


async def on_ban_user(message: Message, bot: Bot) -> None:
    """Команда /ban — забанить пользователя."""
    if message.chat.id != SUPPORT_GROUP_ID:
        return
    
    topic_id = message.message_thread_id
    if not topic_id:
        await message.answer("❌ Эта команда работает только в топике пользователя.")
        return
    
    user_id = get_user_by_topic(topic_id)
    if not user_id:
        await message.answer("❌ Пользователь не найден для этого топика.")
        return
    
    set_user_status(user_id, "banned")
    await message.answer(f"🚫 Пользователь {user_id} заблокирован!")
    
    # Уведомляем пользователя
    try:
        await bot.send_message(
            chat_id=user_id,
            text="🚫 Ваш аккаунт заблокирован.\n\nОбратитесь к администратору для разблокировки.",
        )
    except Exception:
        pass


async def on_unban_user(message: Message, bot: Bot) -> None:
    """Команда /unban — разбанить пользователя."""
    if message.chat.id != SUPPORT_GROUP_ID:
        return
    
    topic_id = message.message_thread_id
    if not topic_id:
        await message.answer("❌ Эта команда работает только в топике пользователя.")
        return
    
    user_id = get_user_by_topic(topic_id)
    if not user_id:
        await message.answer("❌ Пользователь не найден для этого топика.")
        return
    
    status = get_user_status(user_id)
    if status != "banned":
        await message.answer("ℹ️ Пользователь не заблокирован.")
        return
    
    set_user_status(user_id, "approved")
    await message.answer(f"✅ Пользователь {user_id} разблокирован!")
    
    # Уведомляем пользователя
    try:
        await bot.send_message(
            chat_id=user_id,
            text=(
                "✅ Ваш аккаунт разблокирован!\n\n"
                "Нажмите /start чтобы продолжить."
            ),
        )
    except Exception:
        pass


async def on_admin_command(message: Message, state: FSMContext) -> None:
    """Обработка секретной команды для получения админ-доступа."""
    # Только в личном чате с ботом
    if message.chat.type != "private":
        return
    
    user = message.from_user
    if not user:
        return

    save_admin(user.id)
    await state.clear()
    await message.answer(
        "✅ Админ-доступ активирован!\n\n"
        "Теперь тебе доступны функции управления базой данных.",
        reply_markup=get_admin_keyboard(),
    )


async def on_chatid(message: Message) -> None:
    """Показывает ID чата (для настройки группы поддержки)."""
    chat = message.chat
    topic_id = message.message_thread_id
    
    text = f"📍 **Информация о чате:**\n\n"
    text += f"Chat ID: `{chat.id}`\n"
    text += f"Тип: {chat.type}\n"
    if chat.title:
        text += f"Название: {chat.title}\n"
    if topic_id:
        text += f"Topic ID: `{topic_id}`\n"
    
    await message.answer(text, parse_mode="Markdown")


async def on_get_online(message: Message) -> None:
    """Показывает количество пользователей бота (только для группы админов)."""
    # Только в группе поддержки
    if message.chat.id != SUPPORT_GROUP_ID:
        return
    
    count = get_users_count()
    await message.answer(
        f"📊 Статистика бота:\n\n"
        f"👥 Всего пользователей: {count}"
    )


async def on_download_db(message: Message) -> None:
    """Выгрузка всей базы данных (только для группы админов)."""
    # Только в группе поддержки
    if message.chat.id != SUPPORT_GROUP_ID:
        return
    
    await message.answer("⏳ Собираю базу данных...")
    
    try:
        file_buffer, filename = await asyncio.to_thread(_create_full_excel)
        document = BufferedInputFile(file_buffer.read(), filename=filename)
        await message.answer_document(
            document=document,
            caption="📤 Полная база данных"
        )
    except Exception as e:
        await message.answer(f"❌ Ошибка при выгрузке: {e}")


async def on_download_lead(message: Message) -> None:
    """Выгрузка базы лидов (только для топика Лиды авто)."""
    # Только в топике "Лиды авто"
    if message.chat.id != SUPPORT_GROUP_ID or message.message_thread_id != LEADS_TOPIC_ID:
        return
    
    await message.answer("⏳ Собираю базу лидов...")
    
    try:
        file_buffer, filename = await asyncio.to_thread(_create_leads_excel)
        document = BufferedInputFile(file_buffer.read(), filename=filename)
        await message.answer_document(
            document=document,
            caption="📤 База лидов"
        )
    except Exception as e:
        await message.answer(f"❌ Ошибка при выгрузке: {e}")


async def on_download_lead_day(message: Message) -> None:
    """Выгрузка базы лидов за текущий день (только для топика Лиды авто)."""
    if message.chat.id != SUPPORT_GROUP_ID or message.message_thread_id != LEADS_TOPIC_ID:
        return

    today = get_current_lead_day()
    await message.answer(f"⏳ Собираю лиды за день {today}...")

    try:
        file_buffer, filename = await asyncio.to_thread(_create_daily_leads_excel, today)
        document = BufferedInputFile(file_buffer.read(), filename=filename)
        await message.answer_document(
            document=document,
            caption=f"📤 Лиды за {today}"
        )
    except Exception as e:
        await message.answer(f"❌ Ошибка при выгрузке: {e}")


async def on_stats(message: Message) -> None:
    """Статистика свободных контактов и выданных за периоды (только для группы админов)."""
    # Только в группе поддержки
    if message.chat.id != SUPPORT_GROUP_ID:
        return
    
    def _count_stats() -> tuple:
        from datetime import timedelta
        
        now = datetime.now(timezone.utc)
        day_ago = now - timedelta(days=1)
        week_ago = now - timedelta(days=7)
        month_ago = now - timedelta(days=30)
        
        free_stats = []  # (name, free, total)
        issued_stats = []  # (name, day, week, month)
        
        for key, info in BASE_TYPES.items():
            csv_path = info["csv"]
            rows = _read_csv(csv_path)
            total = len(rows) - 1  # Минус заголовок
            free = sum(1 for r in rows[1:] if len(r) < 2 or not r[1])
            free_stats.append((info["name"], free, total))
            
            # Считаем выданные за периоды
            day_count = 0
            week_count = 0
            month_count = 0
            
            for row in rows[1:]:
                if len(row) >= 4 and row[3]:  # Есть дата выдачи
                    try:
                        # Формат: "YYYY.MM.DD HH:MM:SS" — считаем UTC для корректного сравнения
                        issued_date = datetime.strptime(row[3], "%Y.%m.%d %H:%M:%S").replace(tzinfo=timezone.utc)
                        if issued_date >= day_ago:
                            day_count += 1
                        if issued_date >= week_ago:
                            week_count += 1
                        if issued_date >= month_ago:
                            month_count += 1
                    except ValueError:
                        pass
            
            issued_stats.append((info["name"], day_count, week_count, month_count))
        
        return free_stats, issued_stats
    
    free_stats, issued_stats = await asyncio.to_thread(_count_stats)
    
    # Свободные контакты
    lines = ["📊 **Свободные контакты:**\n"]
    total_free = 0
    total_all = 0
    
    for name, free, total in free_stats:
        if free == 0:
            status = "🔴"
        elif free < 100:
            status = "🟡"
        else:
            status = "🟢"
        lines.append(f"{status} **{name}**: {free} / {total}")
        total_free += free
        total_all += total
    
    lines.append(f"\n📦 **Итого**: {total_free} свободных / {total_all} всего")
    
    # Выданные за периоды
    lines.append("\n\n📈 **Выдано контактов:**\n")
    lines.append("```")
    lines.append(f"{'Тип':<25} {'Сутки':>7} {'Неделя':>7} {'Месяц':>7}")
    lines.append("-" * 48)
    
    total_day = 0
    total_week = 0
    total_month = 0
    
    for name, day, week, month in issued_stats:
        # Обрезаем длинные названия
        short_name = name[:24] if len(name) > 24 else name
        lines.append(f"{short_name:<25} {day:>7} {week:>7} {month:>7}")
        total_day += day
        total_week += week
        total_month += month
    
    lines.append("-" * 48)
    lines.append(f"{'ИТОГО':<25} {total_day:>7} {total_week:>7} {total_month:>7}")
    lines.append("```")
    
    await message.answer("\n".join(lines), parse_mode="Markdown")


async def on_leadstats(message: Message) -> None:
    """Статистика по лидам (только для топика Лиды авто)."""
    # Только в топике "Лиды авто"
    if message.chat.id != SUPPORT_GROUP_ID or message.message_thread_id != LEADS_TOPIC_ID:
        return
    
    def _count_lead_stats() -> tuple:
        from datetime import timedelta
        
        now = datetime.now(timezone.utc)
        day_ago = now - timedelta(days=1)
        week_ago = now - timedelta(days=7)
        month_ago = now - timedelta(days=30)
        
        lead_stats = []  # (name, total, day, week, month)
        
        for key, info in LEAD_TYPES.items():
            csv_path = info["csv"]
            rows = _read_csv(csv_path)
            total = len(rows) - 1  # Минус заголовок
            
            # Считаем добавленные за периоды
            day_count = 0
            week_count = 0
            month_count = 0
            
            for row in rows[1:]:
                if len(row) >= 4 and row[3]:  # Есть дата добавления
                    try:
                        # Формат: "YYYY.MM.DD HH:MM:SS"
                        added_date = datetime.strptime(row[3], "%Y.%m.%d %H:%M:%S").replace(tzinfo=timezone.utc)
                        if added_date >= day_ago:
                            day_count += 1
                        if added_date >= week_ago:
                            week_count += 1
                        if added_date >= month_ago:
                            month_count += 1
                    except ValueError:
                        pass
            
            lead_stats.append((info["name"], total, day_count, week_count, month_count))
        
        return lead_stats
    
    lead_stats = await asyncio.to_thread(_count_lead_stats)
    
    # Формируем статистику
    lines = ["📊 **Статистика лидов:**\n"]
    lines.append("```")
    lines.append(f"{'Тип':<25} {'Всего':>7} {'Сутки':>7} {'Неделя':>7} {'Месяц':>7}")
    lines.append("-" * 59)
    
    total_all = 0
    total_day = 0
    total_week = 0
    total_month = 0
    
    for name, total, day, week, month in lead_stats:
        # Обрезаем длинные названия
        short_name = name[:24] if len(name) > 24 else name
        lines.append(f"{short_name:<25} {total:>7} {day:>7} {week:>7} {month:>7}")
        total_all += total
        total_day += day
        total_week += week
        total_month += month
    
    lines.append("-" * 59)
    lines.append(f"{'ИТОГО':<25} {total_all:>7} {total_day:>7} {total_week:>7} {total_month:>7}")
    lines.append("```")
    
    await message.answer("\n".join(lines), parse_mode="Markdown")


# ============ РУЧНОЕ ДОБАВЛЕНИЕ ЛИДОВ ============

async def on_add_lead_start(message: Message, state: FSMContext) -> None:
    """Начало ручного добавления лида (только топик Лиды авто)."""
    # Только в топике "Лиды авто"
    if message.chat.id != SUPPORT_GROUP_ID or message.message_thread_id != LEADS_TOPIC_ID:
        return
    
    await state.set_state(ManualLeadStates.waiting_contact)
    await message.answer(
        "📝 Добавление лида вручную\n\n"
        "Отправьте контакт лида: @username, номер телефона или ссылку.",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="⬅️ Отмена")]],
            resize_keyboard=True,
        ),
    )


async def on_add_lead_contact(message: Message, state: FSMContext) -> None:
    """Получен контакт — запрашиваем категорию."""
    if not message.text or not message.text.strip():
        return
    
    # Сохраняем контакт в FSM
    contact = message.text.strip()
    await state.update_data(lead_contact=contact)
    await state.set_state(ManualLeadStates.waiting_category)
    
    await message.answer(
        f"Контакт: {contact}\n\n"
        "Выберите категорию для добавления лида:",
        reply_markup=get_lead_category_inline_keyboard(),
    )


# Маппинг callback_data -> тип лида
LEAD_CATEGORY_CALLBACK_MAP = {
    "lead_cat_telegram": "telegram",
    "lead_cat_whatsapp": "whatsapp",
    "lead_cat_max": "max",
    "lead_cat_viber": "viber",
    "lead_cat_instagram": "instagram",
    "lead_cat_vk": "vk",
    "lead_cat_ok": "ok",
    "lead_cat_email": "email",
    "lead_cat_avito": "avito",
    "lead_cat_yula": "yula",
    "lead_cat_kwork": "kwork",
    "lead_cat_self": "self",
}


async def on_add_lead_category_callback(callback: CallbackQuery, state: FSMContext, bot: Bot) -> None:
    """Выбрана категория через inline-кнопку — добавляем лид."""
    data_text = callback.data
    if not data_text or data_text == "lead_cat_cancel":
        await callback.answer("Отменено")
        await state.clear()
        await callback.message.edit_text("❌ Добавление лида отменено.")
        return
    
    lead_type = LEAD_CATEGORY_CALLBACK_MAP.get(data_text)
    if not lead_type:
        await callback.answer("Неверная категория")
        return
    
    data = await state.get_data()
    contact = data.get("lead_contact", "")
    
    if not contact:
        await callback.answer("Ошибка: контакт не найден.")
        await state.clear()
        return
    
    user = callback.from_user
    if not user:
        await state.clear()
        return
    
    await callback.answer()
    
    # Проверяем дубликат
    duplicate = check_lead_duplicate(contact)
    if duplicate:
        dup_type, dup_user_id, dup_username = duplicate
        await callback.message.edit_text(
            f"⚠️ Лид уже существует!\n\n"
            f"📋 Лид: {contact}\n"
            f"📦 Тип: {LEAD_TYPES[dup_type]['name']}\n"
            f"🆔 Добавлен пользователем: {dup_user_id} (@{dup_username})",
        )
        await state.clear()
        return
    
    # Добавляем лид: проверяем, из базы или самостоятельный
    in_base = determine_contact_type(contact, user.id) == lead_type
    source = "база" if in_base else f"самостоятельный {LEAD_TYPES[lead_type]['name'].lower()}"
    success = add_lead(contact, lead_type, user.id, user.username or "admin", source=source)
    
    if success:
        try:
            await callback.message.edit_text(
                f"✅ Лид добавлен!\n\n"
                f"📋 Контакт: {contact}\n"
                f"📦 Категория: {LEAD_TYPES[lead_type]['name']}",
            )
        except Exception:
            await callback.message.answer(
                f"✅ Лид добавлен!\n\n"
                f"📋 Контакт: {contact}\n"
                f"📦 Категория: {LEAD_TYPES[lead_type]['name']}",
            )
        
        # Уведомление в топик
        await bot.send_message(
            chat_id=SUPPORT_GROUP_ID,
            message_thread_id=LEADS_TOPIC_ID,
            text=(
                f"➕ Лид добавлен вручную\n\n"
                f"📋 Контакт: {contact}\n"
                f"📦 Категория: {LEAD_TYPES[lead_type]['name']}\n"
                f"👤 Добавил: {user.full_name} (@{user.username or 'нет'})"
            ),
        )
    else:
        await callback.message.edit_text("❌ Ошибка при добавлении лида.")
    
    await state.clear()


async def on_add_lead_cancel(message: Message, state: FSMContext) -> None:
    """Отмена добавления лида."""
    await state.clear()
    await message.answer(
        "Отмена добавления лида.",
        reply_markup=ReplyKeyboardRemove(remove_keyboard=True),
    )


# ============ УДАЛЕНИЕ ЛИДОВ ============

async def on_delete_lead_start(message: Message, state: FSMContext) -> None:
    """Начало удаления лида (только топик Лиды авто)."""
    # Только в топике "Лиды авто"
    if message.chat.id != SUPPORT_GROUP_ID or message.message_thread_id != LEADS_TOPIC_ID:
        return
    
    await state.set_state(DeleteLeadStates.waiting_contact)
    await message.answer(
        "🗑 Удаление лида\n\n"
        "Отправьте контакт лида для удаления: @username, номер телефона или ссылку.",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="⬅️ Отмена")]],
            resize_keyboard=True,
        ),
    )


async def on_delete_lead_contact(message: Message, state: FSMContext, bot: Bot) -> None:
    """Получен контакт — удаляем лид."""
    if not message.text or not message.text.strip():
        return
    
    contact = message.text.strip()
    user = message.from_user
    
    # Удаляем лид
    result = delete_lead(contact)
    
    if result:
        lead_type, found_value = result
        await message.answer(
            f"✅ Лид удалён!\n\n"
            f"📋 Контакт: {found_value}\n"
            f"📦 Категория: {LEAD_TYPES[lead_type]['name']}",
            reply_markup=ReplyKeyboardRemove(remove_keyboard=True),
        )
        
        # Уведомление в топик
        if user:
            await bot.send_message(
                chat_id=SUPPORT_GROUP_ID,
                message_thread_id=LEADS_TOPIC_ID,
                text=(
                    f"🗑 Лид удалён\n\n"
                    f"📋 Контакт: {found_value}\n"
                    f"📦 Категория: {LEAD_TYPES[lead_type]['name']}\n"
                    f"👤 Удалил: {user.full_name} (@{user.username or 'нет'})"
                ),
            )
    else:
        await message.answer(
            f"❌ Лид не найден: {contact}\n\n"
            "Проверьте правильность написания контакта.",
            reply_markup=ReplyKeyboardRemove(remove_keyboard=True),
        )
    
    await state.clear()


async def on_delete_lead_cancel(message: Message, state: FSMContext) -> None:
    """Отмена удаления лида."""
    await state.clear()
    await message.answer(
        "Отмена удаления лида.",
        reply_markup=ReplyKeyboardRemove(remove_keyboard=True),
    )


async def on_get_base(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if not user or not is_user_approved(user.id):
        await message.answer("❌ У вас нет доступа к этой функции.")
        return
    
    await state.clear()
    await message.answer(
        "Выбери, какую базу хочешь получить:",
        reply_markup=get_user_choice_keyboard(),
    )


async def on_back(message: Message, state: FSMContext, bot: Bot) -> None:
    await state.clear()
    await on_start(message, state, bot)


async def on_user_base_choice(message: Message, state: FSMContext, bot: Bot) -> None:
    """Обработка выбора типа базы пользователем."""
    user = message.from_user
    if not user:
        await message.answer("Не удалось определить пользователя.")
        return
    
    if not is_user_approved(user.id):
        await message.answer("❌ У вас нет доступа к этой функции.")
        return

    text = message.text
    if text not in USER_BUTTON_MAP:
        return

    base_key = USER_BUTTON_MAP[text]
    info = BASE_TYPES[base_key]
    user_id = user.id
    username = user.username or ""

    # Защита от повторных нажатий
    if user_id in processing_users[base_key]:
        await message.answer("Подожди, твой запрос уже обрабатывается...")
        return

    processing_users[base_key].add(user_id)
    try:
        values, reason = await allocate_for_user(base_key, user_id, username)
    except Exception:
        await message.answer("Произошла ошибка. Попробуй позже.")
        return
    finally:
        processing_users[base_key].discard(user_id)

    if reason == "already_got":
        await message.answer(
            f"Ты уже получил контакты из «{info['name']}».\n"
            f"Лимит: {info['limit']} контактов. Обратитесь к менеджеру."
        )
        return

    if reason == "not_enough":
        await message.answer(
            f"❌ К сожалению, контакты «{info['name']}» на данный момент отсутствуют.\n"
            "Обратитесь к менеджеру."
        )
        # Уведомление в General (группу поддержки)
        try:
            await bot.send_message(
                chat_id=SUPPORT_GROUP_ID,
                text=(
                    f"⚠️ ВНИМАНИЕ: Контакты закончились!\n\n"
                    f"📦 Тип: {info['name']}\n"
                    f"👤 Пользователь: {user.full_name} (@{user.username or 'нет'})\n"
                    f"🆔 ID: {user_id}\n\n"
                    f"Необходимо загрузить новые контакты!"
                ),
            )
        except Exception:
            pass  # Не прерываем, если не удалось отправить
        return

    if not values:
        await message.answer("Произошла ошибка при выдаче контактов.")
        return

    # Отправляем контакты сообщением (не файлом)
    contacts_text = "\n".join(values)
    
    # Telegram ограничивает длину сообщения 4096 символами
    if len(contacts_text) <= 4000:
        await message.answer(
            f"✅ Выдано из «{info['name']}»: {len(values)} контактов\n\n"
            f"{contacts_text}"
        )
    else:
        # Если слишком длинный, разбиваем на части
        await message.answer(f"✅ Выдано из «{info['name']}»: {len(values)} контактов")
        
        # Отправляем по частям (пауза — защита от Flood control)
        chunk = ""
        for val in values:
            if len(chunk) + len(val) + 1 > 4000:
                await message.answer(chunk)
                await asyncio.sleep(FLOOD_DELAY)
                chunk = val
            else:
                chunk = chunk + "\n" + val if chunk else val
        if chunk:
            await message.answer(chunk)
    
    # Подсказка и переход в главное меню
    await message.answer(
        "Когда выполните работу, нажмите «Отчёт по лидам» и пришлите скриншот + ссылку на лида.",
        reply_markup=get_main_keyboard(),
    )
    
    # Проверяем, осталось ли меньше 5% свободных контактов
    try:
        csv_path = info["csv"]
        rows = _read_csv(csv_path)
        total = len(rows) - 1  # Минус заголовок
        free = sum(1 for r in rows[1:] if len(r) < 2 or not r[1])
        
        if total > 0:
            percent = (free / total) * 100
            if percent < 5:
                await bot.send_message(
                    chat_id=SUPPORT_GROUP_ID,
                    text=(
                        f"⚠️ ВНИМАНИЕ: Контакты заканчиваются!\n\n"
                        f"📦 Тип: {info['name']}\n"
                        f"📊 Осталось: {free} из {total} ({percent:.1f}%)\n\n"
                        f"Необходимо загрузить новые контакты!"
                    ),
                )
    except Exception:
        pass


# ============ АДМИН-ХЕНДЛЕРЫ ============

async def on_admin_exit(message: Message, state: FSMContext) -> None:
    """Выход из админки."""
    await state.clear()
    await message.answer("Вышел из админки.", reply_markup=get_main_keyboard())


async def on_admin_download(message: Message, state: FSMContext) -> None:
    """Выкачивание всей базы в Excel."""
    user = message.from_user
    if not user or not is_admin(user.id):
        return

    try:
        buffer, filename = _create_full_excel()
    except Exception:
        await message.answer("Ошибка при создании файла.")
        return

    document = BufferedInputFile(buffer.read(), filename=filename)
    await message.answer_document(
        document=document,
        caption="📤 Полная база данных (все типы) в одном Excel-файле.",
    )


async def on_admin_upload_start(message: Message, state: FSMContext) -> None:
    """Начало загрузки базы — показываем выбор типа."""
    user = message.from_user
    if not user or not is_admin(user.id):
        return

    await state.set_state(AdminStates.waiting_upload_choice)
    await message.answer(
        "Выбери, какую базу хочешь загрузить:",
        reply_markup=get_admin_upload_choice_keyboard(),
    )


async def on_admin_upload_cancel(message: Message, state: FSMContext) -> None:
    """Отмена загрузки."""
    await state.clear()
    await message.answer("Загрузка отменена.", reply_markup=get_admin_keyboard())


async def on_admin_upload_choice(message: Message, state: FSMContext) -> None:
    """Обработка выбора типа базы для загрузки."""
    user = message.from_user
    if not user or not is_admin(user.id):
        return

    text = message.text
    if text not in ADMIN_UPLOAD_MAP:
        return

    base_key = ADMIN_UPLOAD_MAP[text]
    await state.update_data(upload_type=base_key)
    await state.set_state(AdminStates.waiting_file)

    if base_key == "all":
        await message.answer(
            "📚 Режим загрузки ВСЕХ листов.\n\n"
            "Отправь Excel-файл (.xlsx) с листами:\n"
            "Тг, Вотсап, Макс, Вайбер, Инст, ВК, Ок, Почта\n\n"
            "Данные будут добавлены в соответствующие базы без дубликатов.",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[[KeyboardButton(text="⬅️ Отмена")]],
                resize_keyboard=True,
            ),
        )
    else:
        info = BASE_TYPES[base_key]
        await message.answer(
            f"📥 Загрузка в базу «{info['name']}»\n\n"
            "Отправь Excel-файл (.xlsx).\n"
            "Данные будут взяты из первого столбца первого листа.\n"
            "Дубликаты автоматически пропускаются.",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[[KeyboardButton(text="⬅️ Отмена")]],
                resize_keyboard=True,
            ),
        )


async def on_admin_file_received(message: Message, state: FSMContext, bot: Bot) -> None:
    """Обработка полученного файла от админа."""
    user = message.from_user
    if not user or not is_admin(user.id):
        return

    if not message.document:
        await message.answer("Пожалуйста, отправь файл Excel (.xlsx).")
        return

    filename = message.document.file_name or ""
    if not filename.lower().endswith(".xlsx"):
        await message.answer("Нужен файл в формате .xlsx (Excel).")
        return

    data = await state.get_data()
    upload_type = data.get("upload_type")
    if not upload_type:
        await message.answer("Ошибка состояния. Начни заново.")
        await state.clear()
        return

    # Скачиваем файл
    await message.answer("⏳ Обрабатываю файл... (большой файл может занять несколько минут)")

    try:
        file_io = await bot.download(message.document)
        if not file_io:
            await message.answer("Не удалось скачать файл.")
            return

        file_bytes = file_io.read()
        results, err = await asyncio.to_thread(_process_excel_upload_sync, file_bytes, upload_type)

        if err:
            await message.answer(f"❌ Ошибка при обработке файла: {err}")
            await state.clear()
            return

        await state.clear()
        await message.answer(
            "📊 Результат загрузки:\n\n" + "\n".join(results),
            reply_markup=get_admin_keyboard(),
        )

    except Exception as e:
        await message.answer(f"❌ Ошибка при обработке файла: {e}")
        await state.clear()


# ============ АДМИН: УДАЛЕНИЕ БАЗЫ ============

def clear_all_databases() -> int:
    """Очищает все CSV-файлы (удаляет данные, оставляет заголовки). Возвращает кол-во очищенных."""
    count = 0
    for key, info in BASE_TYPES.items():
        csv_path = info["csv"]
        if os.path.exists(csv_path):
            with open(csv_path, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Value", "ID", "Username", "Date"])
            count += 1
    return count


async def on_admin_delete_start(message: Message, state: FSMContext) -> None:
    """Начало удаления базы — показываем подтверждение."""
    user = message.from_user
    if not user or not is_admin(user.id):
        return

    await state.set_state(AdminStates.waiting_delete_confirm)
    await message.answer(
        "⚠️ ВНИМАНИЕ!\n\n"
        "Ты собираешься удалить ВСЮ базу данных.\n"
        "Это действие НЕОБРАТИМО!\n\n"
        "Все записи во всех типах баз будут удалены.\n\n"
        "Ты уверен?",
        reply_markup=get_delete_confirm_keyboard(),
    )


async def on_admin_delete_confirm(message: Message, state: FSMContext) -> None:
    """Подтверждение удаления."""
    user = message.from_user
    if not user or not is_admin(user.id):
        return

    async with csv_lock:
        count = await asyncio.to_thread(clear_all_databases)

    await state.clear()
    await message.answer(
        f"🗑 База данных полностью очищена.\n"
        f"Очищено файлов: {count}",
        reply_markup=get_admin_keyboard(),
    )


async def on_admin_delete_cancel(message: Message, state: FSMContext) -> None:
    """Отмена удаления."""
    await state.clear()
    await message.answer(
        "❌ Удаление отменено.",
        reply_markup=get_admin_keyboard(),
    )


# ============ СТАТИСТИКА ЛИДОВ ПОЛЬЗОВАТЕЛЯ ============

def _count_user_leads(user_id: int) -> tuple[int, int, int]:
    """Возвращает (лидов за сегодня, за вчера, за всё время) для пользователя."""
    today = get_current_lead_day()
    yesterday = get_yesterday_lead_day()
    count_today = 0
    count_yesterday = 0
    count_all = 0
    user_id_str = str(user_id)
    for key, info in LEAD_TYPES.items():
        csv_path = info["csv"]
        if os.path.exists(csv_path):
            rows = _read_csv(csv_path)
            for row in rows[1:]:
                if len(row) >= 2 and str(row[1]).strip() == user_id_str:
                    count_all += 1
        for date, cnt in [(today, "today"), (yesterday, "yesterday")]:
            daily_path = _get_daily_leads_path(key, date)
            if daily_path and os.path.exists(daily_path):
                rows = _read_csv(daily_path)
                for row in rows[1:]:
                    if len(row) >= 2 and str(row[1]).strip() == user_id_str:
                        if cnt == "today":
                            count_today += 1
                        else:
                            count_yesterday += 1
    return count_today, count_yesterday, count_all


async def on_user_lead_stats(message: Message) -> None:
    """Личная статистика лидов пользователя: за сегодня и за всё время."""
    user = message.from_user
    if not user or not is_user_approved(user.id):
        await message.answer("❌ У вас нет доступа к этой функции.")
        return

    count_today, count_yesterday, count_all = await asyncio.to_thread(_count_user_leads, user.id)
    await message.answer(
        f"📊 Ваша статистика лидов\n\n"
        f"📈 За весь период: {count_all}\n"
        f"📅 За вчерашний день: {count_yesterday}\n"
        f"📅 За сегодняшний день: {count_today}\n\n"
        "⏰ День обновляется с 20:00. Лиды после 20:00 будут улетать на следующий день.\n\n"
        "💡 Лид не засчитался? Отправьте его через «Отчёт по лидам»: "
        "скриншот переписки + в подписи контакт (@username, ссылка или телефон). "
        "Не забудьте выбрать нужную категорию для каждого лида.\n\n"
        "Возникли вопросы? Пишите в поддержку."
    )


# ============ CHECK_LEADS — ЛИДЫ ПОЛЬЗОВАТЕЛЯ ============

async def on_check_leads(message: Message, bot: Bot) -> None:
    """Команда /check_leads — статистика и Excel лидов пользователя (в чате с ним)."""
    if message.chat.id != SUPPORT_GROUP_ID:
        return

    topic_id = message.message_thread_id
    if not topic_id:
        await message.answer("❌ Используйте команду /check_leads внутри чата с пользователем (в топике).")
        return

    user_id = get_user_by_topic(topic_id)
    if not user_id:
        await message.answer("❌ Не удалось определить пользователя по этому чату.")
        return

    try:
        user_chat = await bot.get_chat(user_id)
        user_name = user_chat.full_name or f"User_{user_id}"
        username = user_chat.username or "нет"
    except Exception:
        user_name = f"User_{user_id}"
        username = "нет"

    await message.answer("⏳ Собираю данные...")

    count_today, count_yesterday, count_all = await asyncio.to_thread(_count_user_leads, user_id)
    today = get_current_lead_day()
    yesterday = get_yesterday_lead_day()

    text = (
        f"📊 Лиды пользователя {user_name}\n"
        f"🆔 ID: {user_id}\n"
        f"📱 @{username}\n\n"
        f"📈 За весь период: {count_all}\n"
        f"📅 За вчерашний день: {count_yesterday}\n"
        f"📅 За сегодняшний день: {count_today}"
    )
    await message.answer(text)
    await asyncio.sleep(FLOOD_DELAY)

    try:
        buf_all, name_all = await asyncio.to_thread(_create_user_leads_excel, user_id, True)
        buf_yesterday, name_yesterday = await asyncio.to_thread(_create_user_leads_excel, user_id, False, yesterday)
        buf_today, name_today = await asyncio.to_thread(_create_user_leads_excel, user_id, False, today)
        doc_all = BufferedInputFile(buf_all.read(), filename=name_all)
        doc_yesterday = BufferedInputFile(buf_yesterday.read(), filename=name_yesterday)
        doc_today = BufferedInputFile(buf_today.read(), filename=name_today)
        await message.answer_document(doc_all, caption="📤 Лиды за весь период")
        await asyncio.sleep(FLOOD_DELAY)
        await message.answer_document(doc_yesterday, caption=f"📤 Лиды за вчерашний день ({yesterday})")
        await asyncio.sleep(FLOOD_DELAY)
        await message.answer_document(doc_today, caption=f"📤 Лиды за сегодняшний день ({today})")
    except Exception as e:
        await message.answer(f"❌ Ошибка при создании файлов: {e}")


# ============ ПОДДЕРЖКА: ХЕНДЛЕРЫ ============

async def on_request_new_contacts(message: Message, bot: Bot) -> None:
    """Пользователь нажал 'Получить новые контакты' — уведомление в админский чат."""
    user = message.from_user
    if not user or not is_user_approved(user.id):
        await message.answer("❌ У вас нет доступа к этой функции.")
        return

    user_link = f'<a href="tg://user?id={user.id}">{user.full_name}</a>'
    username = user.username or "нет"
    topics = load_support_topics()
    topic_id = topics.get(user.id)
    topic_link = ""
    if topic_id:
        chat_id_short = str(SUPPORT_GROUP_ID).replace("-100", "")
        topic_link = f'\n\n📨 <a href="https://t.me/c/{chat_id_short}/{topic_id}">Чат с пользователем</a>'

    await bot.send_message(
        chat_id=SUPPORT_GROUP_ID,
        message_thread_id=LEADS_TOPIC_ID,
        text=(
            f"🆕 Запрос на новые контакты\n\n"
            f"👤 {user_link}\n"
            f"🆔 ID: {user.id}\n"
            f"📱 @{username}"
            f"{topic_link}"
        ),
        parse_mode="HTML",
    )
    await message.answer(
        "✅ Ваш запрос отправлен! Ожидайте, с вами свяжутся для добавления новых контактов."
    )


async def on_support_info(message: Message, state: FSMContext) -> None:
    """Пользователь нажал 'Написать в поддержку' — переводим в режим поддержки."""
    await state.set_state(SupportStates.active)
    await message.answer(
        "💬 Режим поддержки\n\n"
        "Напишите сообщение — оно уйдёт менеджеру, и он ответит вам здесь.\n\n"
        "Нажмите «Назад», когда закончите общение.",
        reply_markup=get_support_keyboard(),
    )


# ============ ОТЧЁТЫ ============

async def on_report_start(message: Message, state: FSMContext) -> None:
    """Пользователь нажал 'Отчёт по лидам' — начинаем сбор файлов."""
    user = message.from_user
    if not user or not is_user_approved(user.id):
        await message.answer("❌ У вас нет доступа к этой функции.")
        return
    
    await state.set_state(ReportStates.waiting_report)
    await state.update_data(report_items=[], report_contact_categories={})
    await message.answer(
        "📋 Отчёт по лидам\n\n"
        "📸 Максимум 5 лидов. Один лид = 1 скриншот + 1 контакт в подписи.\n\n"
        "Формат: 1 скриншот + подпись к нему (@username, ссылка или телефон).\n\n"
        "Не пишите «сам», «самостоятельно» — не нужно.\n\n"
        "Не загружайте в одно сообщение несколько скриншотов. "
        "Не заливайте несколько тегов в один лид — вам не засчитает. "
        "Строго 1 скриншот + 1 подпись (тег, ссылка)\n\n"
        "🔴 Только скриншоты и контакты, без лишнего текста.\n"
        "💬 Вопросы — в поддержку или группу «Работа».\n\n"
        "Не нажимайте кнопку «Отправить отчёт» до того, как загрузите лиды\n\n"
        "✅ Всё загрузили? Жми «Отправить отчёт» 👇",
        reply_markup=get_report_keyboard(),
    )


async def _maybe_show_category_for_item(
    state: FSMContext,
    message: Message,
    bot: Bot,
    item: dict,
    *,
    user_id: int,
    username: str,
    user_name: str,
) -> None:
    """Если в элементе есть контакт — сразу показать выбор категории."""
    data = await state.get_data()
    items = data.get("report_items", [])
    count = len(items)

    source_text = item.get("content", "") or item.get("caption", "") or ""
    if not source_text:
        if count >= REPORT_LEADS_LIMIT:
            text = f"📋 Достигнут лимит {REPORT_LEADS_LIMIT} лидов в отчёте.\n\nНажмите «Отправить отчёт» для отправки."
        else:
            text = f"✅ Добавлено. В отчёте {count} из {REPORT_LEADS_LIMIT} лидов. Можете загрузить следующий лид или нажать «Отправить отчёт»."
        await message.answer(text, reply_markup=get_report_keyboard())
        return

    contacts = extract_contacts_from_text(source_text)
    if not contacts and source_text.strip():
        contacts = [source_text.strip()]

    seen = {}
    unique = []
    for c in contacts:
        norm = normalize_contact(c)
        if norm and norm not in seen:
            seen[norm] = c
            unique.append(c)

    pending = []
    dup_msg = []
    for contact in unique:
        if check_lead_duplicate(contact):
            dup_msg.append(contact)
        else:
            pending.append(contact)

    if dup_msg and not pending:
        if count >= REPORT_LEADS_LIMIT:
            text = f"⚠️ Эти контакты уже в базе: {', '.join(dup_msg)}\n\n📋 Достигнут лимит {REPORT_LEADS_LIMIT} лидов. Нажмите «Отправить отчёт» для отправки."
        else:
            text = f"⚠️ Эти контакты уже в базе: {', '.join(dup_msg)}\n\nВ отчёте {count} из {REPORT_LEADS_LIMIT} лидов. Можете загрузить следующий лид или нажать «Отправить отчёт»."
        await message.answer(text, reply_markup=get_report_keyboard())
        return

    if pending:
        # Показываем выбор категории — лиды добавляются только при «Отправить отчёт»
        topics = load_support_topics()
        topic_id = topics.get(user_id)
        target_topic = topic_id if topic_id else REPORTS_TOPIC_ID
        await state.update_data(
            report_pending_contacts=pending,
            report_idx=0,
            report_user_id=user_id,
            report_username=username,
            report_user_name=user_name,
            report_topic_id=topic_id,
            report_target_topic=target_topic,
            report_message_id=None,
        )
        await state.set_state(ReportStates.waiting_category)
        contact = pending[0]
        total = len(pending)
        dup_note = f"⚠️ Уже в базе: {', '.join(dup_msg)}\n\n" if dup_msg else ""
        await message.answer(
            f"{dup_note}📋 Контакт 1 из {total}: {contact}\n\n"
            "Выберите категорию для добавления лида:",
            reply_markup=get_report_category_inline_keyboard(0),
        )
    elif not dup_msg:
        if count >= REPORT_LEADS_LIMIT:
            text = f"📋 Достигнут лимит {REPORT_LEADS_LIMIT} лидов в отчёте.\n\nНажмите «Отправить отчёт» для отправки."
        else:
            text = f"✅ Добавлено. В отчёте {count} из {REPORT_LEADS_LIMIT} лидов. Можете загрузить следующий лид или нажать «Отправить отчёт»."
        await message.answer(text, reply_markup=get_report_keyboard())


async def on_report_file(
    message: Message, state: FSMContext, bot: Bot,
) -> None:
    """Приём фото/документов для отчёта."""
    user = message.from_user
    if not user:
        return
    
    data = await state.get_data()
    items = data.get("report_items", [])
    
    file_id = None
    file_type = None
    caption = _extract_text_with_urls(message) or (message.caption or "").strip()
    if message.photo:
        file_id = message.photo[-1].file_id
        file_type = "photo"
    elif message.document:
        file_id = message.document.file_id
        file_type = "document"
    elif message.video:
        file_id = message.video.file_id
        file_type = "video"
    
    if file_id and file_type:
        if len(items) >= REPORT_LEADS_LIMIT:
            await message.answer(
                f"📋 Достигнут лимит {REPORT_LEADS_LIMIT} лидов в отчёте.\n\n"
                "Нажмите «Отправить отчёт» для отправки.",
                reply_markup=get_report_keyboard(),
            )
            return
        items.append({"type": file_type, "file_id": file_id, "caption": caption})
        await state.update_data(report_items=items)
        await _maybe_show_category_for_item(
            state, message, bot, items[-1],
            user_id=user.id,
            username=user.username or "",
            user_name=user.full_name or "",
        )


async def on_report_submit(
    message: Message, state: FSMContext, bot: Bot,
) -> None:
    """Пользователь нажал 'Отправить отчёт'."""
    user = message.from_user
    if not user:
        return

    data = await state.get_data()
    items = data.get("report_items", [])

    if not items:
        await message.answer(
            "Вы не добавили ни одного лида. Загрузите скриншоты с контактами, затем нажмите «Отправить отчёт».",
            reply_markup=get_report_keyboard(),
        )
        return

    # Очищаем состояние — защита от двойного нажатия
    await state.clear()
    
    user_id = user.id
    topics = load_support_topics()
    topic_id = topics.get(user_id)
    
    # Файлы идут в обычный чат поддержки пользователя; если его нет — в топик «Отчёты»
    target_topic = topic_id if topic_id else REPORTS_TOPIC_ID
    
    try:
        # Формируем ссылку на пользователя
        user_link = f'<a href="tg://user?id={user_id}">{user.full_name}</a>'
        
        # Формируем текст уведомления
        notification_text = f"📋 Новый отчёт по лидам!\n\n"
        notification_text += f"👤 {user_link}\n"
        notification_text += f"🆔 ID: {user_id}\n"
        if user.username:
            notification_text += f"📱 @{user.username}\n"
        
        # Добавляем ссылку на топик пользователя, если он существует
        if topic_id:
            # Для ссылки на топик нужен chat_id без префикса -100
            chat_id_short = str(SUPPORT_GROUP_ID).replace("-100", "")
            topic_link = f"https://t.me/c/{chat_id_short}/{topic_id}"
            notification_text += f'\n📨 <a href="{topic_link}">Перейти в чат с пользователем</a>'
        else:
            notification_text += "\n📨 Файлы ниже ⬇️"
        
        # Уведомление в топик «Отчёты»
        await bot.send_message(
            chat_id=SUPPORT_GROUP_ID,
            message_thread_id=REPORTS_TOPIC_ID,
            text=notification_text,
            parse_mode="HTML",
        )
        
        # Файлы — в обычный чат поддержки пользователя
        report_msg = await bot.send_message(
            chat_id=SUPPORT_GROUP_ID,
            message_thread_id=target_topic,
            text=f"📋 Отчёт от {user.full_name} (@{user.username or 'нет'}):",
        )
        report_message_id = report_msg.message_id
        
        for item in items:
            ft = item["type"]
            if ft == "text":
                await bot.send_message(
                    chat_id=SUPPORT_GROUP_ID,
                    message_thread_id=target_topic,
                    text=f"💬 {item['content']}",
                )
            elif ft == "photo":
                cap = item.get("caption") or None
                await bot.send_photo(
                    chat_id=SUPPORT_GROUP_ID,
                    message_thread_id=target_topic,
                    photo=item["file_id"],
                    caption=cap,
                )
            elif ft == "document":
                cap = item.get("caption") or None
                await bot.send_document(
                    chat_id=SUPPORT_GROUP_ID,
                    message_thread_id=target_topic,
                    document=item["file_id"],
                    caption=cap,
                )
            elif ft == "video":
                cap = item.get("caption") or None
                await bot.send_video(
                    chat_id=SUPPORT_GROUP_ID,
                    message_thread_id=target_topic,
                    video=item["file_id"],
                    caption=cap,
                )
            await asyncio.sleep(FLOOD_DELAY)

        # Извлекаем контакты из отчёта и добавляем в лиды (как в чате поддержки)
        ensure_leads_csv_exists()
        all_contacts_with_source = []
        for item in items:
            source_text = item.get("content", "") or item.get("caption", "") or ""
            if source_text:
                contacts = extract_contacts_from_text(source_text)
                if not contacts:
                    contacts = [source_text.strip()]
                for c in contacts:
                    if c.strip():
                        all_contacts_with_source.append((c.strip(), source_text))

        seen = {}
        unique_contacts = []
        for contact, _ in all_contacts_with_source:
            norm = normalize_contact(contact)
            if norm and norm not in seen:
                seen[norm] = contact
                unique_contacts.append(contact)

        content_lower = " ".join(
            (item.get("content", "") or item.get("caption", "") or "").lower()
            for item in items
        )
        content_full = " ".join(
            (item.get("content", "") or item.get("caption", "") or "")
            for item in items
        )
        tg_hint = " тг" in content_lower or "тг " in content_lower or " tg" in content_lower or "tg " in content_lower
        kwork_hint = bool(KWORK_LEAD_KEYWORDS.search(content_full))
        username_str = user.username or ""
        cat_map = data.get("report_contact_categories", {})

        duplicates_in_report = []
        for contact in unique_contacts:
            duplicate = check_lead_duplicate(contact)
            if duplicate:
                dup_type, dup_user_id, dup_username = duplicate
                duplicates_in_report.append((contact, dup_type, dup_user_id, dup_username))
                continue
            stored_cat = cat_map.get(normalize_contact(contact))
            if stored_cat == "skip":
                continue
            if stored_cat and stored_cat in LEAD_TYPES:
                contact_type = stored_cat
            else:
                contact_type = determine_contact_type(contact, user_id)
                if not contact_type or contact_type not in LEAD_TYPES:
                    if tg_hint:
                        contact_type = "telegram"
                    elif kwork_hint:
                        contact_type = "kwork"
                    else:
                        contact_type = "self"
            in_base = bool(contact_type) and determine_contact_type(contact, user_id) == contact_type
            src_name = LEAD_TYPES[contact_type]["name"].lower()
            source = "база" if in_base else ("самостоятельный" if contact_type == "self" else f"самостоятельный {src_name}")
            chat_id_short = str(SUPPORT_GROUP_ID).replace("-100", "")
            msg_link = f"https://t.me/c/{chat_id_short}/{target_topic}/{report_message_id}" if target_topic else ""
            try:
                if add_lead(contact, contact_type, user_id, username_str, source=source, message_link=msg_link):
                    user_link = f'<a href="tg://user?id={user_id}">{user.full_name}</a>'
                    report_link = f'\n\n📨 <a href="{msg_link}">Открыть отчёт</a>' if msg_link else ""
                    await bot.send_message(
                        chat_id=SUPPORT_GROUP_ID,
                        message_thread_id=LEADS_TOPIC_ID,
                        text=(
                            f"✅ Лид из отчёта\n\n"
                            f"📋 Контакт: {contact}\n"
                            f"📦 Категория: {LEAD_TYPES[contact_type]['name']}\n"
                            f"👤 От: {user_link}\n"
                            f"🆔 ID: {user_id}\n"
                            f"📱 @{username_str or 'нет'}"
                            f"{report_link}"
                        ),
                        parse_mode="HTML",
                    )
                    await asyncio.sleep(FLOOD_DELAY)
            except Exception as e:
                print(f"Ошибка добавления лида {contact}: {e}")

        if duplicates_in_report:
            dup_text = "\n".join(
                f"• {c} (в базе: {LEAD_TYPES.get(t, {}).get('name', t)}, от {uid} @{un})"
                for c, t, uid, un in duplicates_in_report
            )
            await bot.send_message(
                chat_id=SUPPORT_GROUP_ID,
                message_thread_id=LEADS_TOPIC_ID,
                text=(
                    f"⚠️ Дубликаты в отчёте (не добавлены)\n\n"
                    f"👤 От: {user.full_name} (@{username_str or 'нет'})\n"
                    f"🆔 ID: {user_id}\n\n"
                    f"{dup_text}"
                ),
            )
            await asyncio.sleep(FLOOD_DELAY)
            dup_list = ", ".join(c for c, *_ in duplicates_in_report)
            await message.answer(
                f"✅ Отчёт отправлен!\n\n"
                f"⚠️ Не добавлены (уже в базе): {dup_list}\n\n"
                "💡 Как правильно: скриншот + в подписи контакт (@username, ссылка или телефон). "
                "Выберите категорию для каждого лида.\n\n"
                "Не получилось? Пишите в поддержку.",
                reply_markup=get_main_keyboard(),
            )
        else:
            await message.answer(
                "✅ Отчёт отправлен!",
                reply_markup=get_main_keyboard(),
            )
    except Exception as e:
        await state.clear()
        await message.answer(
            f"❌ Ошибка при отправке: {e}",
            reply_markup=get_main_keyboard(),
        )


async def on_report_category_callback(callback: CallbackQuery, state: FSMContext, bot: Bot) -> None:
    """Пользователь выбрал категорию для лида из отчёта."""
    data_text = callback.data
    if not data_text or not data_text.startswith("report_cat_"):
        await callback.answer()
        return
    
    # report_cat_{idx}_{category}
    parts = data_text.split("_")
    if len(parts) < 4:
        await callback.answer()
        return
    idx = int(parts[2])
    category = "_".join(parts[3:])  # на случай "lead_cat_0_self"
    
    data = await state.get_data()
    pending = data.get("report_pending_contacts", [])
    if idx >= len(pending):
        await callback.answer("Контакты уже обработаны.")
        await state.clear()
        return
    
    contact = pending[idx]
    user_id = data.get("report_user_id")
    username = data.get("report_username", "")
    user_name = data.get("report_user_name", "")
    topic_id = data.get("report_topic_id")
    target_topic = data.get("report_target_topic")
    report_message_id = data.get("report_message_id")
    
    await callback.answer()
    
    if category == "cancel":
        await state.clear()
        await callback.message.edit_text("❌ Отчёт отменён.")
        await callback.message.answer("Отчёт отменён.", reply_markup=get_main_keyboard())
        return
    
    # Сохраняем выбор категории — лиды добавляются только при «Отправить отчёт»
    cat_map = data.get("report_contact_categories", {})
    if category == "skip":
        status = "⏭ Пропущено"
        cat_map[normalize_contact(contact)] = "skip"
    elif category in LEAD_TYPES:
        status = f"✅ Выбрано: {LEAD_TYPES[category]['name']}"
        cat_map[normalize_contact(contact)] = category
    else:
        status = "⏭ Пропущено"
        cat_map[normalize_contact(contact)] = "skip"
    await state.update_data(report_contact_categories=cat_map)
    
    try:
        await callback.message.edit_text(f"{status}\n\n📋 Контакт: {contact}")
    except Exception:
        pass
    
    # Следующий контакт или завершение
    next_idx = idx + 1
    if next_idx < len(pending):
        await state.update_data(report_idx=next_idx)
        next_contact = pending[next_idx]
        total = len(pending)
        await callback.message.answer(
            f"📋 Контакт {next_idx + 1} из {total}: {next_contact}\n\n"
            "Выберите категорию для добавления лида:",
            reply_markup=get_report_category_inline_keyboard(next_idx),
        )
    else:
        # Все контакты из этого лида обработаны — возвращаемся к сбору
        await state.set_state(ReportStates.waiting_report)
        await state.update_data(
            report_pending_contacts=[],
            report_idx=0,
        )
        items = data.get("report_items", [])
        count = len(items)
        if count >= REPORT_LEADS_LIMIT:
            text = (
                f"📋 Достигнут лимит {REPORT_LEADS_LIMIT} лидов в отчёте.\n\n"
                "Нажмите «Отправить отчёт» для отправки.\n\n"
                "👇 Кнопки «Отправить отчёт» и «Отмена» — ниже"
            )
        else:
            text = (
                f"✅ Лид сохранён. В отчёте {count} из {REPORT_LEADS_LIMIT} лидов.\n\n"
                "Можете прислать ещё лид (скриншот + тег) или отправить отчёт.\n\n"
                "👇 Кнопки «Отправить отчёт» и «Отмена» — ниже"
            )
        await callback.message.answer(text, reply_markup=get_report_keyboard())


async def on_report_waiting_category_remind(message: Message, state: FSMContext) -> None:
    """В режиме выбора категории — сохраняем сообщение в отчёт и напоминаем выбрать категорию."""
    # Сохраняем контент в report_items, чтобы лид не потерялся
    content = _extract_text_with_urls(message)
    if message.photo or message.document or message.video:
        file_id = None
        file_type = None
        caption = content or (message.caption or "").strip()
        if message.photo:
            file_id = message.photo[-1].file_id
            file_type = "photo"
        elif message.document:
            file_id = message.document.file_id
            file_type = "document"
        elif message.video:
            file_id = message.video.file_id
            file_type = "video"
        if file_id and file_type:
            data = await state.get_data()
            items = data.get("report_items", [])
            if len(items) < REPORT_LEADS_LIMIT:
                items.append({"type": file_type, "file_id": file_id, "caption": caption})
                await state.update_data(report_items=items)
            else:
                await message.answer(
                    f"📋 Достигнут лимит {REPORT_LEADS_LIMIT} лидов. Нажмите «Отправить отчёт».",
                    reply_markup=get_report_keyboard(),
                )
                return
    elif content:
        data = await state.get_data()
        items = data.get("report_items", [])
        if len(items) < REPORT_LEADS_LIMIT:
            items.append({"type": "text", "content": content})
            await state.update_data(report_items=items)
        else:
            await message.answer(
                f"📋 Достигнут лимит {REPORT_LEADS_LIMIT} лидов. Нажмите «Отправить отчёт».",
                reply_markup=get_report_keyboard(),
            )
            return

    await message.answer(
        "👆 Сначала выберите категорию выше, затем можно загрузить следующий.",
        reply_markup=get_report_keyboard(),
    )


async def on_report_cancel(message: Message, state: FSMContext) -> None:
    """Отмена сдачи отчёта."""
    await state.clear()
    await message.answer("Отмена.", reply_markup=get_main_keyboard())


def _extract_text_with_urls(message: Message) -> str:
    """Извлекает весь текст + URL из entities (на случай скрытых/форматированных ссылок)."""
    text = (message.text or message.caption or "").strip()
    # Добавляем URL из text_link entities (если ссылка под другим текстом)
    urls = []
    for entity in (message.entities or message.caption_entities or []):
        if hasattr(entity, "url") and entity.url:
            urls.append(entity.url)
    if urls:
        text = (text + "\n" + "\n".join(urls)).strip()
    return text


async def on_report_other(message: Message, state: FSMContext, bot: Bot) -> None:
    """Текстовые сообщения в режиме отчёта — добавляем в отчёт."""
    content = _extract_text_with_urls(message)
    if not content:
        return
    
    user = message.from_user
    if not user:
        return
    
    data = await state.get_data()
    items = data.get("report_items", [])
    if len(items) >= REPORT_LEADS_LIMIT:
        await message.answer(
            f"📋 Достигнут лимит {REPORT_LEADS_LIMIT} лидов в отчёте.\n\n"
            "Нажмите «Отправить отчёт» для отправки.",
            reply_markup=get_report_keyboard(),
        )
        return
    items.append({"type": "text", "content": content})
    await state.update_data(report_items=items)
    await _maybe_show_category_for_item(
        state, message, bot, items[-1],
        user_id=user.id,
        username=user.username or "",
        user_name=user.full_name or "",
    )


async def on_user_message_to_support(message: Message, bot: Bot) -> None:
    """Любое сообщение от пользователя пересылается в поддержку."""
    # Только личные чаты
    if message.chat.type != "private":
        return
    
    user = message.from_user
    if not user:
        return
    
    # Проверяем статус пользователя
    if not is_user_approved(user.id):
        # Если пользователь не одобрен — не пересылаем
        return

    topics = load_support_topics()
    topic_id = topics.get(user.id)

    async def create_new_topic():
        """Создаёт новый топик для пользователя."""
        user_name = user.full_name or f"User {user.id}"
        if user.username:
            user_name += f" (@{user.username})"

        forum_topic = await bot.create_forum_topic(
            chat_id=SUPPORT_GROUP_ID,
            name=user_name[:128],
        )
        new_topic_id = forum_topic.message_thread_id
        save_support_topic(user.id, new_topic_id)

        # Приветственное сообщение в топик
        await bot.send_message(
            chat_id=SUPPORT_GROUP_ID,
            message_thread_id=new_topic_id,
            text=(
                f"🆕 Новый диалог!\n\n"
                f"👤 Пользователь: {user.full_name}\n"
                f"🆔 ID: {user.id}\n"
                f"📱 Username: @{user.username or 'нет'}"
            ),
        )
        return new_topic_id

    # Если топика нет — создаём
    if not topic_id:
        try:
            topic_id = await create_new_topic()
        except Exception as e:
            await message.answer(f"❌ Не удалось создать чат с поддержкой: {e}")
            return

    # Сначала пересылаем сообщение в топик, чтобы получить message_id для ссылки
    forwarded_msg_id = None
    try:
        forwarded = await message.forward(
            chat_id=SUPPORT_GROUP_ID,
            message_thread_id=topic_id,
        )
        forwarded_msg_id = forwarded.message_id
        await message.answer("✅ Сообщение отправлено в поддержку.")
    except Exception as e:
        if "thread not found" in str(e).lower() or "message thread not found" in str(e).lower():
            try:
                topic_id = await create_new_topic()
                forwarded = await message.forward(
                    chat_id=SUPPORT_GROUP_ID,
                    message_thread_id=topic_id,
                )
                forwarded_msg_id = forwarded.message_id
                await message.answer("✅ Сообщение отправлено в поддержку.")
            except Exception as e2:
                await message.answer(f"❌ Не удалось отправить сообщение: {e2}")
        else:
            await message.answer(
                f"❌ Не удалось отправить сообщение: {e}"
            )

    # Любые ссылки в сообщении — добавляем как лиды (даже без режима отчёта)
    content = _extract_text_with_urls(message)
    # Фото без подписи — подсказка, чтобы лид засчитался
    if (message.photo or message.document) and not content:
        await message.answer(
            "📷 Чтобы лид попал в базу, укажите контакт в подписи к файлу: "
            "@username, ссылку или номер телефона.",
        )
    elif content:
        contacts = extract_contacts_from_text(content)
        if contacts:
            ensure_leads_csv_exists()
            user_id = user.id
            username = user.username or ""
            msg_link_raw = ""
            if forwarded_msg_id and topic_id:
                chat_short = str(SUPPORT_GROUP_ID).replace("-100", "")
                msg_link_raw = f"https://t.me/c/{chat_short}/{topic_id}/{forwarded_msg_id}"
            content_lower = content.lower()
            tg_hint = " тг" in content_lower or "тг " in content_lower or " tg" in content_lower or "tg " in content_lower
            kwork_hint = bool(KWORK_LEAD_KEYWORDS.search(content))

            for contact in contacts:
                duplicate = check_lead_duplicate(contact)
                if duplicate:
                    dup_type, dup_user_id, dup_username = duplicate
                    await bot.send_message(
                        chat_id=SUPPORT_GROUP_ID,
                        message_thread_id=LEADS_TOPIC_ID,
                        text=(
                            f"⚠️ Дубликат лида (не добавлен)\n\n"
                            f"📋 Контакт: {contact}\n"
                            f"📦 Уже в базе: {LEAD_TYPES.get(dup_type, {}).get('name', dup_type)}\n"
                            f"👤 Отправил: {user.full_name} (@{username or 'нет'})\n"
                            f"🆔 ID: {user_id}\n"
                            f"📌 Добавлен ранее: {dup_user_id} (@{dup_username})"
                        ),
                    )
                    await asyncio.sleep(FLOOD_DELAY)
                    await message.answer(
                        f"⚠️ Контакт {contact} уже есть в базе, повторно не добавлен."
                    )
                    await asyncio.sleep(FLOOD_DELAY)
                    continue
                contact_type = determine_contact_type(contact, user_id)
                if not contact_type or contact_type not in LEAD_TYPES:
                    if tg_hint:
                        contact_type = "telegram"
                    elif kwork_hint:
                        contact_type = "kwork"
                    else:
                        contact_type = "self"
                in_base = bool(contact_type) and determine_contact_type(contact, user_id) == contact_type
                src_name = LEAD_TYPES[contact_type]['name'].lower()
                source = "база" if in_base else ("самостоятельный" if contact_type == "self" else f"самостоятельный {src_name}")
                try:
                    if add_lead(contact, contact_type, user_id, username, source=source, message_link=msg_link_raw):
                        msg_link_html = f'\n\n📨 <a href="{msg_link_raw}">Открыть сообщение</a>' if msg_link_raw else ""
                        await bot.send_message(
                            chat_id=SUPPORT_GROUP_ID,
                            message_thread_id=LEADS_TOPIC_ID,
                            text=(
                                f"✅ Лид из сообщения в поддержку\n\n"
                                f"📋 Контакт: {contact}\n"
                                f"📦 Категория: {LEAD_TYPES[contact_type]['name']}\n"
                                f"👤 От: {user.full_name} (@{username or 'нет'})"
                                f"{msg_link_html}"
                            ),
                            parse_mode="HTML",
                        )
                        await asyncio.sleep(FLOOD_DELAY)
                except Exception as e:
                    print(f"Ошибка добавления лида {contact}: {e}")


async def on_support_admin_reply(message: Message, bot: Bot) -> None:
    """Админ ответил в топике — отправляем пользователю от имени бота."""
    # Проверяем, что это сообщение из группы поддержки
    if message.chat.id != SUPPORT_GROUP_ID:
        return

    # Проверяем, что это ответ в топике (не в General)
    topic_id = message.message_thread_id
    if not topic_id:
        return

    # Игнорируем сообщения от бота
    if message.from_user and message.from_user.is_bot:
        return

    # Находим пользователя по topic_id
    user_id = get_user_by_topic(topic_id)
    if not user_id:
        return

    try:
        # Отправляем сообщение от имени бота (без пересылки)
        if message.text:
            await bot.send_message(
                chat_id=user_id,
                text=f"💬 Поддержка:\n\n{message.text}",
            )
        elif message.photo:
            await bot.send_photo(
                chat_id=user_id,
                photo=message.photo[-1].file_id,
                caption=f"💬 Поддержка:\n\n{message.caption or ''}",
            )
        elif message.document:
            await bot.send_document(
                chat_id=user_id,
                document=message.document.file_id,
                caption=f"💬 Поддержка:\n\n{message.caption or ''}",
            )
        elif message.voice:
            await bot.send_voice(
                chat_id=user_id,
                voice=message.voice.file_id,
                caption="💬 Голосовое от поддержки",
            )
        elif message.video:
            await bot.send_video(
                chat_id=user_id,
                video=message.video.file_id,
                caption=f"💬 Поддержка:\n\n{message.caption or ''}",
            )
        elif message.sticker:
            await bot.send_sticker(
                chat_id=user_id,
                sticker=message.sticker.file_id,
            )
    except Exception:
        # Пользователь мог заблокировать бота
        pass


# ============ МЕНЕДЖЕР: РАЗБЛОКИРОВКА ЛИМИТОВ ============

def get_user_used_types(user_id: int) -> List[str]:
    """Возвращает список типов баз, которые пользователь уже получал."""
    used = []
    for key, info in BASE_TYPES.items():
        csv_path = info["csv"]
        rows = _read_csv(csv_path)
        count = _count_user_records(rows, user_id)
        if count > 0:
            used.append(key)
    return used


def get_user_contacts(user_id: int) -> Dict[str, List[str]]:
    """Возвращает все контакты, выданные пользователю, по типам."""
    result = {}
    for key, info in BASE_TYPES.items():
        csv_path = info["csv"]
        rows = _read_csv(csv_path)
        contacts = []
        for row in rows[1:]:  # Пропускаем заголовок
            if len(row) >= 2 and row[1] == str(user_id):
                value = row[0].strip()
                if value.startswith("="):
                    value = value[1:]
                if value:
                    contacts.append(value)
        if contacts:
            result[key] = contacts
    return result


def _create_user_contacts_excel(user_id: int, contacts: Dict[str, List[str]]) -> tuple[io.BytesIO, str]:
    """Создаёт Excel-файл с контактами пользователя."""
    wb = Workbook()
    first = True
    
    for key, values in contacts.items():
        info = BASE_TYPES[key]
        if first:
            ws = wb.active
            ws.title = info["name"][:31]  # Максимум 31 символ для названия листа
            first = False
        else:
            ws = wb.create_sheet(title=info["name"][:31])
        
        ws.append(["Контакт"])
        for val in values:
            ws.append([val])
    
    if first:
        # Нет контактов — пустой файл
        ws = wb.active
        ws.title = "Пусто"
        ws.append(["Нет выданных контактов"])
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"contacts_user_{user_id}.xlsx"
    return buffer, filename


async def on_contacts_command(message: Message) -> None:
    """Команда /contacts в топике — показать выданные контакты пользователю."""
    # Только в группе поддержки
    if message.chat.id != SUPPORT_GROUP_ID:
        return
    
    topic_id = message.message_thread_id
    if not topic_id:
        await message.answer("❌ Эта команда работает только в топике пользователя.")
        return
    
    # Находим пользователя по топику
    user_id = get_user_by_topic(topic_id)
    if not user_id:
        await message.answer("❌ Пользователь не найден для этого топика.")
        return
    
    await message.answer("⏳ Собираю контакты пользователя...")
    
    # Получаем контакты
    contacts = await asyncio.to_thread(get_user_contacts, user_id)
    
    if not contacts:
        await message.answer("ℹ️ Этому пользователю ещё не выдавались контакты.")
        return
    
    # Формируем статистику
    stats = []
    total = 0
    for key, values in contacts.items():
        info = BASE_TYPES[key]
        stats.append(f"• {info['name']}: {len(values)}")
        total += len(values)
    
    # Создаём Excel
    file_buffer, filename = await asyncio.to_thread(
        _create_user_contacts_excel, user_id, contacts
    )
    document = BufferedInputFile(file_buffer.read(), filename=filename)
    
    await message.answer_document(
        document=document,
        caption=(
            f"📋 Контакты пользователя {user_id}:\n\n"
            + "\n".join(stats) +
            f"\n\n📊 Всего: {total} контактов"
        ),
    )


async def on_clear_command(message: Message, bot: Bot) -> None:
    """Команда /clear в топике — автоматическая разблокировка лимитов."""
    # Проверяем, что это в группе поддержки
    if message.chat.id != SUPPORT_GROUP_ID:
        return
    
    topic_id = message.message_thread_id
    if not topic_id:
        await message.answer("❌ Эта команда работает только в топике пользователя.")
        return
    
    # Находим пользователя по топику
    user_id = get_user_by_topic(topic_id)
    if not user_id:
        await message.answer("❌ Пользователь не найден для этого топика.")
        return
    
    # Находим типы, которые пользователь уже получал
    used_types = await asyncio.to_thread(get_user_used_types, user_id)
    
    if not used_types:
        await message.answer("ℹ️ Пользователь ещё не получал никаких контактов.")
        return
    
    # Разблокируем только те типы, где пользователь использовал весь лимит
    unlocked = []
    unlocked_keys = []
    skipped = []
    
    for key in used_types:
        info = BASE_TYPES[key]
        base_limit = info["limit"]
        extra_limit = get_user_extra_limit(user_id, key)
        total_allowed = base_limit + extra_limit
        
        # Считаем сколько уже получил
        csv_path = info["csv"]
        rows = _read_csv(csv_path)
        current = _count_user_records(rows, user_id)
        
        # Разблокируем только если использовал весь текущий лимит
        if current >= total_allowed:
            # Устанавливаем extra_limit = current, чтобы можно было получить ещё base_limit
            set_user_extra_limit(user_id, key, current)
            unlocked.append(f"• {info['name']} (+{base_limit})")
            unlocked_keys.append(key)
        else:
            remaining = total_allowed - current
            skipped.append(f"• {info['name']} (осталось {remaining})")
    
    if not unlocked:
        await message.answer("ℹ️ Пользователь ещё не использовал текущий лимит.")
        return
    
    await message.answer(f"✅ Разблокировано для пользователя:\n\n" + "\n".join(unlocked))
    
    # Уведомляем пользователя
    try:
        unlocked_names = [BASE_TYPES[k]["name"] for k in unlocked_keys]
        await bot.send_message(
            chat_id=user_id,
            text=(
                "🎉 Менеджер разблокировал тебе контакты!\n\n"
                f"Разблокировано: {', '.join(unlocked_names)}\n\n"
                "Теперь ты можешь получить ещё одну порцию."
            ),
        )
    except Exception:
        pass


# ============ ЗАПУСК ============

async def main() -> None:
    load_dotenv()
    token = os.getenv("BOT_TOKEN")
    if not token:
        raise RuntimeError("Не задан BOT_TOKEN в .env файле")

    # Создаём CSV если нужно
    ensure_csv_exists()
    ensure_leads_csv_exists()

    bot = Bot(token=token)
    storage = MemoryStorage()
    dp = Dispatcher(storage=storage)

    # Базовые команды
    dp.message.register(on_start, CommandStart())
    dp.message.register(on_admin_command, Command(ADMIN_SECRET_COMMAND))
    dp.message.register(on_chatid, Command("chatid"))
    dp.message.register(on_get_online, Command("get_online"))
    dp.message.register(on_download_db, Command("download_db"))
    dp.message.register(on_stats, Command("stats"))
    dp.message.register(on_leadstats, Command("leadstats"))
    dp.message.register(on_download_lead, Command("download_lead"))
    dp.message.register(on_download_lead_day, Command("download_lead_day"))
    dp.message.register(on_add_lead_start, Command("add_lead"))
    dp.message.register(on_delete_lead_start, Command("delete_lead"))
    dp.message.register(on_check_leads, Command("check_leads"), F.chat.id == SUPPORT_GROUP_ID)
    
    # Ручное добавление лида (состояния)
    dp.message.register(
        on_add_lead_cancel,
        ManualLeadStates.waiting_contact,
        F.text == "⬅️ Отмена",
    )
    dp.message.register(
        on_add_lead_cancel,
        ManualLeadStates.waiting_category,
        F.text == "⬅️ Отмена",
    )
    dp.callback_query.register(
        on_add_lead_category_callback,
        StateFilter(ManualLeadStates.waiting_category),
        F.data.startswith("lead_cat_"),
    )
    dp.callback_query.register(
        on_report_category_callback,
        StateFilter(ReportStates.waiting_category),
        F.data.startswith("report_cat_"),
    )
    dp.message.register(
        on_add_lead_contact,
        ManualLeadStates.waiting_contact,
    )
    
    # Удаление лида (состояния)
    dp.message.register(
        on_delete_lead_cancel,
        DeleteLeadStates.waiting_contact,
        F.text == "⬅️ Отмена",
    )
    dp.message.register(
        on_delete_lead_contact,
        DeleteLeadStates.waiting_contact,
    )
    
    # Регистрация пользователя
    dp.message.register(on_send_request, F.text == "✅ Отправить приглашение")
    
    # Команды модерации (в группе поддержки)
    dp.message.register(on_add_user, Command("add"), F.chat.id == SUPPORT_GROUP_ID)
    dp.message.register(on_ban_user, Command("ban"), F.chat.id == SUPPORT_GROUP_ID)
    dp.message.register(on_unban_user, Command("unban"), F.chat.id == SUPPORT_GROUP_ID)

    # Админ: состояние ожидания файла (должно быть выше остальных!)
    dp.message.register(
        on_admin_file_received,
        AdminStates.waiting_file,
        F.document,
    )
    dp.message.register(
        on_admin_upload_cancel,
        AdminStates.waiting_file,
        F.text == "⬅️ Отмена",
    )

    # Админ: состояние выбора типа загрузки
    dp.message.register(
        on_admin_upload_cancel,
        AdminStates.waiting_upload_choice,
        F.text == "⬅️ Отмена",
    )
    dp.message.register(
        on_admin_upload_choice,
        AdminStates.waiting_upload_choice,
    )

    # Админ: состояние подтверждения удаления
    dp.message.register(
        on_admin_delete_confirm,
        AdminStates.waiting_delete_confirm,
        F.text == "✅ Да, удалить всё",
    )
    dp.message.register(
        on_admin_delete_cancel,
        AdminStates.waiting_delete_confirm,
        F.text == "❌ Отмена",
    )

    # Менеджер: команда /clear в группе поддержки (ДО on_support_admin_reply!)
    dp.message.register(
        on_clear_command,
        Command("clear"),
        F.chat.id == SUPPORT_GROUP_ID,
    )
    
    # Менеджер: команда /contacts в группе поддержки (ДО on_support_admin_reply!)
    dp.message.register(
        on_contacts_command,
        Command("contacts"),
        F.chat.id == SUPPORT_GROUP_ID,
    )

    # Поддержка: ответы админов из группы (только supergroup, не личные чаты)
    dp.message.register(
        on_support_admin_reply,
        F.chat.type == "supergroup",
        F.chat.id == SUPPORT_GROUP_ID,
        ~Command("clear"),
        ~Command("contacts"),
        ~Command("add"),
        ~Command("ban"),
        ~Command("unban"),
    )
    
    # Админ: основные кнопки
    dp.message.register(on_admin_download, F.text == "📤 Выкачать Базу данных")
    dp.message.register(on_admin_upload_start, F.text == "📥 Загрузить Базу данных")
    dp.message.register(on_admin_delete_start, F.text == "🗑 Удалить всю базу данных")
    dp.message.register(on_admin_exit, F.text == "⬅️ Выход из админки")

    # Пользователь: навигация
    dp.message.register(on_get_base, F.text == "📦 Получить списки контактов")
    dp.message.register(on_report_start, F.text == "📋 Отчёт по лидам")
    dp.message.register(on_support_info, F.text == "💬 Написать в поддержку")
    dp.message.register(on_user_lead_stats, F.text == "📊 Статистика лидов")
    dp.message.register(on_request_new_contacts, F.text == "🆕 Получить новые контакты")
    dp.message.register(on_back, F.text == "⬅️ Назад")

    # Отчёты: сбор и отправка (ДО on_user_message_to_support!)
    # «Отправить отчёт» работает и в режиме выбора категории — обработаем, не сломаем бота
    dp.message.register(
        on_report_submit,
        StateFilter(ReportStates.waiting_report, ReportStates.waiting_category),
        F.text == "📤 Отправить отчёт",
    )
    dp.message.register(
        on_report_cancel,
        StateFilter(ReportStates.waiting_report),
        F.text == "❌ Отмена",
    )
    dp.message.register(
        on_report_cancel,
        StateFilter(ReportStates.waiting_category),
        F.text == "❌ Отмена",
    )
    dp.message.register(
        on_report_file,
        StateFilter(ReportStates.waiting_report),
        F.photo | F.document | F.video,
    )
    dp.message.register(
        on_report_other,
        StateFilter(ReportStates.waiting_report),
    )
    dp.message.register(
        on_report_waiting_category_remind,
        StateFilter(ReportStates.waiting_category),
        F.photo | F.document | F.video | F.text,
    )

    # Пользователь: выбор типа базы
    for btn_text in USER_BUTTON_MAP:
        dp.message.register(on_user_base_choice, F.text == btn_text)

    # Сообщения в поддержку — ТОЛЬКО когда пользователь нажал «Написать в поддержку»
    dp.message.register(
        on_user_message_to_support,
        StateFilter(SupportStates.active),
        F.chat.type == "private",
    )

    print("Бот запущен!")
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
