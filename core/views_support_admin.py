import os
import threading
import uuid
from datetime import date, datetime, time, timedelta, timezone as dt_utc
from io import BytesIO
from zoneinfo import ZoneInfo

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.utils import OperationalError, ProgrammingError
from django.db.models import Case, CharField, Count, F, IntegerField, Max, Q, Sum, Value, When
from django.db.models.functions import Cast
from django.http import FileResponse, HttpRequest, HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook, load_workbook

from .forms import BaseCategoryUploadForm, BaseExcelUploadForm, LeadRejectForm, LeadReworkForm, LeadsExcelUploadForm
from .models import (
    BaseType,
    BasesImportJob,
    Contact,
    ContactRequest,
    Lead,
    LeadType,
    SiteSettings,
    SupportMessage,
    SupportThread,
    User,
    UserBaseLimit,
    WithdrawalRequest,
    WorkerSelfLead,
    log_balance_change,
)


def _require_support(request: HttpRequest) -> bool:
    user = request.user
    if not user.is_authenticated:
        return False
    # Доступ для ролей поддержки/админов и для staff/superuser
    if getattr(user, "is_support", False) or user.is_staff or user.is_superuser:
        return True
    return False


def _require_support_or_partner(request: HttpRequest) -> bool:
    """Доступ для админов + партнёров (для баз, лидов, контактов)."""
    if _require_support(request):
        return True
    return getattr(request.user, "role", None) == "partner"


def _require_standalone_admin(request: HttpRequest) -> bool:
    """Только роль «Самостоятельный админ». Выдаётся только суперадмином."""
    user = request.user
    if not user.is_authenticated:
        return False
    return getattr(user, "role", None) == "standalone_admin"


@login_required
def admin_users_pending(request: HttpRequest) -> HttpResponse:
    """Список новых пользователей со статусом pending с кнопками одобрения/бана."""

    if not _require_support(request):
        return HttpResponseForbidden("Недостаточно прав.")

    auto_approve_enabled = False
    settings_error = None
    try:
        site_settings = SiteSettings.get_settings()
        auto_approve_enabled = site_settings.auto_approve_users
    except Exception as e:
        site_settings = None
        settings_error = str(e)

    if request.method == "POST":
        if request.POST.get("action") == "toggle_auto_approve":
            if site_settings:
                site_settings.auto_approve_users = not site_settings.auto_approve_users
                site_settings.save(update_fields=["auto_approve_users"])
                if site_settings.auto_approve_users:
                    messages.success(request, "Автоодобрение включено. Новые пользователи будут одобряться автоматически.")
                else:
                    messages.info(request, "Автоодобрение выключено.")
            else:
                messages.error(request, "Не удалось переключить: настройки сайта недоступны. Проверьте, что миграции применены: python manage.py migrate core")
            return redirect("admin_users_pending")
        
        user_id = request.POST.get("user_id")
        action = request.POST.get("action")
        if user_id and action in {"approve", "ban"}:
            user = get_object_or_404(User, pk=user_id, status=User.Status.PENDING)
            if action == "approve":
                user.status = User.Status.APPROVED
                user.is_active = True
                messages.success(request, f"Пользователь @{user.username} одобрен.")
            elif action == "ban":
                user.status = User.Status.BANNED
                user.is_active = False  # Django auto-logout + блок логина
                messages.warning(request, f"Пользователь @{user.username} заблокирован.")
            user.save(update_fields=["status", "is_active"])
        return redirect("admin_users_pending")

    pending_users = User.objects.filter(status=User.Status.PENDING).order_by("-date_joined")

    return render(
        request,
        "core/admin_users_pending.html",
        {
            "pending_users": pending_users,
            "auto_approve_enabled": auto_approve_enabled,
            "settings_error": settings_error,
        },
    )


@login_required
def standalone_admin_ss_leads(request: HttpRequest) -> HttpResponse:
    """Страница самостоятельного админа: одобренные СС-лиды с выбором статуса."""
    if not _require_standalone_admin(request):
        return HttpResponseForbidden("Недостаточно прав. Только роль «Самостоятельный админ».")

    try:
        return _standalone_admin_ss_leads_impl(request)
    except (OperationalError, ProgrammingError) as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception("СС лиды: ошибка БД (миграция?): %s", e)
        messages.error(request, "Ошибка базы данных. Убедитесь, что выполнена миграция: python manage.py migrate core")
        return redirect("dashboard")


def _standalone_admin_ss_leads_impl(request: HttpRequest) -> HttpResponse:
    tab = request.GET.get("tab", "new")
    if tab not in ("new", "rejected", "in_progress", "meeting", "assigned"):
        tab = "new"

    base_qs = Lead.objects.filter(
        status=Lead.Status.APPROVED,
        needs_team_contact=True,
    ).select_related("user", "lead_type", "base_type")

    # Поиск
    search_q = (request.GET.get("q") or "").strip().lstrip("@")[:100]

    from django.db.models import Prefetch
    from .models import LeadAssignment as _LA
    _assignment_prefetch = Prefetch(
        "assignments",
        queryset=_LA.objects.filter(
            worker__standalone_admin_owner=request.user
        ).select_related("worker", "assigned_by").order_by("-created_at"),
    )

    if tab == "assigned":
        leads = (
            base_qs
            .filter(assignments__worker__standalone_admin_owner=request.user)
            .distinct()
            .prefetch_related(_assignment_prefetch)
            .order_by("-updated_at", "-id")
        )
    elif tab == "new":
        leads = base_qs.filter(ss_admin_status__isnull=True).order_by("-reviewed_at", "-id")
    elif tab == "in_progress":
        leads = (
            base_qs.filter(ss_admin_status="in_progress")
            .prefetch_related(_assignment_prefetch)
            .order_by("-updated_at", "-id")
        )
    else:
        leads = base_qs.filter(ss_admin_status=tab).order_by("-updated_at", "-id")

    if search_q:
        from django.db.models.functions import Cast
        from django.db.models import CharField as _CharField
        leads = leads.annotate(_id_str=Cast("id", _CharField()))
        words = [w.strip() for w in search_q.split() if w.strip()]
        for word in words:
            leads = leads.filter(
                Q(user__username__icontains=word)
                | Q(raw_contact__icontains=word)
                | Q(normalized_contact__icontains=word)
                | Q(lead_type__name__icontains=word)
                | Q(comment__icontains=word)
                | Q(_id_str__icontains=word)
            )

    if request.method == "POST":
        lead_id = request.POST.get("lead_id")
        new_status = request.POST.get("ss_admin_status")
        if lead_id and new_status in ("rejected", "in_progress", "meeting", ""):
            lead = base_qs.filter(pk=lead_id).first()
            if lead:
                lead.ss_admin_status = new_status or None
                lead.save(update_fields=["ss_admin_status", "updated_at"])
                # AJAX — JSON ответ без редиректа
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    status_labels = {"rejected": "Отказ", "in_progress": "В работе", "meeting": "Встреча", "": "Новые"}
                    return JsonResponse({"success": True, "message": f"Лид #{lead_id} → {status_labels.get(new_status, new_status)}"})
                messages.success(request, f"Лид #{lead_id} перемещён.")
        # Сохраняем page и q при обычном POST
        page_num_post = request.POST.get("page") or request.GET.get("page") or ""
        search_q_post = request.POST.get("q") or request.GET.get("q") or ""
        redir_url = reverse("standalone_admin_ss_leads") + f"?tab={tab}"
        if page_num_post:
            redir_url += f"&page={page_num_post}"
        if search_q_post:
            from urllib.parse import quote
            redir_url += f"&q={quote(search_q_post)}"
        return redirect(redir_url)

    counts = {
        "new": base_qs.filter(ss_admin_status__isnull=True).count(),
        "rejected": base_qs.filter(ss_admin_status="rejected").count(),
        "in_progress": base_qs.filter(ss_admin_status="in_progress").count(),
        "meeting": base_qs.filter(ss_admin_status="meeting").count(),
        "assigned": base_qs.filter(
            assignments__worker__standalone_admin_owner=request.user
        ).distinct().count(),
    }
    paginator = Paginator(leads, 30)
    try:
        page_num = int(request.GET.get("page", 1))
    except (TypeError, ValueError):
        page_num = 1
    page_obj = paginator.get_page(page_num)

    # Для вкладки «Назначены» — аннотируем статус отчёта и флаг «все одобрены»
    if tab == "assigned":
        from .models import WorkerReport
        for lead in page_obj:
            all_approved = True
            has_any = False
            for a in lead.assignments.all():
                has_any = True
                try:
                    a.report_status = a.report.status
                except WorkerReport.DoesNotExist:
                    a.report_status = None
                if a.report_status != "approved":
                    all_approved = False
            lead.all_assignments_approved = has_any and all_approved

    return render(
        request,
        "core/standalone_admin_ss_leads.html",
        {
            "page_obj": page_obj,
            "tab": tab,
            "counts": counts,
            "search_q": search_q,
        },
    )


@login_required
def support_threads_list(request: HttpRequest) -> HttpResponse:
    if not _require_support(request):
        return HttpResponseForbidden("Недостаточно прав.")

    threads = (
        SupportThread.objects.select_related("user")
        .annotate(
            last_message_at=Max("messages__created_at"),
            messages_count=Count("messages"),
            is_unread=Case(
                When(Q(messages_count=0), then=Value(0)),
                When(Q(last_read_at__isnull=True), then=Value(1)),
                When(updated_at__gt=F("last_read_at"), then=Value(1)),
                default=Value(0),
                output_field=IntegerField(),
            ),
        )
        .order_by("-is_unread", "-updated_at")
    )
    threads_agg = SupportThread.objects.aggregate(m=Max("updated_at"))
    admin_threads_updated_at = threads_agg["m"].isoformat() if threads_agg.get("m") else ""

    return render(
        request,
        "core/support_threads_list.html",
        {
            "threads": threads,
            "admin_threads_updated_at": admin_threads_updated_at,
            "disable_polling": True,
        },
    )


@login_required
def support_thread_detail(request: HttpRequest, pk: int) -> HttpResponse:
    if not _require_support(request):
        return HttpResponseForbidden("Недостаточно прав.")

    thread = get_object_or_404(SupportThread.objects.select_related("user"), pk=pk)

    # Открытие диалога = «прочитано»
    thread.last_read_at = timezone.now()
    thread.save(update_fields=["last_read_at"])

    if request.method == "POST":
        text = (request.POST.get("text") or "").strip()
        if text:
            SupportMessage.objects.create(
                thread=thread,
                sender=request.user,
                is_from_support=True,
                text=text,
            )
            thread.updated_at = timezone.now()
            thread.save(update_fields=["updated_at"])
            messages.success(request, "Ответ отправлен пользователю.")
            return redirect("support_thread_detail", pk=thread.pk)

    messages_qs = thread.messages.select_related("sender").order_by("created_at")
    threads_agg = SupportThread.objects.aggregate(m=Max("updated_at"))
    admin_threads_updated_at = threads_agg["m"].isoformat() if threads_agg.get("m") else ""

    return render(
        request,
        "core/support_thread_detail.html",
        {
            "thread": thread,
            "support_messages": messages_qs,
            "admin_threads_updated_at": admin_threads_updated_at,
            "disable_polling": True,
        },
    )


@login_required
def support_message_delete(request: HttpRequest, pk: int) -> HttpResponse:
    """Удаление сообщения в диалоге (только для сотрудников поддержки)."""
    if not _require_support(request):
        return HttpResponseForbidden("Недостаточно прав.")
    msg = get_object_or_404(SupportMessage.objects.select_related("thread"), pk=pk)
    thread_pk = msg.thread_id
    msg.delete()
    messages.success(request, "Сообщение удалено.")
    return redirect("support_thread_detail", pk=thread_pk)


@login_required
def support_thread_delete(request: HttpRequest, pk: int) -> HttpResponse:
    """Удаление чатов отключено — сохранение истории обращений."""
    return HttpResponseForbidden("Удаление чатов отключено.")


@login_required
def support_thread_by_user(request: HttpRequest, user_id: int) -> HttpResponse:
    """Открыть диалог поддержки с пользователем (редирект на страницу диалога)."""
    if not _require_support(request):
        return HttpResponseForbidden("Недостаточно прав.")
    target_user = get_object_or_404(User, pk=user_id)
    thread, _ = SupportThread.objects.get_or_create(user=target_user, is_closed=False)
    return redirect("support_thread_detail", pk=thread.pk)


def _day_bounds_lead_stats():
    """Границы «дня» для лидов (20:00 МСК), как в боте."""
    tz = ZoneInfo("Europe/Moscow")
    now = datetime.now(tz)
    from_day = now.date()
    if now.hour >= 20:
        from_day = now.date() + timedelta(days=1)

    def bounds(day: date):
        start = datetime.combine(day - timedelta(days=1), time(hour=20), tzinfo=tz)
        end = datetime.combine(day, time(hour=20), tzinfo=tz)
        return start.astimezone(dt_utc.utc), end.astimezone(dt_utc.utc)

    today_start, today_end = bounds(from_day)
    yesterday_start, yesterday_end = bounds(from_day - timedelta(days=1))
    return today_start, today_end, yesterday_start, yesterday_end


@login_required
def admin_user_lead_stats(request: HttpRequest, user_id: int) -> HttpResponse:
    """Статистика лидов по выбранному пользователю (для админа)."""
    if not _require_support(request):
        return HttpResponseForbidden("Недостаточно прав.")
    target_user = get_object_or_404(User, pk=user_id)
    today_start, today_end, yesterday_start, yesterday_end = _day_bounds_lead_stats()
    today_count = Lead.objects.filter(
        user=target_user,
        status=Lead.Status.APPROVED,
        created_at__gte=today_start,
        created_at__lt=today_end,
    ).count()
    yesterday_count = Lead.objects.filter(
        user=target_user,
        status=Lead.Status.APPROVED,
        created_at__gte=yesterday_start,
        created_at__lt=yesterday_end,
    ).count()
    total_count = Lead.objects.filter(user=target_user, status=Lead.Status.APPROVED).count()
    return render(
        request,
        "core/admin_user_lead_stats.html",
        {
            "target_user": target_user,
            "today_count": today_count,
            "yesterday_count": yesterday_count,
            "total_count": total_count,
        },
    )


@login_required
def admin_user_leads_list(request: HttpRequest, user_id: int) -> HttpResponse:
    """Список лидов (отчётов) пользователя по вкладкам: новые, в доработке, принятые, отклонённые."""
    if not _require_support(request):
        return HttpResponseForbidden("Недостаточно прав.")
    target_user = get_object_or_404(User, pk=user_id)
    tab = request.GET.get("tab", "all")
    if tab not in ("all", "new", "rework", "approved", "rejected"):
        tab = "all"
    base = Lead.objects.filter(user=target_user).select_related("lead_type", "base_type", "reviewed_by")
    if tab == "all":
        qs = base.order_by("-created_at")
    elif tab == "new":
        qs = base.filter(status=Lead.Status.PENDING).order_by("created_at")
    elif tab == "rework":
        qs = base.filter(status=Lead.Status.REWORK).order_by("created_at")
    elif tab == "approved":
        qs = base.filter(status=Lead.Status.APPROVED).order_by("-reviewed_at")
    else:
        qs = base.filter(status=Lead.Status.REJECTED).order_by("-reviewed_at")
    paginator = Paginator(qs, 50)
    page_number = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_number)
    lead_approve_reward = getattr(settings, "LEAD_APPROVE_REWARD", 40)
    return render(
        request,
        "core/admin_user_leads_list.html",
        {
            "target_user": target_user,
            "page_obj": page_obj,
            "tab": tab,
            "lead_approve_reward": lead_approve_reward,
        },
    )


@login_required
def admin_leads_all_new(request: HttpRequest) -> HttpResponse:
    """Единая страница «Отчёты»: 4 вкладки — Новые отчёты, В доработке, Принятые, Отклонённые."""
    if not _require_support(request):
        return HttpResponseForbidden("Недостаточно прав.")
    tab = request.GET.get("tab", "new")
    if tab not in ("new", "rework", "approved", "rejected", "all"):
        tab = "new"
    dept = request.GET.get("dept", "")
    try:
        base = Lead.objects.select_related("user", "lead_type", "base_type", "reviewed_by")
        # Исключить дожим-лиды рефов Nastya_Partner (id=260) — она проверяет их сама
        base = base.exclude(lead_type__slug="dozhim", user__partner_owner_id=260)
        if dept == "dozhim":
            base = base.filter(lead_type__slug="dozhim")
        elif dept == "search":
            base = base.exclude(lead_type__slug="dozhim")
        if tab == "new":
            qs = base.filter(status=Lead.Status.PENDING).order_by("created_at")
        elif tab == "rework":
            qs = base.filter(status=Lead.Status.REWORK).order_by("created_at")
        elif tab == "approved":
            qs = base.filter(status=Lead.Status.APPROVED).order_by("-reviewed_at")
        elif tab == "rejected":
            qs = base.filter(status=Lead.Status.REJECTED).order_by("-reviewed_at")
        else:  # all
            qs = base.order_by("-created_at")
            search_q = (request.GET.get("q") or "").strip().lstrip("@")[:100]
            if search_q:
                qs = qs.annotate(id_str=Cast("id", CharField()))
                words = [w.strip() for w in search_q.split() if w.strip()]
                for word in words:
                    date_filt = None
                    time_filt = None
                    try:
                        if len(word) == 10 and word[4] == "-" and word[7] == "-":  # YYYY-MM-DD
                            date_filt = datetime.strptime(word, "%Y-%m-%d").date()
                        elif len(word) == 10 and word[2] == "." and word[5] == ".":  # DD.MM.YYYY
                            date_filt = datetime.strptime(word, "%d.%m.%Y").date()
                        elif len(word) == 5 and word[2] == ".":  # DD.MM
                            date_filt = datetime.strptime(word, "%d.%m").date()
                        elif len(word) == 5 and word[2] == ":" and word.replace(":", "").isdigit():  # HH:MM
                            time_filt = datetime.strptime(word, "%H:%M").time()
                        elif len(word) == 5 and word[2] == "." and word.replace(".", "").isdigit():  # HH.MM
                            time_filt = datetime.strptime(word, "%H.%M").time()
                    except ValueError:
                        pass
                    base_q = (
                        Q(user__username__icontains=word)
                        | Q(raw_contact__icontains=word)
                        | Q(normalized_contact__icontains=word)
                        | Q(contact__value__icontains=word)
                        | Q(source__icontains=word)
                        | Q(comment__icontains=word)
                        | Q(rework_comment__icontains=word)
                        | Q(rejection_reason__icontains=word)
                        | Q(lead_type__name__icontains=word)
                        | Q(reviewed_by__username__icontains=word)
                    )
                    if word.isdigit():
                        base_q = base_q | Q(id_str__icontains=word)
                    if date_filt:
                        if date_filt.year != 1900:
                            base_q = base_q | Q(created_at__date=date_filt)
                        else:
                            base_q = base_q | Q(created_at__month=date_filt.month, created_at__day=date_filt.day)
                    if time_filt:
                        base_q = base_q | Q(created_at__hour=time_filt.hour, created_at__minute=time_filt.minute)
                    qs = qs.filter(base_q)
        paginator = Paginator(qs, 50)
        page_number = request.GET.get("page", 1)
        page_obj = paginator.get_page(page_number)
    except (OperationalError, ProgrammingError) as e:
        import logging
        logging.getLogger(__name__).exception("admin_leads_all_new: DB error (run migrate?): %s", e)
        return HttpResponse(
            "<!DOCTYPE html><html><head><meta charset='utf-8'><title>Ошибка</title></head><body style='font-family:sans-serif;padding:2rem;'>"
            "<h1>Ошибка базы данных</h1>"
            "<p>Страница «Отчёты» не загружается. Часто это из‑за неприменённых миграций после обновления кода.</p>"
            "<p><strong>На сервере выполните:</strong> <code>python manage.py migrate</code></p>"
            "<p><a href='/staff/'>← В кабинет</a></p></body></html>",
            status=500,
            content_type="text/html; charset=utf-8",
        )
    lead_approve_reward = getattr(settings, "LEAD_APPROVE_REWARD", 40)
    search_query = request.GET.get("q", "").strip() if tab == "all" else ""
    return render(
        request,
        "core/admin_leads_all_new.html",
        {
            "page_obj": page_obj,
            "lead_approve_reward": lead_approve_reward,
            "tab": tab,
            "dept": dept,
            "search_query": search_query,
        },
    )


@login_required
def admin_media_storage_status(request: HttpRequest) -> HttpResponse:
    """Диагностика: куда сохраняются медиа (S3 из env или из админки), подключается ли S3. POST с action=test_write — реальная проверка записи."""
    if not _require_support(request):
        return HttpResponseForbidden("Недостаточно прав.")
    from django.core.files.storage import default_storage
    from django.core.files.base import ContentFile
    from core.storage import get_media_storage_diagnostic
    write_result = None
    if request.method == "POST" and request.POST.get("action") == "test_write":
        test_name = "_media_write_test.txt"
        try:
            default_storage.save(test_name, ContentFile(b"test-write-check"))
            exists = default_storage.exists(test_name)
            default_storage.delete(test_name)
            if exists:
                write_result = {"ok": True, "message": "Запись в хранилище прошла успешно. Файл создан и удалён. Если это S3 — проверьте бакет в панели (возможно, обновление счётчика с задержкой)."}
            else:
                write_result = {"ok": False, "message": "Файл записан, но сразу не найден (exists() вернул False)."}
        except Exception as e:
            import traceback
            write_result = {"ok": False, "message": str(e), "detail": traceback.format_exc()}
    diag = get_media_storage_diagnostic()
    return render(
        request,
        "core/admin_media_storage_status.html",
        {"diag": diag, "write_result": write_result},
    )


LEAD_APPROVE_REWARD = getattr(settings, "LEAD_APPROVE_REWARD", 40)


@login_required
@require_http_methods(["POST"])
def admin_lead_approve(request: HttpRequest, user_id: int, lead_id: int) -> HttpResponse:
    """Одобрить лид: начислить пользователю LEAD_APPROVE_REWARD руб., статус — одобрен."""
    if not _require_support(request):
        return HttpResponseForbidden("Недостаточно прав.")
    from .models import LeadReviewLog
    with transaction.atomic():
        lead = Lead.objects.select_for_update().filter(pk=lead_id, user_id=user_id).select_related("user").first()
        if not lead:
            return redirect("admin_user_leads_list", user_id=user_id)
        if lead.status == Lead.Status.APPROVED:
            messages.info(request, "Лид уже одобрен.")
            return redirect("admin_user_leads_list", user_id=user_id)
        if lead.status not in (Lead.Status.PENDING, Lead.Status.REWORK):
            messages.warning(request, "Можно одобрять только лиды на проверке или на доработке.")
            return redirect("admin_user_leads_list", user_id=user_id)
        lead.status = Lead.Status.APPROVED
        lead.rejection_reason = ""
        lead.rework_comment = ""
        lead.reviewed_at = timezone.now()
        lead.reviewed_by = request.user
        lead.save(update_fields=["status", "rejection_reason", "rework_comment", "reviewed_at", "reviewed_by"])
        # Динамическая награда: дожим → 30р, поиск → 40р
        is_dozhim = lead.lead_type and lead.lead_type.slug == "dozhim"
        reward = getattr(settings, "DOZHIM_APPROVE_REWARD", 40) if is_dozhim else LEAD_APPROVE_REWARD
        partner_earning = 0
        # Партнёрская экономика — только для лидов из Отдела поиска (не дожим)
        if not is_dozhim:
            partner_owner_id = lead.user.partner_owner_id
            if partner_owner_id:
                from .models import PartnerEarning
                partner = User.objects.select_for_update().get(pk=partner_owner_id)
                if partner.role == User.Role.PARTNER:
                    # Старая система (role=partner, Настя — фикс ставка)
                    partner_earning = partner.partner_rate or 10
                else:
                    # Реферальная система: per-user override → ставка с ссылки → 20 по умолчанию
                    if lead.user.ref_lead_reward is not None:
                        ref_reward = max(1, min(LEAD_APPROVE_REWARD - 1, lead.user.ref_lead_reward))
                    else:
                        link = lead.user.partner_link
                        ref_reward = max(1, min(LEAD_APPROVE_REWARD - 1, link.ref_reward if link else 20))
                    reward = ref_reward  # реф получит ref_reward вместо стандартных 40
                    partner_earning = LEAD_APPROVE_REWARD - ref_reward
                PartnerEarning.objects.create(partner=partner, lead=lead, amount=partner_earning)
                from .models import log_balance_change
                _old_pb = partner.balance or 0
                partner.balance = _old_pb + partner_earning
                partner.save(update_fields=["balance"])
                log_balance_change(partner, "balance", _old_pb, partner.balance, f"partner_earning lead#{lead_id}", request.user)
                # Авто-аккредитация для партнёра: баланс был в минусе и перешёл в плюс
                if not partner.is_accredited and _old_pb < 0 and partner.balance >= 0:
                    partner.is_accredited = True
                    partner.save(update_fields=["is_accredited"])
        lead_owner = User.objects.select_for_update().get(pk=lead.user_id)
        from .models import log_balance_change
        if is_dozhim:
            _old = lead_owner.dozhim_balance or 0
            lead_owner.dozhim_balance = _old + reward
            lead_owner.save(update_fields=["dozhim_balance"])
            log_balance_change(lead_owner, "dozhim_balance", _old, lead_owner.dozhim_balance, f"lead_approve#{lead_id} dozhim +{reward}", request.user)
        else:
            _old = lead_owner.balance or 0
            lead_owner.balance = _old + reward
            lead_owner.save(update_fields=["balance"])
            log_balance_change(lead_owner, "balance", _old, lead_owner.balance, f"lead_approve#{lead_id} +{reward}", request.user)
        # Авто-аккредитация: баланс был в минусе и перешёл в плюс
        if not lead_owner.is_accredited and _old < 0 and (lead_owner.balance if not is_dozhim else lead_owner.dozhim_balance) >= 0:
            lead_owner.is_accredited = True
            lead_owner.save(update_fields=["is_accredited"])
        # Сохраняем текущую ставку баланс-админа в логе
        _ba_rate = None
        _ba_user = User.objects.filter(role=User.Role.BALANCE_ADMIN).first()
        if _ba_user:
            _ba_rate = _ba_user.balance_admin_rate
        LeadReviewLog.objects.create(
            lead=lead, admin=request.user,
            action=LeadReviewLog.Action.APPROVED,
            balance_admin_rate_snapshot=_ba_rate,
        )
    msg = f"Лид #{lead_id} одобрен. Пользователю начислено {reward} руб."
    messages.success(request, msg)
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse({"success": True, "message": msg})
    next_url = request.GET.get("next") or request.POST.get("next")
    if next_url:
        from django.utils.http import url_has_allowed_host_and_scheme
        if url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
            return redirect(next_url)
    return redirect("admin_user_leads_list", user_id=user_id)


@login_required
def admin_lead_reject(request: HttpRequest, user_id: int, lead_id: int) -> HttpResponse:
    """Отклонить лид — форма с причиной отклонения. Поддерживает и одобренные: списывает баланс."""
    if not _require_support(request):
        return HttpResponseForbidden("Недостаточно прав.")
    from .models import LeadReviewLog
    lead = get_object_or_404(Lead, pk=lead_id, user_id=user_id)
    if request.method == "POST":
        form = LeadRejectForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                lead_refresh = Lead.objects.select_for_update().select_related("user").get(pk=lead_id, user_id=user_id)
                was_approved = lead_refresh.status == Lead.Status.APPROVED
                lead_refresh.status = Lead.Status.REJECTED
                lead_refresh.rejection_reason = form.cleaned_data["rejection_reason"].strip()
                lead_refresh.rework_comment = ""
                lead_refresh.reviewed_at = timezone.now()
                lead_refresh.reviewed_by = request.user
                lead_refresh.save(update_fields=["status", "rejection_reason", "rework_comment", "reviewed_at", "reviewed_by"])
                if was_approved:
                    _is_dz = lead_refresh.lead_type and lead_refresh.lead_type.slug == "dozhim"
                    reward = getattr(settings, "DOZHIM_APPROVE_REWARD", 40) if _is_dz else getattr(settings, "LEAD_APPROVE_REWARD", 40)
                    # Откат партнёрского заработка (только для лидов поиска)
                    if not _is_dz:
                        from .models import PartnerEarning
                        pe = PartnerEarning.objects.filter(lead=lead_refresh).select_related("partner").first()
                        if pe:
                            partner = User.objects.select_for_update().get(pk=pe.partner_id)
                            if partner.role != User.Role.PARTNER:
                                reward = LEAD_APPROVE_REWARD - pe.amount
                            _old_pb = partner.balance or 0
                            partner.balance = _old_pb - pe.amount
                            partner.save(update_fields=["balance"])
                            log_balance_change(partner, "balance", _old_pb, partner.balance, f"lead_reject#{lead_id} partner_rollback -{pe.amount}", request.user)
                            pe.delete()
                    lead_owner = User.objects.select_for_update().get(pk=lead_refresh.user_id)
                    if _is_dz:
                        _old = lead_owner.dozhim_balance or 0
                        lead_owner.dozhim_balance = _old - reward
                        lead_owner.save(update_fields=["dozhim_balance"])
                        log_balance_change(lead_owner, "dozhim_balance", _old, lead_owner.dozhim_balance, f"lead_reject#{lead_id} dozhim -{reward}", request.user)
                    else:
                        _old = lead_owner.balance or 0
                        lead_owner.balance = _old - reward
                        lead_owner.save(update_fields=["balance"])
                        log_balance_change(lead_owner, "balance", _old, lead_owner.balance, f"lead_reject#{lead_id} -{reward}", request.user)
                LeadReviewLog.objects.create(lead=lead_refresh, admin=request.user, action=LeadReviewLog.Action.REJECTED)
            msg = f"Лид #{lead_id} отклонён." + (" Баланс уменьшен." if was_approved else "")
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": True, "message": msg})
            messages.success(request, msg)
            from django.utils.http import url_has_allowed_host_and_scheme
            next_url = request.GET.get("next") or request.POST.get("next")
            if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
                return redirect(next_url)
            return redirect("admin_user_leads_list", user_id=user_id)
        else:
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": False, "errors": form.errors}, status=400)
    else:
        form = LeadRejectForm()
    return render(
        request,
        "core/admin_lead_reject.html",
        {"lead": lead, "target_user": lead.user, "form": form, "next_url": request.GET.get("next")},
    )


@login_required
def admin_lead_rework(request: HttpRequest, user_id: int, lead_id: int) -> HttpResponse:
    """Отправить лид на доработку — форма с указанием, что доработать. Поддерживает одобренные: списывает баланс."""
    if not _require_support(request):
        return HttpResponseForbidden("Недостаточно прав.")
    from .models import LeadReviewLog
    lead = get_object_or_404(Lead, pk=lead_id, user_id=user_id)
    if request.method == "POST":
        form = LeadReworkForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                lead_refresh = Lead.objects.select_for_update().select_related("user").get(pk=lead_id, user_id=user_id)
                was_approved = lead_refresh.status == Lead.Status.APPROVED
                lead_refresh.status = Lead.Status.REWORK
                lead_refresh.rework_comment = form.cleaned_data["rework_comment"].strip()
                lead_refresh.rejection_reason = ""
                lead_refresh.reviewed_at = timezone.now()
                lead_refresh.reviewed_by = request.user
                lead_refresh.save(update_fields=["status", "rework_comment", "rejection_reason", "reviewed_at", "reviewed_by"])
                if was_approved:
                    _is_dz = lead_refresh.lead_type and lead_refresh.lead_type.slug == "dozhim"
                    reward = getattr(settings, "DOZHIM_APPROVE_REWARD", 40) if _is_dz else getattr(settings, "LEAD_APPROVE_REWARD", 40)
                    if not _is_dz:
                        from .models import PartnerEarning
                        pe = PartnerEarning.objects.filter(lead=lead_refresh).select_related("partner").first()
                        if pe:
                            partner = User.objects.select_for_update().get(pk=pe.partner_id)
                            if partner.role != User.Role.PARTNER:
                                reward = LEAD_APPROVE_REWARD - pe.amount
                            _old_pb = partner.balance or 0
                            partner.balance = _old_pb - pe.amount
                            partner.save(update_fields=["balance"])
                            log_balance_change(partner, "balance", _old_pb, partner.balance, f"lead_rework#{lead_id} partner_rollback", request.user)
                            pe.delete()
                    lead_owner = User.objects.select_for_update().get(pk=lead_refresh.user_id)
                    if _is_dz:
                        _old = lead_owner.dozhim_balance or 0
                        lead_owner.dozhim_balance = _old - reward
                        lead_owner.save(update_fields=["dozhim_balance"])
                        log_balance_change(lead_owner, "dozhim_balance", _old, lead_owner.dozhim_balance, f"lead_rework#{lead_id} dozhim -{reward}", request.user)
                    else:
                        _old = lead_owner.balance or 0
                        lead_owner.balance = _old - reward
                        lead_owner.save(update_fields=["balance"])
                        log_balance_change(lead_owner, "balance", _old, lead_owner.balance, f"lead_rework#{lead_id} -{reward}", request.user)
                LeadReviewLog.objects.create(lead=lead_refresh, admin=request.user, action=LeadReviewLog.Action.REWORK)
            msg = f"Лид #{lead_id} отправлен на доработку." + (" Баланс уменьшен." if was_approved else "")
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": True, "message": msg})
            messages.success(request, msg)
            from django.utils.http import url_has_allowed_host_and_scheme
            next_url = request.GET.get("next") or request.POST.get("next")
            if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
                return redirect(next_url)
            return redirect("admin_user_leads_list", user_id=user_id)
        else:
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": False, "message": "Укажите причину доработки."}, status=400)
    else:
        form = LeadReworkForm()
    return render(
        request,
        "core/admin_lead_rework.html",
        {"lead": lead, "target_user": lead.user, "form": form, "next_url": request.GET.get("next")},
    )


MIME_BY_EXT = {
    "mp4": "video/mp4", "m4v": "video/mp4", "mov": "video/mp4",
    "webm": "video/webm", "3gp": "video/3gpp",
    "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
    "gif": "image/gif", "webp": "image/webp", "bmp": "image/bmp",
}


def _serve_lead_attachment(lead, request=None):
    """Отдаёт вложение лида: прямой S3 URL (JSON для AJAX) или FileResponse с правильным Content‑Type."""
    import logging
    name = lead.attachment.name or ""
    ext = name.rsplit(".", 1)[-1].lower() if "." in name else ""
    content_type = MIME_BY_EXT.get(ext, "application/octet-stream")

    # Если файл в S3 — отдаём прямой URL
    try:
        url = lead.attachment.url
        if url and ("s3" in url or "twc" in url or "timeweb" in url or url.startswith("http")):
            # AJAX-запрос (из видео-плеера) — вернуть JSON с прямым URL
            if request and request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"url": url})
            # Обычный запрос (открыть в новой вкладке / скачать) — redirect
            from django.shortcuts import redirect
            return redirect(url)
    except Exception:
        pass

    try:
        f = lead.attachment.open("rb")
    except Exception as e:
        from django.utils.html import escape as html_escape
        lead_pk = getattr(lead, "id", None) or getattr(lead, "pk", None)
        # WorkerSelfLead has .worker, Lead has .user, fake objects may have .user
        owner = getattr(lead, "user", None) or getattr(lead, "worker", None)
        username = getattr(owner, "username", "unknown") if owner else "unknown"
        logging.getLogger(__name__).warning(
            "Lead attachment open failed: lead_id=%s path=%s err=%s",
            lead_pk, lead.attachment.name, e,
        )
        html = (
            "<!DOCTYPE html><html><head><meta charset='utf-8'><title>Файл недоступен</title></head><body style='font-family:sans-serif;padding:2rem;'>"
            "<h1>Файл не найден</h1>"
            "<p>Вложение этого отчёта в базе есть, но файл на сервере отсутствует.</p>"
            "<p><strong>Что сделать:</strong> попросите пользователя <strong>%s</strong> заново загрузить видео через «Мои лиды» → доработка отчёта.</p>"
            "<p><a href='javascript:history.back()'>← Назад</a></p></body></html>"
        ) % (html_escape(username),)
        return HttpResponse(html, status=404, content_type="text/html; charset=utf-8")
    filename = name.split("/")[-1] if name else "attachment"
    response = FileResponse(f, content_type=content_type)
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


@login_required
def admin_lead_attachment(request: HttpRequest, user_id: int, lead_id: int) -> HttpResponse:
    """Отдаёт вложение лида (фото/видео). Доступ: staff или владелец лида."""
    lead = get_object_or_404(Lead, pk=lead_id, user_id=user_id)
    if not lead.attachment:
        return HttpResponseForbidden("У этого лида нет вложения.")
    if not _require_support(request) and request.user.id != lead.user_id:
        return HttpResponseForbidden("Нет доступа к этому файлу.")
    return _serve_lead_attachment(lead, request=request)


@login_required
def standalone_admin_lead_attachment(request: HttpRequest, lead_id: int) -> HttpResponse:
    """Отдаёт вложение лида для самостоятельного админа (только одобренные СС-лиды)."""
    if not _require_standalone_admin(request):
        return HttpResponseForbidden("Недостаточно прав.")
    lead = get_object_or_404(
        Lead,
        pk=lead_id,
        needs_team_contact=True,
        status=Lead.Status.APPROVED,
    )
    if not lead.attachment:
        return HttpResponseForbidden("У этого лида нет вложения.")
    return _serve_lead_attachment(lead, request=request)


@login_required
def admin_user_leads_export(request: HttpRequest, user_id: int, period: str) -> HttpResponse:
    """Выгрузка лидов пользователя за сегодня, вчера или все (Excel). В колонке «Скриншот» — ссылка на файл."""
    if not _require_support(request):
        return HttpResponseForbidden("Недостаточно прав.")
    if period not in ("today", "yesterday", "all"):
        return HttpResponseForbidden("Недопустимый период.")
    target_user = get_object_or_404(User, pk=user_id)
    leads_qs = Lead.objects.filter(user=target_user).select_related("lead_type", "base_type").order_by("-created_at")
    if period in ("today", "yesterday"):
        today_start, today_end, yesterday_start, yesterday_end = _day_bounds_lead_stats()
        if period == "today":
            start, end = today_start, today_end
        else:
            start, end = yesterday_start, yesterday_end
        leads_qs = leads_qs.filter(created_at__gte=start, created_at__lt=end)
    leads = list(leads_qs)
    wb = Workbook()
    ws = wb.active
    ws.title = "Лиды"
    ws.append(
        [
            "ID",
            "Пользователь",
            "Тип лида",
            "Тип базы",
            "Контакт (raw)",
            "Источник",
            "Комментарий",
            "Создан",
            "Скриншот (ссылка)",
        ]
    )
    for lead in leads:
        screenshot_url = ""
        if lead.attachment:
            screenshot_url = request.build_absolute_uri(
                reverse("admin_lead_attachment", args=[lead.user_id, lead.id])
            )
        ws.append(
            [
                lead.id,
                lead.user.username,
                lead.lead_type.name if lead.lead_type else "",
                lead.base_type.name if lead.base_type else "",
                lead.raw_contact,
                lead.source,
                lead.comment,
                lead.created_at.isoformat(),
                screenshot_url,
            ]
        )
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = f"leads_{target_user.username}_{period}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def admin_user_limits(request: HttpRequest, user_id: int) -> HttpResponse:
    """Редирект: окошко выдачи лимитов убрано, используйте «Выдача контактов по запросу»."""
    if not _require_support(request):
        return HttpResponseForbidden("Недостаточно прав.")
    return redirect("admin_all_users")


@login_required
def admin_user_balance(request: HttpRequest, user_id: int) -> HttpResponse:
    """Начисление или списание рублей пользователю (баланс). Только main_admin."""
    if getattr(request.user, "role", None) != "main_admin":
        return HttpResponseForbidden("Только для главного админа.")
    target_user = get_object_or_404(User, pk=user_id)
    if request.method == "POST":
        try:
            amount = int(request.POST.get("amount") or 0)
            action = request.POST.get("action")
        except (TypeError, ValueError):
            amount = 0
            action = None
        if amount > 0 and action in ("add", "subtract"):
            dept = request.POST.get("dept", "search")
            with transaction.atomic():
                user_locked = User.objects.select_for_update().get(pk=user_id)
                if dept == "dozhim":
                    current = user_locked.dozhim_balance or 0
                    if action == "add":
                        user_locked.dozhim_balance = current + amount
                    else:
                        user_locked.dozhim_balance = current - amount
                    user_locked.save(update_fields=["dozhim_balance"])
                    new_bal = user_locked.dozhim_balance
                else:
                    current = user_locked.balance or 0
                    if action == "add":
                        user_locked.balance = current + amount
                    else:
                        user_locked.balance = current - amount
                    user_locked.save(update_fields=["balance"])
                    new_bal = user_locked.balance
            dept_label = "Дожим" if dept == "dozhim" else "Поиск"
            msg = f"Начислено {amount} руб." if action == "add" else f"Списано {amount} руб."
            messages.success(request, f"[{dept_label}] @{target_user.username}: {msg}. Баланс: {new_bal} руб.")
            return redirect("admin_all_users")
        messages.warning(request, "Укажите положительное число и действие (начислить / списать).")
    return render(
        request,
        "core/admin_user_balance.html",
        {"target_user": target_user},
    )


@login_required
def admin_user_search(request: HttpRequest) -> HttpResponse:
    """Поиск пользователей по нику (или части) — для выпадающего списка."""
    if not _require_support(request):
        return HttpResponseForbidden("Недостаточно прав.")
    q = (request.GET.get("q") or "").strip().lstrip("@")[:50]
    if not q:
        return JsonResponse({"users": []})
    users = (
        User.objects.filter(username__icontains=q)
        .values("id", "username")[:15]
        .order_by("username")
    )
    return JsonResponse({"users": list(users)})


@login_required
def admin_all_users(request: HttpRequest) -> HttpResponse:
    """Список всех пользователей сайта с вкладками: все, активные за сегодня, активные за вчера."""
    if not _require_support(request) and getattr(request.user, "role", None) != User.Role.BALANCE_ADMIN:
        return HttpResponseForbidden("Недостаточно прав.")
    show = request.GET.get("show", "all")
    today_start, today_end, yesterday_start, yesterday_end = _day_bounds_lead_stats()
    total_users_count = User.objects.count()
    active_today_count = User.objects.filter(
        leads__created_at__gte=today_start, leads__created_at__lt=today_end
    ).distinct().count()
    active_yesterday_count = User.objects.filter(
        leads__created_at__gte=yesterday_start,
        leads__created_at__lt=yesterday_end,
    ).distinct().count()
    if show == "today":
        users_list = (
            User.objects.filter(
                leads__created_at__gte=today_start, leads__created_at__lt=today_end
            )
            .distinct()
            .order_by("-date_joined")
        )
    elif show == "yesterday":
        users_list = (
            User.objects.filter(
                leads__created_at__gte=yesterday_start,
                leads__created_at__lt=yesterday_end,
            )
            .distinct()
            .order_by("-date_joined")
        )
    else:
        users_list = User.objects.all().order_by("-date_joined")
    total_balance = (User.objects.aggregate(s=Sum("balance"))["s"] or 0) + (User.objects.aggregate(s=Sum("dozhim_balance"))["s"] or 0)
    return render(
        request,
        "core/admin_all_users.html",
        {
            "users_list": users_list,
            "show": show,
            "total_users_count": total_users_count,
            "active_today_count": active_today_count,
            "active_yesterday_count": active_yesterday_count,
            "total_balance": total_balance,
        },
    )


@login_required
def admin_stats(request: HttpRequest) -> HttpResponse:
    if not _require_support(request):
        return HttpResponseForbidden("Недостаточно прав.")

    # Статистика по базам (один запрос с аннотацией вместо 2N запросов)
    base_stats = [
        {
            "base": bt,
            "total": bt._total,
            "free": bt._free,
            "issued": bt._total - bt._free,
        }
        for bt in BaseType.objects.annotate(
            _total=Count("contacts"),
            _free=Count("contacts", filter=Q(contacts__assigned_to__isnull=True, contacts__is_active=True)),
        ).order_by("order")
    ]

    # Статистика по лидам: за неделю, месяц, всё время
    now = timezone.now()
    week_ago = now - timedelta(days=7)
    month_ago = now - timedelta(days=30)

    total_leads_week = Lead.objects.filter(
        status=Lead.Status.APPROVED, created_at__gte=week_ago
    ).count()
    total_leads_month = Lead.objects.filter(
        status=Lead.Status.APPROVED, created_at__gte=month_ago
    ).count()
    total_leads_all = Lead.objects.filter(status=Lead.Status.APPROVED).count()

    q_week = Lead.objects.filter(
        status=Lead.Status.APPROVED, created_at__gte=week_ago
    ).values("lead_type__name").annotate(c=Count("id"))
    q_month = Lead.objects.filter(
        status=Lead.Status.APPROVED, created_at__gte=month_ago
    ).values("lead_type__name").annotate(c=Count("id"))
    q_all = Lead.objects.filter(status=Lead.Status.APPROVED).values("lead_type__name").annotate(c=Count("id"))

    leads_week = {x["lead_type__name"] or "Без категории": x["c"] for x in q_week}
    leads_month = {x["lead_type__name"] or "Без категории": x["c"] for x in q_month}
    leads_all = {x["lead_type__name"] or "Без категории": x["c"] for x in q_all}

    # Для админской статистики показываем все типы лидов, включая «Самостоятельные лиды»,
    # чтобы не терять уже существующие данные.
    type_names = list(LeadType.objects.values_list("name", flat=True).order_by("name"))
    lead_type_stats = [
        {
            "name": n,
            "week": leads_week.get(n, 0),
            "month": leads_month.get(n, 0),
            "all": leads_all.get(n, 0),
        }
        for n in type_names
    ]
    # Категории, которые есть в лидах, но нет в LeadType (на всякий случай)
    for name in set(leads_week.keys()) | set(leads_month.keys()) | set(leads_all.keys()):
        if name not in type_names:
            lead_type_stats.append({
                "name": name,
                "week": leads_week.get(name, 0),
                "month": leads_month.get(name, 0),
                "all": leads_all.get(name, 0),
            })

    # Кто проверял лиды: сводка по админам (одобрено / отклонено / на доработку)
    from .models import LeadReviewLog
    reviewed_qs = (
        LeadReviewLog.objects.values("admin__username", "admin__id")
        .annotate(
            approved=Count("id", filter=Q(action=LeadReviewLog.Action.APPROVED)),
            rejected=Count("id", filter=Q(action=LeadReviewLog.Action.REJECTED)),
            rework=Count("id", filter=Q(action=LeadReviewLog.Action.REWORK)),
        )
    )
    reviewed_by_stats = [
        {
            "username": x["admin__username"] or "—",
            "user_id": x["admin__id"],
            "approved": x["approved"],
            "rejected": x["rejected"],
            "rework": x["rework"],
            "total": x["approved"] + x["rejected"] + x["rework"],
        }
        for x in reviewed_qs.order_by("-approved", "-rejected")
    ]

    return render(
        request,
        "core/admin_stats.html",
        {
            "base_stats": base_stats,
            "lead_type_stats": lead_type_stats,
            "total_leads_week": total_leads_week,
            "total_leads_month": total_leads_month,
            "total_leads_all": total_leads_all,
            "reviewed_by_stats": reviewed_by_stats,
        },
    )


def _allocate_contacts_to_user(target_user, base_type, count: int) -> int:
    """Выдаёт пользователю до count контактов из базы. Возвращает количество выданных."""
    from django.db import transaction

    with transaction.atomic():
        free_qs = (
            Contact.objects.select_for_update()
            .filter(base_type=base_type, assigned_to__isnull=True, is_active=True)
            .order_by("id")[:count]
        )
        contacts_to_give = list(free_qs)
        if not contacts_to_give:
            return 0
        now = timezone.now()
        for c in contacts_to_give:
            c.assigned_to = target_user
            c.assigned_at = now
            c.save(update_fields=["assigned_to", "assigned_at", "updated_at"])
        return len(contacts_to_give)


@login_required
def admin_contact_requests(request: HttpRequest) -> HttpResponse:
    """Список заявок на контакты. «Выдать контакты» — сбрасывает лимит (добавляет доп. лимит), пользователь сам нажимает «Получить контакты»."""
    if not _require_support_or_partner(request):
        return HttpResponseForbidden("Недостаточно прав.")
    if request.method == "POST" and request.POST.get("action") == "refresh":
        return redirect("admin_contact_requests")
    if request.method == "POST":
        req_id = request.POST.get("request_id")
        if req_id:
            req = get_object_or_404(ContactRequest, pk=req_id, status="pending")
            target_user = req.user
            bases_to_give = [req.base_type] if req.base_type else list(BaseType.objects.all().order_by("order"))
            for base in bases_to_give:
                obj, _ = UserBaseLimit.objects.get_or_create(
                    user=target_user, base_type=base, defaults={"extra_daily_limit": 0}
                )
                obj.extra_daily_limit += base.default_daily_limit
                obj.save(update_fields=["extra_daily_limit"])
            req.status = "resolved"
            req.resolved_at = timezone.now()
            req.resolved_by = request.user
            req.save(update_fields=["status", "resolved_at", "resolved_by", "updated_at"])
            messages.success(
                request,
                f"@{target_user.username}: добавлен доп. лимит. Пользователь может нажать «Получить контакты» на странице контактов.",
            )
            return redirect("admin_contact_requests")
    pending = ContactRequest.objects.filter(status="pending").select_related("user", "base_type").order_by("-created_at")
    return render(
        request,
        "core/admin_contact_requests.html",
        {"pending_requests": pending},
    )


@login_required
def balance_admin_contact_requests(request: HttpRequest) -> HttpResponse:
    """Выдача контактов баланс-админом — та же логика, что у обычных админов."""
    from .views import _is_balance_admin
    if not _is_balance_admin(request.user):
        return HttpResponseForbidden("Недостаточно прав.")
    if request.method == "POST" and request.POST.get("action") == "refresh":
        return redirect("balance_admin_contact_requests")
    if request.method == "POST":
        req_id = request.POST.get("request_id")
        if req_id:
            req = get_object_or_404(ContactRequest, pk=req_id, status="pending")
            target_user = req.user
            bases_to_give = [req.base_type] if req.base_type else list(BaseType.objects.all().order_by("order"))
            for base in bases_to_give:
                obj, _ = UserBaseLimit.objects.get_or_create(
                    user=target_user, base_type=base, defaults={"extra_daily_limit": 0}
                )
                obj.extra_daily_limit += base.default_daily_limit
                obj.save(update_fields=["extra_daily_limit"])
            req.status = "resolved"
            req.resolved_at = timezone.now()
            req.resolved_by = request.user
            req.save(update_fields=["status", "resolved_at", "resolved_by", "updated_at"])
            messages.success(
                request,
                f"@{target_user.username}: добавлен доп. лимит. Пользователь может нажать «Получить контакты».",
            )
            return redirect("balance_admin_contact_requests")
    pending = ContactRequest.objects.filter(status="pending").select_related("user", "base_type").order_by("-created_at")
    return render(
        request,
        "core/balance_admin_contact_requests.html",
        {"pending_requests": pending},
    )


@login_required
def admin_withdrawal_requests(request: HttpRequest) -> HttpResponse:
    """Список заявок на вывод. Только main_admin."""
    if getattr(request.user, "role", None) != "main_admin":
        return HttpResponseForbidden("Только для главного админа.")
    if request.method == "POST":
        action = request.POST.get("action")
        # ── одиночное действие ──────────────────────────────────────────
        req_id = request.POST.get("request_id")
        if req_id and action in ("approve", "reject"):
            with transaction.atomic():
                wreq = (
                    WithdrawalRequest.objects.select_for_update()
                    .select_related("user")
                    .filter(pk=req_id, status="pending")
                    .first()
                )
                if not wreq:
                    messages.warning(request, "Заявка уже обработана или не найдена.")
                    return redirect("admin_withdrawal_requests")
                now = timezone.now()
                if action == "approve":
                    wreq.status = "approved"
                    messages.success(
                        request,
                        f"Вывод @{wreq.user.username} на {wreq.amount} руб. одобрен.",
                    )
                else:
                    _is_dozhim_wr = wreq.payout_details and wreq.payout_details.startswith("[Дожим]")
                    if _is_dozhim_wr:
                        _old = wreq.user.dozhim_balance or 0
                        wreq.user.dozhim_balance = _old + wreq.amount
                        wreq.user.save(update_fields=["dozhim_balance"])
                        log_balance_change(wreq.user, "dozhim_balance", _old, wreq.user.dozhim_balance, f"withdrawal_reject#{wreq.pk} +{wreq.amount}", request.user)
                    else:
                        _old = wreq.user.balance or 0
                        wreq.user.balance = _old + wreq.amount
                        wreq.user.save(update_fields=["balance"])
                        log_balance_change(wreq.user, "balance", _old, wreq.user.balance, f"withdrawal_reject#{wreq.pk} +{wreq.amount}", request.user)
                    wreq.status = "rejected"
                    messages.info(request, f"Заявка от @{wreq.user.username} отклонена. Баланс восстановлен.")
                wreq.processed_at = now
                wreq.processed_by = request.user
                wreq.save(update_fields=["status", "processed_at", "processed_by"])
            return redirect("admin_withdrawal_requests")
        # ── массовое действие ───────────────────────────────────────────
        bulk_ids = request.POST.getlist("bulk_ids")
        if bulk_ids and action in ("bulk_approve", "bulk_reject"):
            bulk_ids = [i for i in bulk_ids if i.isdigit()]
            approved_count = rejected_count = 0
            now = timezone.now()
            with transaction.atomic():
                wreqs = (
                    WithdrawalRequest.objects.select_for_update()
                    .select_related("user")
                    .filter(pk__in=bulk_ids, status="pending")
                )
                for wreq in wreqs:
                    if action == "bulk_approve":
                        wreq.status = "approved"
                        approved_count += 1
                    else:
                        _is_dozhim_wr = wreq.payout_details and wreq.payout_details.startswith("[Дожим]")
                        if _is_dozhim_wr:
                            wreq.user.dozhim_balance = (wreq.user.dozhim_balance or 0) + wreq.amount
                            wreq.user.save(update_fields=["dozhim_balance"])
                        else:
                            wreq.user.balance = (wreq.user.balance or 0) + wreq.amount
                            wreq.user.save(update_fields=["balance"])
                        wreq.status = "rejected"
                        rejected_count += 1
                    wreq.processed_at = now
                    wreq.processed_by = request.user
                    wreq.save(update_fields=["status", "processed_at", "processed_by"])
            if approved_count:
                messages.success(request, f"Одобрено заявок: {approved_count}.")
            if rejected_count:
                messages.info(request, f"Отклонено заявок: {rejected_count}. Балансы восстановлены.")
            return redirect("admin_withdrawal_requests")
    all_pending = WithdrawalRequest.objects.filter(status="pending").select_related("user").order_by("created_at")
    pending_regular = [r for r in all_pending if not r.user.partner_owner_id]
    pending_referrals = [r for r in all_pending if r.user.partner_owner_id]
    history = (
        WithdrawalRequest.objects.exclude(status="pending")
        .select_related("user", "processed_by")
        .order_by("-created_at")[:200]
    )
    total_user_withdrawals = (
        WithdrawalRequest.objects.filter(status="approved").aggregate(s=Sum("amount"))["s"] or 0
    )
    from .models import WorkerWithdrawalRequest
    total_worker_withdrawals = (
        WorkerWithdrawalRequest.objects.filter(status="approved").aggregate(s=Sum("amount"))["s"] or 0
    )
    return render(
        request,
        "core/admin_withdrawal_requests.html",
        {
            "pending_regular": pending_regular,
            "pending_referrals": pending_referrals,
            "pending_requests": list(all_pending),
            "history_requests": history,
            "total_user_withdrawals": total_user_withdrawals,
            "total_worker_withdrawals": total_worker_withdrawals,
            "total_approved_withdrawals": total_user_withdrawals + total_worker_withdrawals,
        },
    )


# Значения первой строки/заголовка — не считаем контактом (как в боте)
EXCEL_HEADER_VALUES = frozenset(("value", "значение", "контакт", "данные"))


def _excel_contact_value(cell_value) -> str | None:
    """Возвращает значение контакта из ячейки или None (пусто/заголовок)."""
    if cell_value is None:
        return None
    value = str(cell_value).strip()
    if not value:
        return None
    if value.lower() in EXCEL_HEADER_VALUES:
        return None
    return value


def _excel_row_is_assigned(row: tuple) -> bool:
    """Строка «отработана»: в колонках ID/Username/Date (справа от Value) есть данные."""
    if not row or len(row) < 2:
        return False
    for i in (1, 2, 3):
        if i < len(row) and row[i] is not None and str(row[i]).strip():
            return True
    return False


BULK_CREATE_BATCH_SIZE = 1000

# Телефонные базы: загрузка в одну → автоматическая репликация во все остальные
PHONE_BASE_SLUGS = ("whatsapp", "max", "viber")


def _replicate_to_phone_bases(values: list[str], source_slug: str) -> dict[str, int]:
    """Реплицирует контакты из телефонной базы во все остальные телефонные базы.
    Возвращает {slug: кол-во созданных} для каждой целевой базы."""
    if source_slug not in PHONE_BASE_SLUGS or not values:
        return {}
    result = {}
    for slug in PHONE_BASE_SLUGS:
        if slug == source_slug:
            continue
        try:
            bt = BaseType.objects.get(slug=slug)
        except BaseType.DoesNotExist:
            continue
        count_before = Contact.objects.filter(base_type=bt).count()
        for i in range(0, len(values), BULK_CREATE_BATCH_SIZE):
            chunk = values[i : i + BULK_CREATE_BATCH_SIZE]
            Contact.objects.bulk_create(
                [Contact(base_type=bt, value=v) for v in chunk],
                ignore_conflicts=True,
            )
        count_after = Contact.objects.filter(base_type=bt).count()
        result[slug] = count_after - count_before
    return result
# Лимит строк на одну загрузку, чтобы не превышать таймаут воркера (gunicorn)
MAX_UPLOAD_ROWS = 40_000

EXCEL_SHEET_MAP = {
    # Короткие названия
    "Тг": "telegram",
    "ТГ": "telegram",
    "Вотсап": "whatsapp",
    "Макс": "max",
    "Вайбер": "viber",
    "Инст": "instagram",
    "ВК": "vk",
    "ВКонтакте": "vk",
    "Вконтакте": "vk",
    "вконтакте": "vk",
    "Ок": "ok",
    "Почта": "email",
    # Полные/латинские
    "Telegram": "telegram",
    "telegram": "telegram",
    "WhatsApp": "whatsapp",
    "Whatsapp": "whatsapp",
    "whatsapp": "whatsapp",
    "Max": "max",
    "max": "max",
    "Viber": "viber",
    "viber": "viber",
    "Нельзяграм": "instagram",
    "Нельзяграм (там где Reels)": "instagram",
    "Instagram": "instagram",
    "instagram": "instagram",
    "VK": "vk",
    "Одноклассники": "ok",
    "одноклассники": "ok",
    "OK": "ok",
    "Ok": "ok",
    "Email": "email",
    "email": "email",
    "Почты": "email",
}


def _run_bases_import_background(file_path: str, job_id: int) -> None:
    """Выполняет импорт всех листов в фоне. Обновляет BasesImportJob по завершении. Удаляет файл."""
    try:
        wb = load_workbook(file_path, read_only=True)
        created, skipped, details = _process_excel_all_sheets(wb, max_rows=None)
        wb.close()
        msg = f"Добавлено контактов: {created}, пропущено (дубликаты): {skipped}.\n\n" + "\n".join(details)
        BasesImportJob.objects.filter(pk=job_id).update(status=BasesImportJob.Status.SUCCESS, message=msg)
    except Exception as e:
        BasesImportJob.objects.filter(pk=job_id).update(
            status=BasesImportJob.Status.ERROR,
            message=f"Ошибка: {e}",
        )
    finally:
        try:
            if os.path.isfile(file_path):
                os.remove(file_path)
        except OSError:
            pass


@login_required
def upload_bases_excel(request: HttpRequest) -> HttpResponse:
    """Редирект на единую страницу загрузки/выгрузки баз."""
    if not _require_support_or_partner(request):
        return HttpResponseForbidden("Недостаточно прав.")
    return redirect("bases_excel")


def _process_excel_all_sheets(wb, max_rows: int | None = MAX_UPLOAD_ROWS) -> tuple[int, int, list]:
    """Обработка книги Excel: все листы по EXCEL_SHEET_MAP. Свободные контакты (без ID/User/Date).
    max_rows: лимит строк (защита от таймаута при синхронной загрузке); None — без лимита (фоновый импорт)."""
    total_created = 0
    total_skipped = 0
    details = []
    total_rows_processed = 0
    limit_reached = False
    for sheet in wb.worksheets:
        if limit_reached:
            details.append(f"Лист «{sheet.title}» — пропущен (достигнут лимит {max_rows} строк)")
            continue
        base_slug = EXCEL_SHEET_MAP.get(sheet.title)
        if not base_slug:
            details.append(f"Лист «{sheet.title}» — неизвестный тип, пропущен")
            continue
        try:
            base_type = BaseType.objects.get(slug=base_slug)
        except BaseType.DoesNotExist:
            details.append(f"Лист «{sheet.title}» — база не найдена")
            continue
        free_values = []
        for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True):
            value = _excel_contact_value(row[0] if row else None)
            if not value:
                continue
            if _excel_row_is_assigned(row):
                continue
            free_values.append(value)
        if max_rows is not None:
            remaining = max_rows - total_rows_processed
            if remaining <= 0:
                limit_reached = True
                details.append(f"Лист «{sheet.title}» — пропущен (достигнут лимит {max_rows} строк)")
                continue
            to_process = free_values[:remaining]
            skipped_by_limit = len(free_values) - len(to_process)
        else:
            to_process = free_values
            skipped_by_limit = 0
        count_before = Contact.objects.filter(base_type=base_type).count()
        for i in range(0, len(to_process), BULK_CREATE_BATCH_SIZE):
            chunk = to_process[i : i + BULK_CREATE_BATCH_SIZE]
            Contact.objects.bulk_create(
                [Contact(base_type=base_type, value=v) for v in chunk],
                ignore_conflicts=True,
            )
        count_after = Contact.objects.filter(base_type=base_type).count()
        sheet_created = count_after - count_before
        sheet_skipped = len(to_process) - sheet_created
        total_created += sheet_created
        total_skipped += sheet_skipped
        total_rows_processed += len(to_process)
        if max_rows is not None and skipped_by_limit > 0:
            limit_reached = True
            details.append(
                f"«{base_type.name}» — обработано {len(to_process)} из {len(free_values)} (лимит {max_rows}), "
                f"добавлено {sheet_created}, дубликатов {sheet_skipped}. Остальные листы не загружены."
            )
        else:
            details.append(
                f"«{base_type.name}» — свободных строк {len(free_values)}, добавлено {sheet_created}, дубликатов {sheet_skipped}"
            )
        # Репликация телефонных номеров во все телефонные базы
        replicated = _replicate_to_phone_bases(to_process, base_slug)
        for rslug, rcount in replicated.items():
            if rcount > 0:
                total_created += rcount
                details.append(f"  ↳ репликация → «{rslug}»: добавлено {rcount}")
        if skipped_by_limit > 0:
            break
    return total_created, total_skipped, details


def _process_excel_single_sheet(wb, base_type: BaseType) -> tuple[int, int]:
    """Обработка книги Excel: первый лист, первый столбец. Только свободные контакты (без ID/User/Date)."""
    ws = wb.active
    free_values = []
    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        value = _excel_contact_value(row[0] if row else None)
        if not value:
            continue
        if _excel_row_is_assigned(row):
            continue
        free_values.append(value)
    if len(free_values) > MAX_UPLOAD_ROWS:
        raise ValueError(
            f"В файле слишком много строк: {len(free_values)}. "
            f"Максимум за одну загрузку: {MAX_UPLOAD_ROWS}. Разбейте файл на части или загрузите «все листы»."
        )
    count_before = Contact.objects.filter(base_type=base_type).count()
    for i in range(0, len(free_values), BULK_CREATE_BATCH_SIZE):
        chunk = free_values[i : i + BULK_CREATE_BATCH_SIZE]
        Contact.objects.bulk_create(
            [Contact(base_type=base_type, value=v) for v in chunk],
            ignore_conflicts=True,
        )
    count_after = Contact.objects.filter(base_type=base_type).count()
    created = count_after - count_before
    return created, len(free_values) - (count_after - count_before)


@login_required
def bases_excel(request: HttpRequest) -> HttpResponse:
    """Одна страница: загрузка по категории, загрузка всех листов (шаблон бота), выгрузка."""
    if not _require_support_or_partner(request):
        return HttpResponseForbidden("Недостаточно прав.")

    base_types = list(BaseType.objects.all().order_by("order"))
    form_all = BaseExcelUploadForm()
    form_category = BaseCategoryUploadForm()

    if request.method == "POST":
        if "upload_all" in request.POST:
            form_all = BaseExcelUploadForm(request.POST, request.FILES)
            if form_all.is_valid():
                file = form_all.cleaned_data["file"]
                if not (file and file.name and file.name.lower().endswith(".xlsx")):
                    messages.error(request, "Нужен файл в формате .xlsx")
                else:
                    # Фоновый импорт без лимита строк: сохраняем файл и запускаем поток
                    import_dir = settings.MEDIA_ROOT / "imports"
                    import_dir.mkdir(parents=True, exist_ok=True)
                    safe_name = f"{uuid.uuid4().hex}.xlsx"
                    file_path = import_dir / safe_name
                    with open(file_path, "wb") as f:
                        for chunk in file.chunks():
                            f.write(chunk)
                    job = BasesImportJob.objects.create(
                        status=BasesImportJob.Status.RUNNING,
                        started_by=request.user,
                    )
                    thread = threading.Thread(
                        target=_run_bases_import_background,
                        args=(str(file_path), job.pk),
                        daemon=True,
                    )
                    thread.start()
                    messages.success(
                        request,
                        "Импорт запущен в фоне. Файл будет обработан полностью (без лимита строк). "
                        "Обновите страницу через несколько минут, чтобы увидеть результат.",
                    )
                    return redirect("bases_excel")
        elif "upload_category" in request.POST:
            form_category = BaseCategoryUploadForm(request.POST, request.FILES)
            if form_category.is_valid():
                base_type = form_category.cleaned_data["base_type"]
                file = form_category.cleaned_data["file"]
                if not (file and file.name and file.name.lower().endswith(".xlsx")):
                    messages.error(request, "Нужен файл в формате .xlsx")
                elif base_type == BaseCategoryUploadForm.PHONE_BASES_VALUE:
                    # Загрузка номеров сразу во все телефонные базы
                    try:
                        wb = load_workbook(file, read_only=True)
                        total_created = 0
                        total_skipped = 0
                        for slug in PHONE_BASE_SLUGS:
                            bt = BaseType.objects.get(slug=slug)
                            created, skipped = _process_excel_single_sheet(wb, bt)
                            total_created += created
                            total_skipped += skipped
                        wb.close()
                        messages.success(
                            request,
                            f"Номера загружены в WhatsApp, Max, Viber: добавлено {total_created}, дубликатов {total_skipped}.",
                        )
                        return redirect("bases_excel")
                    except Exception as e:
                        messages.error(request, f"Ошибка при обработке файла: {e}")
                else:
                    try:
                        wb = load_workbook(file, read_only=True)
                        created, skipped = _process_excel_single_sheet(wb, base_type)
                        wb.close()
                        messages.success(
                            request,
                            f"База «{base_type.name}»: добавлено {created} контактов, пропущено (дубликаты) {skipped}.",
                        )
                        return redirect("bases_excel")
                    except Exception as e:
                        messages.error(request, f"Ошибка при обработке файла: {e}")

    sheet_names_hint = "Тг, Вотсап, Макс, Вайбер, Инст, ВК, Ок, Почта (или Telegram, WhatsApp и т.д.)"
    latest_import_job = BasesImportJob.objects.order_by("-created_at").first()
    return render(
        request,
        "core/bases_excel.html",
        {
            "form_all": form_all,
            "form_category": form_category,
            "base_types": base_types,
            "sheet_names_hint": sheet_names_hint,
            "latest_import_job": latest_import_job,
        },
    )


def _make_bases_excel_response(wb, filename: str) -> HttpResponse:
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def download_bases_excel(request: HttpRequest) -> HttpResponse:
    """Выгрузка всех баз контактов в один Excel (по листам)."""

    if not _require_support_or_partner(request):
        return HttpResponseForbidden("Недостаточно прав.")

    wb = Workbook()
    first_sheet = True

    for base in BaseType.objects.all().order_by("order"):
        contacts = Contact.objects.filter(base_type=base).select_related("assigned_to").order_by("id")

        if first_sheet:
            ws = wb.active
            ws.title = base.name[:31]
            first_sheet = False
        else:
            ws = wb.create_sheet(title=base.name[:31])

        ws.append(["Value", "User", "Assigned at"])
        for c in contacts:
            ws.append(
                [
                    c.value,
                    c.assigned_to.username if c.assigned_to else "",
                    c.assigned_at.isoformat() if c.assigned_at else "",
                ]
            )

    return _make_bases_excel_response(wb, "bases.xlsx")


@login_required
def download_bases_excel_category(request: HttpRequest, base_type_id: int) -> HttpResponse:
    """Выгрузка одной базы контактов по категории."""

    if not _require_support_or_partner(request):
        return HttpResponseForbidden("Недостаточно прав.")

    base_type = get_object_or_404(BaseType, pk=base_type_id)
    contacts = Contact.objects.filter(base_type=base_type).select_related("assigned_to").order_by("id")

    wb = Workbook()
    ws = wb.active
    ws.title = base_type.name[:31]
    ws.append(["Value", "User", "Assigned at"])
    for c in contacts:
        ws.append(
            [
                c.value,
                c.assigned_to.username if c.assigned_to else "",
                c.assigned_at.isoformat() if c.assigned_at else "",
            ]
        )

    safe_name = base_type.slug or "base"
    return _make_bases_excel_response(wb, f"bases_{safe_name}.xlsx")


@login_required
def download_leads_excel(request: HttpRequest) -> HttpResponse:
    """Выгрузка всех лидов в один Excel."""

    if not _require_support_or_partner(request):
        return HttpResponseForbidden("Недостаточно прав.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Лиды"

    ws.append(
        [
            "ID",
            "Пользователь",
            "Тип лида",
            "Тип базы",
            "Контакт (raw)",
            "Источник",
            "Комментарий",
            "Создан",
        ]
    )

    leads = (
        Lead.objects.select_related("user", "lead_type", "base_type")
        .all()
        .order_by("-created_at")
    )

    for lead in leads:
        ws.append(
            [
                lead.id,
                lead.user.username,
                lead.lead_type.name if lead.lead_type else "",
                lead.base_type.name if lead.base_type else "",
                lead.raw_contact,
                lead.source,
                lead.comment,
                lead.created_at.isoformat(),
            ]
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="leads.xlsx"'
    return response


@login_required
def admin_site_settings(request: HttpRequest) -> HttpResponse:
    """Настройки сайта: пример видео-отчёта и другие параметры. Только main_admin."""
    if getattr(request.user, "role", None) != "main_admin":
        return HttpResponseForbidden("Только для главного админа.")
    
    try:
        site_settings = SiteSettings.get_settings()
    except Exception as e:
        messages.error(request, f"Не удалось загрузить настройки. Возможно, нужно выполнить миграцию: python manage.py migrate core. Ошибка: {e}")
        return redirect("dashboard")
    
    if request.method == "POST":
        description = request.POST.get("example_video_description", "").strip()
        if description:
            site_settings.example_video_description = description
        
        if "example_video" in request.FILES:
            video_file = request.FILES["example_video"]
            allowed_ext = ("mp4", "mov", "webm", "m4v")
            ext = video_file.name.rsplit(".", 1)[-1].lower() if "." in video_file.name else ""
            if ext not in allowed_ext:
                messages.error(request, f"Разрешены только видео: {', '.join(allowed_ext)}")
            elif video_file.size > 100 * 1024 * 1024:
                messages.error(request, "Максимальный размер видео — 100 МБ.")
            else:
                if site_settings.example_video:
                    try:
                        site_settings.example_video.delete(save=False)
                    except Exception:
                        pass
                site_settings.example_video = video_file
                messages.success(request, "Видео-пример успешно загружено.")
        
        if request.POST.get("delete_video") == "1" and site_settings.example_video:
            try:
                site_settings.example_video.delete(save=False)
            except Exception:
                pass
            site_settings.example_video = None
            messages.success(request, "Видео-пример удалено.")
        
        site_settings.save()
        return redirect("admin_site_settings")
    
    return render(
        request,
        "core/admin_site_settings.html",
        {"site_settings": site_settings},
    )


# ──────────────────────────────────────────────────────────────
# SS Admin: Worker Sub-System
# ──────────────────────────────────────────────────────────────

def _serve_worker_report_attachment(report, request=None) -> HttpResponse:
    """Отдаёт вложение к отчёту воркера (аналог _serve_lead_attachment)."""

    class _Proxy:
        pass

    obj = _Proxy()
    obj.attachment = report.attachment
    obj.user = report.worker
    obj.pk = report.pk
    return _serve_lead_attachment(obj, request=request)


@login_required
def standalone_admin_ref_links(request: HttpRequest) -> HttpResponse:
    """Управление реферальными ссылками самостоятельного админа."""
    if not _require_standalone_admin(request):
        return HttpResponseForbidden("Недостаточно прав.")
    from .models import ReferralLink
    import secrets

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "create":
            note = (request.POST.get("note") or "").strip()[:100]
            code = secrets.token_urlsafe(8)[:12]
            ReferralLink.objects.create(standalone_admin=request.user, code=code, note=note)
            messages.success(request, f"Ссылка создана: /ref/{code}/")
            return redirect("standalone_admin_ref_links")
        elif action == "deactivate":
            link_id = request.POST.get("link_id")
            if link_id:
                ReferralLink.objects.filter(pk=link_id, standalone_admin=request.user).update(is_active=False)
                messages.success(request, "Ссылка деактивирована.")
            return redirect("standalone_admin_ref_links")
        elif action == "activate":
            link_id = request.POST.get("link_id")
            if link_id:
                ReferralLink.objects.filter(pk=link_id, standalone_admin=request.user).update(is_active=True)
                messages.success(request, "Ссылка активирована.")
            return redirect("standalone_admin_ref_links")

    links = ReferralLink.objects.filter(standalone_admin=request.user).order_by("-created_at")
    return render(request, "core/standalone_admin_ref_links.html", {"links": links})


@login_required
def standalone_admin_workers(request: HttpRequest) -> HttpResponse:
    """Список исполнителей самостоятельного админа со статистикой."""
    if not _require_standalone_admin(request):
        return HttpResponseForbidden("Недостаточно прав.")
    from .models import WorkerReport
    workers = (
        User.objects.filter(standalone_admin_owner=request.user, role="worker")
        .annotate(
            reports_total=Count("worker_reports", distinct=True),
            reports_approved=Count("worker_reports", filter=Q(worker_reports__status="approved"), distinct=True),
            reports_pending=Count("worker_reports", filter=Q(worker_reports__status="pending"), distinct=True),
            self_leads_total=Count("self_leads", distinct=True),
            self_leads_approved=Count("self_leads", filter=Q(self_leads__status="approved"), distinct=True),
        )
        .order_by("-date_joined")
    )
    return render(request, "core/standalone_admin_workers.html", {"workers": workers})


@login_required
def standalone_admin_assign_lead(request: HttpRequest, lead_id: int) -> HttpResponse:
    """Назначить лид исполнителю. POST: worker_id, task_description."""
    if not _require_standalone_admin(request):
        return HttpResponseForbidden("Недостаточно прав.")
    from .models import LeadAssignment
    lead = get_object_or_404(Lead, pk=lead_id, needs_team_contact=True, status=Lead.Status.APPROVED)

    if request.method == "POST":
        worker_id = request.POST.get("worker_id")
        task_description = (request.POST.get("task_description") or "").strip()
        if not worker_id:
            messages.error(request, "Выберите исполнителя.")
        else:
            worker = get_object_or_404(User, pk=worker_id, standalone_admin_owner=request.user, role="worker")
            existing = LeadAssignment.objects.filter(lead=lead).exclude(worker=worker).select_related("worker").first()
            if existing:
                messages.warning(request, f"Лид #{lead_id} уже назначен @{existing.worker.username}. Один лид — один исполнитель.")
            else:
                assignment, created = LeadAssignment.objects.get_or_create(
                    lead=lead,
                    worker=worker,
                    defaults={"assigned_by": request.user, "task_description": task_description},
                )
                if created:
                    if lead.ss_admin_status is None:
                        lead.ss_admin_status = "in_progress"
                        lead.save(update_fields=["ss_admin_status", "updated_at"])
                    messages.success(request, f"Лид #{lead_id} назначен @{worker.username}.")
                else:
                    assignment.task_description = task_description
                    assignment.save(update_fields=["task_description", "updated_at"])
                    messages.info(request, f"Назначение обновлено для @{worker.username}.")
        next_url = request.POST.get("next") or request.GET.get("next")
        if next_url:
            from django.utils.http import url_has_allowed_host_and_scheme
            if url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
                return redirect(next_url)
        return redirect("standalone_admin_ss_leads")

    # GET: show assignment form
    workers = User.objects.filter(standalone_admin_owner=request.user, role="worker").order_by("username")
    from .models import WorkerReport
    existing_assignments = list(
        lead.assignments.select_related("worker").prefetch_related("report").all()
    )
    # Annotate each assignment with safe report status (None if no report yet)
    for a in existing_assignments:
        try:
            a.report_status = a.report.status
        except Exception:
            a.report_status = None
    next_url = request.GET.get("next", "")
    return render(request, "core/standalone_admin_assign_lead.html", {
        "lead": lead,
        "workers": workers,
        "existing_assignments": existing_assignments,
        "next_url": next_url,
    })


@login_required
def standalone_admin_worker_reports(request: HttpRequest) -> HttpResponse:
    """Список отчётов исполнителей для самостоятельного админа (вкладки по статусам)."""
    if not _require_standalone_admin(request):
        return HttpResponseForbidden("Недостаточно прав.")
    from .models import WorkerReport
    tab = request.GET.get("tab", "pending")
    tab_map = {
        "pending": WorkerReport.Status.PENDING,
        "approved": WorkerReport.Status.APPROVED,
        "rejected": WorkerReport.Status.REJECTED,
        "rework": WorkerReport.Status.REWORK,
    }
    status = tab_map.get(tab, WorkerReport.Status.PENDING)
    search_q = (request.GET.get("q") or "").strip().lstrip("@")[:100]
    reports_qs = (
        WorkerReport.objects.filter(standalone_admin=request.user, status=status)
        .select_related("worker", "assignment__lead", "assignment__lead__lead_type")
        .order_by("-created_at")
    )
    if search_q:
        words = [w.strip() for w in search_q.split() if w.strip()]
        for word in words:
            reports_qs = reports_qs.filter(
                Q(worker__username__icontains=word)
                | Q(raw_contact__icontains=word)
                | Q(comment__icontains=word)
            )
    paginator = Paginator(reports_qs, 30)
    try:
        page_number = int(request.GET.get("page", 1))
    except (TypeError, ValueError):
        page_number = 1
    page_obj = paginator.get_page(page_number)
    counts = {
        t: WorkerReport.objects.filter(standalone_admin=request.user, status=s).count()
        for t, s in tab_map.items()
    }
    return render(request, "core/standalone_admin_worker_reports.html", {
        "page_obj": page_obj,
        "tab": tab,
        "counts": counts,
        "search_q": search_q,
    })


@login_required
def standalone_admin_report_approve(request: HttpRequest, report_id: int) -> HttpResponse:
    """Одобрить отчёт воркера: +reward к балансу воркера."""
    if not _require_standalone_admin(request):
        return HttpResponseForbidden("Недостаточно прав.")
    from .models import WorkerReport
    if request.method != "POST":
        return redirect("standalone_admin_worker_reports")
    report = get_object_or_404(WorkerReport, pk=report_id, standalone_admin=request.user)
    if report.status not in (WorkerReport.Status.PENDING, WorkerReport.Status.REWORK):
        messages.warning(request, "Отчёт уже обработан.")
        return redirect("standalone_admin_worker_reports")
    with transaction.atomic():
        report_refresh = WorkerReport.objects.select_for_update().get(pk=report_id, standalone_admin=request.user)
        report_refresh.status = WorkerReport.Status.APPROVED
        report_refresh.reviewed_at = timezone.now()
        report_refresh.reviewed_by = request.user
        report_refresh.rework_comment = ""
        report_refresh.rejection_reason = ""
        report_refresh.save(update_fields=["status", "reviewed_at", "reviewed_by", "rework_comment", "rejection_reason", "updated_at"])
        reward = report_refresh.reward or 40
        worker = User.objects.select_for_update().get(pk=report_refresh.worker_id)
        worker.balance = (worker.balance or 0) + reward
        worker.save(update_fields=["balance"])
    messages.success(request, f"Отчёт одобрен. @{report_refresh.worker.username} +{reward} руб.")
    return redirect("standalone_admin_worker_reports")


@login_required
def standalone_admin_report_reject(request: HttpRequest, report_id: int) -> HttpResponse:
    """Отклонить отчёт воркера — форма с причиной."""
    if not _require_standalone_admin(request):
        return HttpResponseForbidden("Недостаточно прав.")
    from .models import WorkerReport
    from .forms import LeadRejectForm
    report = get_object_or_404(WorkerReport, pk=report_id, standalone_admin=request.user)
    if request.method == "POST":
        form = LeadRejectForm(request.POST)
        if form.is_valid():
            report.status = WorkerReport.Status.REJECTED
            report.rejection_reason = form.cleaned_data["rejection_reason"].strip()
            report.rework_comment = ""
            report.reviewed_at = timezone.now()
            report.reviewed_by = request.user
            report.save(update_fields=["status", "rejection_reason", "rework_comment", "reviewed_at", "reviewed_by", "updated_at"])
            messages.success(request, f"Отчёт #{report_id} отклонён.")
            return redirect("standalone_admin_worker_reports")
    else:
        form = LeadRejectForm()
    return render(request, "core/standalone_admin_report_reject.html", {"report": report, "form": form})


@login_required
def standalone_admin_report_rework(request: HttpRequest, report_id: int) -> HttpResponse:
    """Отправить отчёт воркера на доработку — форма с комментарием."""
    if not _require_standalone_admin(request):
        return HttpResponseForbidden("Недостаточно прав.")
    from .models import WorkerReport
    from .forms import LeadReworkForm
    report = get_object_or_404(WorkerReport, pk=report_id, standalone_admin=request.user)
    if request.method == "POST":
        form = LeadReworkForm(request.POST)
        if form.is_valid():
            report.status = WorkerReport.Status.REWORK
            report.rework_comment = form.cleaned_data["rework_comment"].strip()
            report.rejection_reason = ""
            report.reviewed_at = timezone.now()
            report.reviewed_by = request.user
            report.save(update_fields=["status", "rework_comment", "rejection_reason", "reviewed_at", "reviewed_by", "updated_at"])
            messages.success(request, f"Отчёт #{report_id} отправлен на доработку.")
            return redirect("standalone_admin_worker_reports")
    else:
        form = LeadReworkForm()
    return render(request, "core/standalone_admin_report_rework.html", {"report": report, "form": form})


@login_required
def standalone_admin_worker_report_attachment(request: HttpRequest, report_id: int) -> HttpResponse:
    """Отдаёт вложение к отчёту воркера (только для владельца — самостоятельного админа)."""
    if not _require_standalone_admin(request):
        return HttpResponseForbidden("Недостаточно прав.")
    from .models import WorkerReport
    report = get_object_or_404(WorkerReport, pk=report_id, standalone_admin=request.user)
    if not report.attachment:
        return HttpResponseForbidden("Вложение отсутствует.")
    return _serve_worker_report_attachment(report, request=request)


@login_required
def standalone_admin_worker_withdrawal_requests(request: HttpRequest) -> HttpResponse:
    """Список заявок воркеров на вывод. Одобрить — подтвердить. Отклонить — вернуть баланс."""
    if not _require_standalone_admin(request):
        return HttpResponseForbidden("Недостаточно прав.")
    from .models import WorkerWithdrawalRequest

    if request.method == "POST":
        req_id = request.POST.get("request_id")
        action = request.POST.get("action")
        if req_id and action in ("approve", "reject"):
            with transaction.atomic():
                wreq = (
                    WorkerWithdrawalRequest.objects
                    .select_for_update(of=("self",))
                    .select_related("worker")
                    .filter(pk=req_id, standalone_admin=request.user, status="pending")
                    .first()
                )
                if not wreq:
                    messages.warning(request, "Заявка уже обработана или не найдена.")
                else:
                    now = timezone.now()
                    if action == "approve":
                        wreq.status = "approved"
                        wreq.processed_at = now
                        wreq.processed_by = request.user
                        wreq.save()
                        messages.success(request, f"Вывод @{wreq.worker.username} на {wreq.amount} руб. одобрен.")
                    else:
                        wreq.worker.balance = (wreq.worker.balance or 0) + wreq.amount
                        wreq.worker.save(update_fields=["balance"])
                        wreq.status = "rejected"
                        wreq.processed_at = now
                        wreq.processed_by = request.user
                        wreq.save()
                        messages.info(request, f"Заявка @{wreq.worker.username} отклонена. Баланс восстановлен.")
        return redirect("standalone_admin_worker_withdrawal_requests")

    pending = (
        WorkerWithdrawalRequest.objects
        .filter(standalone_admin=request.user, status="pending")
        .select_related("worker")
        .order_by("created_at")
    )
    history = (
        WorkerWithdrawalRequest.objects
        .filter(standalone_admin=request.user)
        .exclude(status="pending")
        .select_related("worker", "processed_by")
        .order_by("-created_at")[:100]
    )
    total_worker_approved = (
        WorkerWithdrawalRequest.objects
        .filter(standalone_admin=request.user, status="approved")
        .aggregate(s=Sum("amount"))["s"] or 0
    )
    total_user_approved = (
        WithdrawalRequest.objects
        .filter(status="approved")
        .aggregate(s=Sum("amount"))["s"] or 0
    )
    return render(request, "core/standalone_admin_worker_withdrawal_requests.html", {
        "pending_requests": pending,
        "history_requests": history,
        "total_worker_approved": total_worker_approved,
        "total_user_approved": total_user_approved,
        "total_approved_all": total_worker_approved + total_user_approved,
    })


@login_required
def standalone_admin_worker_withdrawal_debug(request: HttpRequest) -> HttpResponse:
    """Диагностика: показывает статус заявок на вывод."""
    import traceback as tb_mod
    try:
        if not _require_standalone_admin(request):
            return HttpResponse("NOT standalone_admin", content_type="text/plain")

        from .models import WorkerWithdrawalRequest
        pending_count = WorkerWithdrawalRequest.objects.filter(
            standalone_admin=request.user, status="pending"
        ).count()
        history_count = WorkerWithdrawalRequest.objects.filter(
            standalone_admin=request.user
        ).exclude(status="pending").count()
        total_w = (
            WorkerWithdrawalRequest.objects
            .filter(standalone_admin=request.user, status="approved")
            .aggregate(s=Sum("amount"))["s"] or 0
        )
        total_u = WithdrawalRequest.objects.filter(status="approved").aggregate(s=Sum("amount"))["s"] or 0
        lines = [
            f"OK — view работает v2",
            f"pending: {pending_count}",
            f"history: {history_count}",
            f"total_worker_approved: {total_w}",
            f"total_user_approved: {total_u}",
        ]
        return HttpResponse("\n".join(lines), content_type="text/plain; charset=utf-8")
    except Exception:
        return HttpResponse(
            f"CRASH:\n\n{tb_mod.format_exc()}",
            content_type="text/plain; charset=utf-8",
        )


# ──────────────────────────────────────────────────────────────
# SS Admin: Worker Self-Leads
# ──────────────────────────────────────────────────────────────

@login_required
def standalone_admin_worker_self_leads(request: HttpRequest) -> HttpResponse:
    """Список самостоятельных лидов от исполнителей для проверки СС-админом."""
    if not _require_standalone_admin(request):
        return HttpResponseForbidden("Недостаточно прав.")
    tab = request.GET.get("tab", "pending")
    valid_tabs = ("pending", "approved", "rejected", "rework")
    if tab not in valid_tabs:
        tab = "pending"

    base_qs = WorkerSelfLead.objects.filter(standalone_admin=request.user).select_related("worker")
    search_q = (request.GET.get("q") or "").strip().lstrip("@")[:100]
    leads = base_qs.filter(status=tab).order_by("-created_at")
    if search_q:
        words = [w.strip() for w in search_q.split() if w.strip()]
        for word in words:
            leads = leads.filter(
                Q(worker__username__icontains=word)
                | Q(raw_contact__icontains=word)
                | Q(comment__icontains=word)
            )

    from django.core.paginator import Paginator
    paginator = Paginator(leads, 30)
    page_obj = paginator.get_page(request.GET.get("page"))

    counts = {t: base_qs.filter(status=t).count() for t in valid_tabs}

    return render(request, "core/standalone_admin_worker_self_leads.html", {
        "tab": tab,
        "page_obj": page_obj,
        "counts": counts,
        "search_q": search_q,
    })


@login_required
def standalone_admin_worker_self_lead_approve(request: HttpRequest, self_lead_id: int) -> HttpResponse:
    """Одобрить самостоятельный лид исполнителя и начислить вознаграждение."""
    if not _require_standalone_admin(request):
        return HttpResponseForbidden("Недостаточно прав.")
    if request.method != "POST":
        return HttpResponseForbidden("Только POST.")
    self_lead = get_object_or_404(WorkerSelfLead, pk=self_lead_id, standalone_admin=request.user)
    if self_lead.status not in (WorkerSelfLead.Status.PENDING, WorkerSelfLead.Status.REWORK, WorkerSelfLead.Status.REJECTED):
        from django.contrib import messages
        messages.warning(request, "Лид уже обработан.")
        return redirect("standalone_admin_worker_self_leads")

    from django.db import transaction
    from django.utils import timezone
    with transaction.atomic():
        self_lead.status = WorkerSelfLead.Status.APPROVED
        self_lead.rejection_reason = ""
        self_lead.rework_comment = ""
        self_lead.reviewed_at = timezone.now()
        self_lead.reviewed_by = request.user
        self_lead.save(update_fields=["status", "rejection_reason", "rework_comment", "reviewed_at", "reviewed_by_id", "updated_at"])

        worker = User.objects.select_for_update().get(pk=self_lead.worker_id)
        worker.balance = (worker.balance or 0) + self_lead.reward
        worker.save(update_fields=["balance"])

    from django.contrib import messages
    messages.success(request, f"Лид одобрен. Исполнителю @{self_lead.worker.username} начислено {self_lead.reward} руб.")
    return redirect("standalone_admin_worker_self_leads")


@login_required
def standalone_admin_worker_self_lead_reject(request: HttpRequest, self_lead_id: int) -> HttpResponse:
    """Отклонить самостоятельный лид исполнителя (включая уже одобренные — с возвратом вознаграждения)."""
    if not _require_standalone_admin(request):
        return HttpResponseForbidden("Недостаточно прав.")
    if request.method != "POST":
        return HttpResponseForbidden("Только POST.")
    self_lead = get_object_or_404(WorkerSelfLead, pk=self_lead_id, standalone_admin=request.user)
    if self_lead.status == WorkerSelfLead.Status.REJECTED:
        from django.contrib import messages
        messages.warning(request, "Лид уже отклонён.")
        return redirect("standalone_admin_worker_self_leads")

    rejection_reason = (request.POST.get("rejection_reason") or "").strip()
    was_approved = self_lead.status == WorkerSelfLead.Status.APPROVED

    from django.db import transaction
    from django.utils import timezone
    with transaction.atomic():
        self_lead.status = WorkerSelfLead.Status.REJECTED
        self_lead.rejection_reason = rejection_reason
        self_lead.reviewed_at = timezone.now()
        self_lead.reviewed_by = request.user
        self_lead.save(update_fields=["status", "rejection_reason", "reviewed_at", "reviewed_by_id", "updated_at"])

        if was_approved:
            worker = User.objects.select_for_update().get(pk=self_lead.worker_id)
            worker.balance = (worker.balance or 0) - self_lead.reward
            worker.save(update_fields=["balance"])

    from django.contrib import messages
    from django.urls import reverse
    if was_approved:
        messages.success(request, f"Лид отклонён. С баланса @{self_lead.worker.username} списано {self_lead.reward} руб.")
        return redirect(reverse("standalone_admin_worker_self_leads") + "?tab=approved")
    messages.success(request, "Лид отклонён.")
    return redirect("standalone_admin_worker_self_leads")


@login_required
def standalone_admin_worker_self_lead_rework(request: HttpRequest, self_lead_id: int) -> HttpResponse:
    """Отправить самостоятельный лид исполнителя на доработку (включая уже одобренные — с возвратом вознаграждения)."""
    if not _require_standalone_admin(request):
        return HttpResponseForbidden("Недостаточно прав.")
    if request.method != "POST":
        return HttpResponseForbidden("Только POST.")
    self_lead = get_object_or_404(WorkerSelfLead, pk=self_lead_id, standalone_admin=request.user)
    if self_lead.status == WorkerSelfLead.Status.REWORK:
        from django.contrib import messages
        from django.urls import reverse
        messages.warning(request, "Лид уже на доработке.")
        return redirect(reverse("standalone_admin_worker_self_leads") + "?tab=rework")

    rework_comment = (request.POST.get("rework_comment") or "").strip()
    was_approved = self_lead.status == WorkerSelfLead.Status.APPROVED

    from django.db import transaction
    from django.utils import timezone
    with transaction.atomic():
        self_lead.status = WorkerSelfLead.Status.REWORK
        self_lead.rework_comment = rework_comment
        self_lead.reviewed_at = timezone.now()
        self_lead.reviewed_by = request.user
        self_lead.save(update_fields=["status", "rework_comment", "reviewed_at", "reviewed_by_id", "updated_at"])

        if was_approved:
            worker = User.objects.select_for_update().get(pk=self_lead.worker_id)
            worker.balance = (worker.balance or 0) - self_lead.reward
            worker.save(update_fields=["balance"])

    from django.contrib import messages
    if was_approved:
        messages.success(request, f"Лид отправлен на доработку. С баланса @{self_lead.worker.username} списано {self_lead.reward} руб.")
    else:
        messages.success(request, "Лид отправлен на доработку.")
    return redirect("standalone_admin_worker_self_leads")


@login_required
def standalone_admin_worker_self_lead_attachment(request: HttpRequest, self_lead_id: int) -> HttpResponse:
    """Отдаёт вложение самостоятельного лида для СС-админа."""
    if not _require_standalone_admin(request):
        return HttpResponseForbidden("Недостаточно прав.")
    self_lead = get_object_or_404(WorkerSelfLead, pk=self_lead_id, standalone_admin=request.user)
    if not self_lead.attachment:
        return HttpResponseForbidden("Вложение отсутствует.")
    return _serve_lead_attachment(self_lead, request=request)


# ──────────────────────────────────────────────────────────
#  Сброс пароля (админ + СС-админ)
# ──────────────────────────────────────────────────────────

def _generate_password(length: int = 10) -> str:
    import secrets
    import string
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))


@login_required
def admin_reset_password(request: HttpRequest) -> HttpResponse:
    """Сброс пароля пользователя (для обычного админа/саппорта)."""
    if not _require_support(request):
        return HttpResponseForbidden("Недостаточно прав.")

    found_user = None
    new_password = None
    not_found = False
    query = ""

    if request.method == "POST":
        action = request.POST.get("action", "")
        query = request.POST.get("username", "").strip()

        _admin_roles = ("admin", "main_admin", "support", "balance_admin", "standalone_admin")
        _is_main = getattr(request.user, "role", None) == "main_admin"

        if action == "search" and query:
            try:
                found_user = User.objects.get(username=query)
                # Обычные админы не могут сбрасывать пароль другим админам
                if not _is_main and found_user.role in _admin_roles:
                    found_user = None
                    not_found = True
            except User.DoesNotExist:
                not_found = True

        elif action == "reset" and query:
            try:
                target = User.objects.get(username=query)
                if not _is_main and target.role in _admin_roles:
                    not_found = True
                else:
                    new_password = _generate_password()
                    target.set_password(new_password)
                    target.save(update_fields=["password"])
                    found_user = target
            except User.DoesNotExist:
                not_found = True

    return render(request, "core/admin_reset_password.html", {
        "found_user": found_user,
        "new_password": new_password,
        "not_found": not_found,
        "query": query,
    })


@login_required
def standalone_admin_reset_password(request: HttpRequest) -> HttpResponse:
    """Сброс пароля воркера (для СС-админа — только свои воркеры)."""
    if not _require_standalone_admin(request):
        return HttpResponseForbidden("Недостаточно прав.")

    found_user = None
    new_password = None
    not_found = False
    query = ""

    if request.method == "POST":
        action = request.POST.get("action", "")
        query = request.POST.get("username", "").strip()

        if action == "search" and query:
            try:
                found_user = User.objects.get(
                    username=query, role="worker",
                )
            except User.DoesNotExist:
                not_found = True

        elif action == "reset" and query:
            try:
                target = User.objects.get(
                    username=query, role="worker",
                )
                new_password = _generate_password()
                target.set_password(new_password)
                target.save(update_fields=["password"])
                found_user = target
            except User.DoesNotExist:
                not_found = True

    return render(request, "core/standalone_admin_reset_password.html", {
        "found_user": found_user,
        "new_password": new_password,
        "not_found": not_found,
        "query": query,
    })


# ──────────────────────────────────────────────────────────
#  Статистика начислений админов
# ──────────────────────────────────────────────────────────

@login_required
def admin_earnings_stats(request: HttpRequest) -> HttpResponse:
    """Страница статистики начислений админов (2.5р за действие). Только main_admin."""
    if getattr(request.user, "role", None) != "main_admin":
        return HttpResponseForbidden("Только для главного админа.")
    from .models import LeadReviewLog
    from decimal import Decimal

    admins = User.objects.filter(role="admin").order_by("username")
    admin_stats = []
    for admin in admins:
        actions = LeadReviewLog.objects.filter(admin=admin).count()
        earned = int(actions * Decimal("2.5"))
        withdrawn = WithdrawalRequest.objects.filter(user=admin, status__in=("pending", "approved")).aggregate(s=Sum("amount")).get("s") or 0
        admin_stats.append({
            "user": admin,
            "actions": actions,
            "earned": earned,
            "withdrawn": withdrawn,
            "available": max(0, earned - withdrawn),
        })

    selected_admin = None
    logs = []
    admin_id = request.GET.get("admin_id")
    if admin_id:
        selected_admin = User.objects.filter(pk=admin_id, role="admin").first()
        if selected_admin:
            logs = (
                LeadReviewLog.objects.filter(admin=selected_admin)
                .select_related("lead", "lead__user", "lead__lead_type")
                .order_by("-created_at")[:200]
            )

    return render(request, "core/admin_earnings_stats.html", {
        "admin_stats": admin_stats,
        "selected_admin": selected_admin,
        "logs": logs,
    })




# ─── Аккредитация пользователя ─────────────────────────────────────────────

@login_required
@require_http_methods(["POST"])
def admin_toggle_accredited(request: HttpRequest, user_id: int) -> HttpResponse:
    """Переключить галочку аккредитации пользователя. Только main_admin."""
    if getattr(request.user, "role", None) != User.Role.MAIN_ADMIN:
        return HttpResponseForbidden("Только для главного админа.")
    target = get_object_or_404(User, pk=user_id)
    target.is_accredited = not target.is_accredited
    target.save(update_fields=["is_accredited"])
    status = "аккредитирован" if target.is_accredited else "не аккредитирован"
    messages.success(request, f"@{target.username} — {status}.")
    return redirect("admin_all_users")


# ─── Отказы исполнителей (СС-админ) ──────────────────────────────────────────

@login_required
def standalone_admin_refused(request: HttpRequest) -> HttpResponse:
    """Список заданий, где исполнитель отметил отказ лида."""
    if not _require_standalone_admin(request):
        return HttpResponseForbidden("Только для самостоятельного админа.")
    from .models import LeadAssignment
    refused_qs = (
        LeadAssignment.objects.filter(
            worker__standalone_admin_owner=request.user,
            refused=True,
        )
        .select_related("lead", "lead__lead_type", "worker")
        .order_by("-refused_at")
    )
    return render(request, "core/standalone_admin_refused.html", {
        "refused_list": refused_qs,
    })


# ─── Бан/разбан пользователя ─────────────────────────────────────────────────

@login_required
@require_http_methods(["POST"])
def admin_toggle_ban(request: HttpRequest, user_id: int) -> HttpResponse:
    """Забанить или разбанить пользователя."""
    if not _require_support(request) and getattr(request.user, "role", None) != User.Role.BALANCE_ADMIN:
        return HttpResponseForbidden("Недостаточно прав.")
    target = get_object_or_404(User, pk=user_id)
    if target.role in ("main_admin", "admin", "support"):
        messages.error(request, "Нельзя банить админов и саппортов.")
        return redirect("admin_all_users")
    if target.status == User.Status.BANNED:
        target.status = User.Status.APPROVED
        target.is_active = True
        target.save(update_fields=["status", "is_active"])
        messages.success(request, f"@{target.username} разбанен.")
    else:
        target.status = User.Status.BANNED
        target.is_active = False  # Django auto-logout + запрет повторного логина
        target.save(update_fields=["status", "is_active"])
        messages.success(request, f"@{target.username} забанен.")
    return redirect("admin_all_users")


# ─── Оплата (баланс-админ) ───────────────────────────────────────────────────

def _is_balance_admin_check(request: HttpRequest) -> bool:
    return getattr(request.user, "role", None) == User.Role.BALANCE_ADMIN


@login_required
def balance_admin_payment_list(request: HttpRequest) -> HttpResponse:
    """Список пользователей для оплаты (баланс-админ)."""
    if not _is_balance_admin_check(request):
        return HttpResponseForbidden("Только для баланс-админа.")
    q = (request.GET.get("q") or "").strip().lstrip("@")[:50]
    users_qs = User.objects.filter(role="user").order_by("-date_joined")
    if q:
        users_qs = users_qs.filter(username__icontains=q)
    from django.core.paginator import Paginator
    paginator = Paginator(users_qs, 30)
    page_obj = paginator.get_page(request.GET.get("page", 1))
    return render(request, "core/balance_admin_payment_list.html", {
        "page_obj": page_obj,
        "q": q,
    })


@login_required
def balance_admin_payment_detail(request: HttpRequest, user_id: int) -> HttpResponse:
    """Детали пользователя для оплаты: баланс, выводы, кнопки."""
    if not _is_balance_admin_check(request):
        return HttpResponseForbidden("Только для баланс-админа.")
    target = get_object_or_404(User, pk=user_id, role="user")
    pending_withdrawals = WithdrawalRequest.objects.filter(user=target, status="pending").order_by("-created_at")
    prev_balance = request.session.get(f"prev_balance_{user_id}")
    return render(request, "core/balance_admin_payment_detail.html", {
        "target": target,
        "pending_withdrawals": pending_withdrawals,
        "prev_balance": prev_balance,
    })


@login_required
@require_http_methods(["POST"])
def balance_admin_payment_multiply(request: HttpRequest, user_id: int) -> HttpResponse:
    """Умножить баланс пользователя на 1.5."""
    if not _is_balance_admin_check(request):
        return HttpResponseForbidden()
    with transaction.atomic():
        target = User.objects.select_for_update().get(pk=user_id, role="user")
        _old = target.balance
        request.session[f"prev_balance_{user_id}"] = _old
        target.balance = round(_old * 1.5)
        target.save(update_fields=["balance"])
        log_balance_change(target, "balance", _old, target.balance, "payment_multiply x1.5", request.user)
    messages.success(request, f"@{target.username}: баланс ×1.5 → {target.balance} руб.")
    return redirect("balance_admin_payment_detail", user_id=user_id)


@login_required
@require_http_methods(["POST"])
def balance_admin_payment_subtract(request: HttpRequest, user_id: int) -> HttpResponse:
    """Вычесть 25000 из баланса пользователя."""
    if not _is_balance_admin_check(request):
        return HttpResponseForbidden()
    with transaction.atomic():
        target = User.objects.select_for_update().get(pk=user_id, role="user")
        _old = target.balance
        request.session[f"prev_balance_{user_id}"] = _old
        target.balance = _old - 25000
        target.save(update_fields=["balance"])
        log_balance_change(target, "balance", _old, target.balance, "payment_subtract -25000", request.user)
    messages.success(request, f"@{target.username}: баланс −25000 → {target.balance} руб.")
    return redirect("balance_admin_payment_detail", user_id=user_id)


@login_required
@require_http_methods(["POST"])
def balance_admin_payment_revert(request: HttpRequest, user_id: int) -> HttpResponse:
    """Вернуть баланс к предыдущему значению."""
    if not _is_balance_admin_check(request):
        return HttpResponseForbidden()
    prev = request.session.get(f"prev_balance_{user_id}")
    if prev is None:
        messages.error(request, "Нет сохранённого предыдущего баланса.")
        return redirect("balance_admin_payment_detail", user_id=user_id)
    with transaction.atomic():
        target = User.objects.select_for_update().get(pk=user_id, role="user")
        _old = target.balance
        target.balance = prev
        target.save(update_fields=["balance"])
        log_balance_change(target, "balance", _old, prev, "payment_revert", request.user)
    del request.session[f"prev_balance_{user_id}"]
    messages.success(request, f"@{target.username}: баланс возвращён → {prev} руб.")
    return redirect("balance_admin_payment_detail", user_id=user_id)


# ─── СМЗ заявки ──────────────────────────────────────────────────────────────

@login_required
def admin_smz_requests(request: HttpRequest) -> HttpResponse:
    """Список заявок на СМЗ-верификацию для одобрения главным админом."""
    if getattr(request.user, "role", None) != "main_admin":
        return HttpResponseForbidden("Только для главного админа.")

    if request.method == "POST":
        user_id = request.POST.get("user_id")
        action = request.POST.get("action")
        target = User.objects.filter(pk=user_id).first()
        if target:
            if action == "approve":
                target.smz_status = "approved"
                target.smz_reject_reason = ""
                target.save(update_fields=["smz_status", "smz_reject_reason"])
                messages.success(request, f"СМЗ @{target.username} одобрена.")
            elif action == "reject":
                reason = (request.POST.get("reason") or "").strip()
                target.smz_status = "rejected"
                target.smz_reject_reason = reason
                target.save(update_fields=["smz_status", "smz_reject_reason"])
                messages.success(request, f"СМЗ @{target.username} отклонена.")
        return redirect("admin_smz_requests")

    tab = request.GET.get("tab", "pending")
    if tab == "pending":
        qs = User.objects.filter(smz_status="pending").order_by("-smz_submitted_at")
    elif tab == "approved":
        qs = User.objects.filter(smz_status="approved").exclude(smz_fio="").order_by("-smz_submitted_at")
    elif tab == "rejected":
        qs = User.objects.filter(smz_status="rejected").order_by("-smz_submitted_at")
    else:
        qs = User.objects.exclude(smz_status="none").order_by("-smz_submitted_at")

    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get("page", 1))
    pending_count = User.objects.filter(smz_status="pending").count()

    return render(request, "core/admin_smz_requests.html", {
        "page_obj": page_obj,
        "tab": tab,
        "pending_count": pending_count,
    })


# ─── Чеки ────────────────────────────────────────────────────────────────────

@login_required
def admin_receipts(request: HttpRequest) -> HttpResponse:
    """Список загруженных чеков для проверки."""
    if getattr(request.user, "role", None) != "main_admin":
        return HttpResponseForbidden("Только для главного админа.")

    if request.method == "POST":
        wr_id = request.POST.get("wr_id")
        action = request.POST.get("action", "approve")
        wr = WithdrawalRequest.objects.filter(pk=wr_id).first()
        if wr:
            if action == "approve":
                wr.receipt_status = "approved"
                wr.receipt_checked = True
                wr.receipt_reject_reason = ""
                wr.save(update_fields=["receipt_status", "receipt_checked", "receipt_reject_reason", "updated_at"])
                messages.success(request, f"Чек #{wr.pk} (@{wr.user.username}) одобрен.")
            elif action == "reject":
                reason = (request.POST.get("reason") or "").strip()
                wr.receipt_status = "rejected"
                wr.receipt_reject_reason = reason
                wr.receipt_checked = False
                wr.save(update_fields=["receipt_status", "receipt_reject_reason", "receipt_checked", "updated_at"])
                messages.success(request, f"Чек #{wr.pk} (@{wr.user.username}) отклонён.")
        return redirect("admin_receipts")

    tab = request.GET.get("tab", "unchecked")
    qs = WithdrawalRequest.objects.filter(receipt_status__in=["pending", "approved", "rejected"]).select_related("user")
    if tab == "unchecked":
        qs = qs.filter(receipt_status="pending")
    qs = qs.order_by("-receipt_uploaded_at")

    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get("page", 1))
    unchecked_count = WithdrawalRequest.objects.filter(receipt_status="pending").count()

    return render(request, "core/admin_receipts.html", {
        "page_obj": page_obj,
        "tab": tab,
        "unchecked_count": unchecked_count,
    })


# ─── Модерация по админам (кто что проверял) ─────────────────────────────────

@login_required
def admin_moderation_by_admin_list(request: HttpRequest) -> HttpResponse:
    """Сводная таблица: сколько отчётов каждый админ одобрил / отклонил / отправил на доработку."""
    if not _require_support(request):
        return HttpResponseForbidden("Недостаточно прав.")

    from .models import SearchReport

    # Лиды (основная биржа) — агрегируем по reviewed_by
    lead_stats = {}
    for row in (
        Lead.objects.filter(reviewed_by__isnull=False)
        .values("reviewed_by_id", "status")
        .annotate(c=Count("id"))
    ):
        d = lead_stats.setdefault(row["reviewed_by_id"], {"approved": 0, "rejected": 0, "rework": 0})
        if row["status"] in d:
            d[row["status"]] = row["c"]

    # Отчёты SearchLink — агрегируем по reviewed_by
    sr_stats = {}
    for row in (
        SearchReport.objects.filter(reviewed_by__isnull=False)
        .values("reviewed_by_id", "status")
        .annotate(c=Count("id"))
    ):
        d = sr_stats.setdefault(row["reviewed_by_id"], {"approved": 0, "rejected": 0, "rework": 0})
        if row["status"] in d:
            d[row["status"]] = row["c"]

    admin_ids = set(lead_stats.keys()) | set(sr_stats.keys())
    admins = {u.id: u for u in User.objects.filter(id__in=admin_ids)}

    rows = []
    for aid in admin_ids:
        admin = admins.get(aid)
        if not admin:
            continue
        ls = lead_stats.get(aid, {"approved": 0, "rejected": 0, "rework": 0})
        ss = sr_stats.get(aid, {"approved": 0, "rejected": 0, "rework": 0})
        total = sum(ls.values()) + sum(ss.values())
        rows.append({
            "admin": admin,
            "lead_approved": ls["approved"],
            "lead_rejected": ls["rejected"],
            "lead_rework": ls["rework"],
            "sr_approved": ss["approved"],
            "sr_rejected": ss["rejected"],
            "sr_rework": ss["rework"],
            "total": total,
        })
    rows.sort(key=lambda r: r["total"], reverse=True)

    return render(request, "core/admin_moderation_by_admin_list.html", {
        "rows": rows,
    })


@login_required
def admin_moderation_by_admin_detail(request: HttpRequest, admin_id: int) -> HttpResponse:
    """Детализация: что именно проверял конкретный админ — лиды + SearchReport."""
    if not _require_support(request):
        return HttpResponseForbidden("Недостаточно прав.")

    from .models import SearchReport

    target_admin = get_object_or_404(User, pk=admin_id)
    kind = request.GET.get("kind", "lead")  # lead | searchreport
    tab = request.GET.get("tab", "rework")  # approved | rejected | rework | all
    if kind not in ("lead", "searchreport"):
        kind = "lead"
    if tab not in ("approved", "rejected", "rework", "all"):
        tab = "rework"

    if kind == "lead":
        qs = Lead.objects.filter(reviewed_by=target_admin).select_related("user", "lead_type", "base_type")
        if tab == "approved":
            qs = qs.filter(status=Lead.Status.APPROVED)
        elif tab == "rejected":
            qs = qs.filter(status=Lead.Status.REJECTED)
        elif tab == "rework":
            qs = qs.filter(status=Lead.Status.REWORK)
        qs = qs.order_by("-reviewed_at")
    else:
        qs = SearchReport.objects.filter(reviewed_by=target_admin).select_related("user", "search_link")
        if tab == "approved":
            qs = qs.filter(status=SearchReport.Status.APPROVED)
        elif tab == "rejected":
            qs = qs.filter(status=SearchReport.Status.REJECTED)
        elif tab == "rework":
            qs = qs.filter(status=SearchReport.Status.REWORK)
        qs = qs.order_by("-reviewed_at")

    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    # Счётчики для бейджей вкладок
    if kind == "lead":
        base = Lead.objects.filter(reviewed_by=target_admin)
        counts = {
            "approved": base.filter(status=Lead.Status.APPROVED).count(),
            "rejected": base.filter(status=Lead.Status.REJECTED).count(),
            "rework": base.filter(status=Lead.Status.REWORK).count(),
        }
    else:
        base = SearchReport.objects.filter(reviewed_by=target_admin)
        counts = {
            "approved": base.filter(status=SearchReport.Status.APPROVED).count(),
            "rejected": base.filter(status=SearchReport.Status.REJECTED).count(),
            "rework": base.filter(status=SearchReport.Status.REWORK).count(),
        }
    counts["all"] = counts["approved"] + counts["rejected"] + counts["rework"]

    return render(request, "core/admin_moderation_by_admin_detail.html", {
        "target_admin": target_admin,
        "page_obj": page_obj,
        "kind": kind,
        "tab": tab,
        "counts": counts,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# Zvonok.com — настройки + cron поллинга входящих звонков
# ═══════════════════════════════════════════════════════════════════════════════

from django.views.decorators.csrf import csrf_exempt as _csrf_exempt_zv
from .robocall import (
    get_or_create_webhook_secret as _zv_get_secret,
    poll_incoming_calls as _zv_poll_incoming,
)


@login_required
def admin_robocall_test(request: HttpRequest) -> HttpResponse:
    """Страница главного админа: настройки zvonok + URL cron-эндпоинта поллинга."""
    if getattr(request.user, "role", None) != "main_admin":
        return HttpResponseForbidden("Только для главного админа.")

    st = SiteSettings.get_settings()

    if request.method == "POST" and request.POST.get("action") == "save_config":
        st.zvonok_public_key = (request.POST.get("public_key") or "").strip()[:255]
        st.zvonok_incoming_campaign_id = (request.POST.get("incoming_campaign_id") or "").strip()[:64]
        st.save(update_fields=["zvonok_public_key", "zvonok_incoming_campaign_id"])
        messages.success(request, "Настройки zvonok.com сохранены.")
        return redirect("admin_robocall_test")

    secret = _zv_get_secret()
    cron_url = request.build_absolute_uri(reverse("zvonok_poll_cron") + f"?secret={secret}")

    return render(request, "core/admin_robocall_test.html", {
        "site_settings": st,
        "cron_url": cron_url,
    })


@_csrf_exempt_zv
def zvonok_poll_cron(request: HttpRequest) -> HttpResponse:
    """Cron-endpoint: бот-сервер раз в час дёргает это, чтобы поллить входящие звонки.

    URL: /api/cron/poll-incoming-calls/?secret=<secret>
    """
    st = SiteSettings.get_settings()
    expected_secret = st.zvonok_webhook_secret
    provided_secret = request.GET.get("secret", "") or request.headers.get("X-Webhook-Secret", "")
    if not expected_secret or provided_secret != expected_secret:
        return HttpResponseForbidden("Bad secret")

    return JsonResponse({"poll": _zv_poll_incoming()})


# ═══════════════════════════════════════════════════════════════════════════════
# Phone-reports dashboard — страница со списком всех телефон-отчётов
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
def admin_phone_reports(request: HttpRequest) -> HttpResponse:
    """Страница главного админа/админа: все phone_callback отчёты со статусом поллинга."""
    if not (_require_support(request) or getattr(request.user, "role", None) == "main_admin"):
        return HttpResponseForbidden("Недостаточно прав.")

    from .models import SearchReport

    reports_qs = (
        SearchReport.objects
        .filter(report_type=SearchReport.ReportType.PHONE_CALLBACK)
        .select_related("user", "search_link", "reviewed_by")
        .order_by("-created_at")
    )

    q = (request.GET.get("q") or "").strip()
    tab = request.GET.get("tab", "all")
    if q:
        reports_qs = reports_qs.filter(
            Q(client_phone__icontains=q) | Q(user__username__icontains=q) | Q(search_link__code__icontains=q)
        )
    if tab == "confirmed":
        reports_qs = reports_qs.filter(callback_confirmed_at__isnull=False)
    elif tab == "unconfirmed":
        reports_qs = reports_qs.filter(callback_confirmed_at__isnull=True)
    elif tab == "approved":
        reports_qs = reports_qs.filter(status=SearchReport.Status.APPROVED)
    elif tab == "rejected":
        reports_qs = reports_qs.filter(status=SearchReport.Status.REJECTED)

    paginator = Paginator(reports_qs, 50)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    totals = SearchReport.objects.filter(report_type=SearchReport.ReportType.PHONE_CALLBACK).aggregate(
        total=Count("id"),
        confirmed=Count("id", filter=Q(callback_confirmed_at__isnull=False)),
        approved=Count("id", filter=Q(status=SearchReport.Status.APPROVED)),
        rejected=Count("id", filter=Q(status=SearchReport.Status.REJECTED)),
        pending_callback=Count("id", filter=Q(status=SearchReport.Status.PENDING_CALLBACK)),
    )

    return render(request, "core/admin_phone_reports.html", {
        "page_obj": page_obj,
        "q": q,
        "tab": tab,
        "totals": totals,
    })

